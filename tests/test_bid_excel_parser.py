import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook

from services.excel_row_parser import classify_rows, find_header_rows
from services.excel_row_pipeline import (
    attach_category,
    build_logical_records,
    build_normalized_records,
    build_schema,
    clean_code_preserve_text,
    merge_header_rows,
)
from services.export_service import export_bid_records


TARGET_SHEET_INDEX = 1
TARGET_SHEET_NAME = "分部分项工程项目清单计价表"


def _write_mock_bid_workbook(path: Path) -> None:
    wb = Workbook()
    wb.active.title = "工程项目清单汇总表"
    ws = wb.create_sheet(TARGET_SHEET_NAME)
    ws.append(["序号", "项目编码", None, "项目名称", "项目特征描述", None, None, "计量单位", "工程量", "金额(元)", None, None])
    ws.append([None, None, None, None, None, None, None, None, None, "综合单价", None, "合价"])
    ws.append([None, None, None, "混凝土工程", None, None, None, None, None, None, None, None])
    ws.append([1, "10501001001", None, "现浇构混凝土垫层 C15", "1、混凝土强度等级:C15", None, None, "m3", 12.5, 100, None, 1250])
    ws.append([
        2,
        10506001001,
        None,
        "现浇构件钢筋",
        "1、钢筋种类、规格:Ⅲ级螺纹钢 Φ10以内\n2、综合单价不含措施筋，其工程量并入清单量计算\n3、清单工程量按图纸计算",
        None,
        None,
        "t",
        22.22,
        3936.12,
        None,
        87460.59,
    ])
    ws.append([3, "010506001001", None, "现浇构混凝土桩承台基础 C35 P8", "1、抗渗等级:P8", None, None, "m3", 5, 200, None, 1000])
    wb.save(path)


def _parse_mock_workbook(path: Path):
    excel_file = pd.ExcelFile(path)
    source_sheet_name = excel_file.sheet_names[TARGET_SHEET_INDEX]
    header_rows, skip_rows = find_header_rows(path, TARGET_SHEET_INDEX)
    schema = build_schema(merge_header_rows(header_rows))
    rows = classify_rows(path, TARGET_SHEET_INDEX, schema, skip_rows)
    logical_records = build_logical_records(attach_category(rows, schema), schema)
    return build_normalized_records(
        logical_records,
        batch_id="test-batch",
        source_file_name=path.name,
        source_sheet_index=TARGET_SHEET_INDEX,
        source_sheet_name=source_sheet_name,
    )


def test_simple_excel_without_page_info_does_not_fail(tmp_path):
    xlsx = tmp_path / "mock_bid.xlsx"
    _write_mock_bid_workbook(xlsx)

    records = _parse_mock_workbook(xlsx)

    assert len(records) == 3
    assert all(record["project_name"] == "" for record in records)


def test_item_name_preserves_internal_spaces(tmp_path):
    xlsx = tmp_path / "mock_bid.xlsx"
    _write_mock_bid_workbook(xlsx)

    names = {record["item_name"] for record in _parse_mock_workbook(xlsx)}

    assert "现浇构混凝土垫层 C15" in names
    assert "现浇构混凝土桩承台基础 C35 P8" in names
    assert "现浇构混凝土垫层C15" not in names
    assert "现浇构混凝土桩承台基础C35P8" not in names


def test_item_code_preserves_text_shape():
    assert clean_code_preserve_text(10506001001) == "10506001001"
    assert clean_code_preserve_text(10506001001.0) == "10506001001"
    assert clean_code_preserve_text("1.0506001001E+10") == "10506001001"
    assert clean_code_preserve_text("010506001001") == "010506001001"


def test_long_rebar_feature_does_not_make_valid_row_header_or_subheader(tmp_path):
    xlsx = tmp_path / "mock_bid.xlsx"
    _write_mock_bid_workbook(xlsx)

    records = _parse_mock_workbook(xlsx)
    codes = {record["item_code"] for record in records}

    assert "10506001001" in codes


def test_source_sheet_and_excel_row_number_are_human_checkable(tmp_path):
    xlsx = tmp_path / "mock_bid.xlsx"
    _write_mock_bid_workbook(xlsx)

    records = _parse_mock_workbook(xlsx)
    by_code = {record["item_code"]: record for record in records}

    assert by_code["10506001001"]["source_sheet_name"] == TARGET_SHEET_NAME
    assert by_code["10506001001"]["source_sheet_index"] == 1
    assert by_code["10506001001"]["source_row_index"] == 4
    assert by_code["10506001001"]["source_excel_row_no"] == 5
    assert by_code["10506001001"]["parse_status"] == "parsed"
    assert by_code["10506001001"]["parse_warnings"] == ""


def test_export_includes_parse_and_source_columns(tmp_path):
    conn = sqlite3.connect(tmp_path / "bid.sqlite")
    conn.execute(
        """
        CREATE TABLE import_bid_records (
            id INTEGER PRIMARY KEY,
            batch_id TEXT,
            source_file_name TEXT,
            source_sheet_name TEXT,
            source_sheet_index INTEGER,
            source_row_index INTEGER,
            source_excel_row_no INTEGER,
            mapping_version TEXT,
            parser_confidence NUMERIC,
            review_status TEXT,
            parse_status TEXT,
            parse_warnings TEXT,
            project_name TEXT,
            category TEXT,
            serial_number TEXT,
            item_code TEXT,
            item_name TEXT,
            feature TEXT,
            unit TEXT,
            quantity NUMERIC,
            unit_price NUMERIC,
            total_price NUMERIC,
            imported_at TEXT
        )
        """
    )
    conn.execute(
        """
        INSERT INTO import_bid_records
        (batch_id, source_file_name, source_sheet_name, source_sheet_index,
         source_row_index, source_excel_row_no, mapping_version, parser_confidence, review_status,
         parse_status, parse_warnings, project_name, category, serial_number,
         item_code, item_name, feature, unit, quantity, unit_price, total_price,
         imported_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            "batch",
            "mock.xlsx",
            TARGET_SHEET_NAME,
            1,
            4,
            5,
            "v1.0",
            1.0,
            "pending",
            "parsed",
            "",
            "",
            "混凝土工程",
            "2",
            "10506001001",
            "现浇构件钢筋",
            "长项目特征",
            "t",
            22.22,
            3936.12,
            87460.59,
            "2026-07-04",
        ),
    )
    conn.commit()

    output = tmp_path / "bid_records_temp.xlsx"
    export_bid_records(db_conn=conn, output_file=output)

    wb = load_workbook(output, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    data = [cell.value for cell in ws[2]]

    assert "Excel原始行号" in headers
    assert "解析状态" in headers
    assert "解析警告" in headers
    assert "项目编码" in headers
    assert "项目名称" in headers
    assert "导入时间" in headers
    assert data[headers.index("Excel原始行号")] == 5
    assert data[headers.index("解析状态")] == "parsed"
    assert "现浇构件钢筋" in data
