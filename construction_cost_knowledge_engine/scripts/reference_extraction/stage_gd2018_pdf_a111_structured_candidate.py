#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Stage GD2018-PDF-A111-STRUCTURED-CANDIDATE-1.

Extracts official-PDF sourced structured candidate data for A.1.1 土石方工程.

This stage is a bounded A.1.1 pilot only. It does not write databases,
migrations, schemas.py, src/cost_engine, web_collab_prototype, approved data,
internal_price_library, quota_to_bill_mapping, or source/baseline files.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter, defaultdict
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, DefaultDict, Dict, Iterable, List, Optional, Sequence, Tuple

import pdfplumber
from openpyxl import load_workbook
from pypdf import PdfReader


PROJECT_ROOT_DEFAULT = Path(r"E:\workspace\01_Projects\ai-construction-system")
ENGINE_REL = Path("construction_cost_knowledge_engine")
SOURCE_STANDARDS_REL = ENGINE_REL / "data" / "private" / "reference_extraction" / "source_standards"
SOURCE_EXCELS_REL = ENGINE_REL / "data" / "private" / "reference_extraction" / "source_excels"
PRIOR_RUN_REL = (
    ENGINE_REL
    / "data"
    / "private"
    / "reference_extraction"
    / "runs"
    / "GD2018_PDF_VS_XLSX_DIFFICULTY_ASSESSMENT_1"
)
OUTPUT_DIR_REL = (
    ENGINE_REL
    / "data"
    / "private"
    / "reference_extraction"
    / "runs"
    / "GD2018_PDF_A111_STRUCTURED_CANDIDATE_1"
)

SOURCE_VOLUME = "上册"
CHAPTER_CODE = "A.1.1"
CHAPTER_NAME = "土石方工程"
REVIEW_STATUS = "pending"
SUPPLEMENTAL_CODES = [
    "A1-1-56-1",
    "A1-1-56-2",
    "A1-1-56-3",
    "A1-1-56-4",
    "A1-1-118-1",
    "A1-1-118-2",
]

QUOTA_CODE_RE = re.compile(r"\bA1-1-\d+(?:-\d+)?\b")
RESOURCE_CODE_RE = re.compile(r"\b\d{8,9}(?:-\d{4})?\b")
PRIVATE_USE_RE = re.compile(r"[\ue000-\uf8ff]")
NUMBER_RE = re.compile(r"-?\d+(?:\.\d+)?")

PAGE_SPAN_FIELDS = [
    "source_file",
    "pdf_page_no",
    "book_page_no",
    "detected_chapter_code",
    "detected_chapter_name",
    "detected_section_code",
    "detected_section_name",
    "quota_code_count",
    "resource_code_count",
    "work_content_marker_count",
    "unit_marker_count",
    "is_in_A111_scope",
    "detect_basis",
    "remark",
]

STRUCTURED_FIELDS = [
    "source_file",
    "source_volume",
    "pdf_page_no",
    "book_page_no",
    "source_block_id",
    "chapter_code",
    "chapter_name",
    "section_code",
    "section_name",
    "subsection_code",
    "subsection_name",
    "quota_group_title_1",
    "quota_group_title_2",
    "quota_group_title_3",
    "quota_source_code",
    "quota_name_from_pdf",
    "quota_name_full_from_pdf",
    "quota_unit_raw",
    "quota_unit_normalized",
    "base_price",
    "labor_fee",
    "material_fee",
    "machine_fee",
    "management_fee",
    "total_fee_calculated",
    "price_extract_method",
    "table_header_group_id",
    "column_index_in_table",
    "is_supplemental_quota_code",
    "table_continuation_status",
    "parse_confidence",
    "review_status",
    "reviewer",
    "review_comment",
    "raw_text_sample",
    "remark",
]

WORK_CONTENT_FIELDS = [
    "source_file",
    "pdf_page_no",
    "book_page_no",
    "source_block_id",
    "chapter_code",
    "chapter_name",
    "section_code",
    "section_name",
    "subsection_code",
    "subsection_name",
    "quota_source_code_start",
    "quota_source_code_end",
    "applicable_quota_codes_json",
    "work_content_raw",
    "work_content_normalized",
    "parse_confidence",
    "review_status",
    "remark",
]

QUANTITY_RULE_FIELDS = [
    "source_file",
    "pdf_page_no",
    "book_page_no",
    "source_block_id",
    "chapter_code",
    "chapter_name",
    "rule_no",
    "rule_text_raw",
    "rule_text_normalized",
    "applicable_section",
    "applicable_quota_code_range",
    "parse_confidence",
    "review_status",
    "remark",
]

RESOURCE_DETAIL_FIELDS = [
    "source_file",
    "source_volume",
    "pdf_page_no",
    "book_page_no",
    "source_block_id",
    "chapter_code",
    "chapter_name",
    "section_code",
    "section_name",
    "subsection_code",
    "subsection_name",
    "quota_source_code",
    "quota_name_from_pdf",
    "resource_category_raw",
    "resource_category_normalized",
    "resource_code",
    "resource_name",
    "resource_spec",
    "resource_unit_raw",
    "resource_unit_normalized",
    "resource_unit_price",
    "resource_consumption",
    "resource_fee_calculated",
    "resource_row_index",
    "table_header_group_id",
    "column_index_in_table",
    "is_cross_page_continuation",
    "parse_confidence",
    "review_status",
    "reviewer",
    "review_comment",
    "raw_row_json",
    "remark",
]

RESOURCE_DISPLAY_FIELDS = [
    "source_file",
    "source_volume",
    "pdf_page_no",
    "book_page_no",
    "chapter_code",
    "chapter_name",
    "section_code",
    "section_name",
    "subsection_code",
    "subsection_name",
    "quota_source_code",
    "quota_name_from_pdf",
    "quota_unit_normalized",
    "resource_display_order",
    "resource_category_raw",
    "resource_category_normalized",
    "resource_code",
    "resource_name",
    "resource_spec",
    "resource_unit_raw",
    "resource_unit_normalized",
    "resource_consumption",
    "resource_unit_price",
    "resource_market_price_ex_tax",
    "resource_market_price_tax_included",
    "tax_rate",
    "resource_fee_calculated",
    "is_temporary_price",
    "is_main_material",
    "is_equipment",
    "is_machine",
    "is_labor",
    "table_header_group_id",
    "column_index_in_table",
    "parse_confidence",
    "review_status",
    "raw_row_json",
    "remark",
]

FEE_SUMMARY_FIELDS = [
    "quota_source_code",
    "quota_name_from_pdf",
    "quota_unit_normalized",
    "resource_labor_fee_sum",
    "resource_material_fee_sum",
    "resource_machine_fee_sum",
    "resource_equipment_fee_sum",
    "resource_main_material_fee_sum",
    "resource_other_fee_sum",
    "resource_total_fee_sum",
    "quota_labor_fee_from_main_table",
    "quota_material_fee_from_main_table",
    "quota_machine_fee_from_main_table",
    "quota_management_fee_from_main_table",
    "quota_base_price_from_main_table",
    "delta_labor",
    "delta_material",
    "delta_machine",
    "delta_resource_total_vs_base_price",
    "resource_row_count",
    "resource_reconciliation_status",
    "issue_type",
    "remark",
]

RECONCILIATION_FIELDS = [
    "quota_source_code",
    "quota_name_from_pdf",
    "quota_name_from_xlsx",
    "unit_from_pdf",
    "unit_from_xlsx",
    "base_price_from_pdf",
    "total_from_xlsx",
    "labor_fee_from_pdf",
    "labor_fee_from_xlsx",
    "material_fee_from_pdf",
    "material_fee_from_xlsx",
    "machine_fee_from_pdf",
    "machine_fee_from_xlsx",
    "management_fee_from_pdf",
    "management_fee_from_xlsx",
    "delta_total",
    "delta_labor",
    "delta_material",
    "delta_machine",
    "delta_management",
    "match_status",
    "issue_type",
    "remark",
]

ISSUE_FIELDS = [
    "issue_id",
    "issue_type",
    "severity",
    "pdf_page_no",
    "book_page_no",
    "quota_source_code",
    "resource_code",
    "field_name",
    "issue_detail",
    "suggested_action",
    "raw_text_sample",
]


def compact_text(value: Any, limit: Optional[int] = None) -> str:
    if value is None:
        return ""
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text[:limit] if limit else text


def clean_cell(value: Any) -> str:
    return compact_text(value)


def json_dump(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def code_sort_key(code: str) -> Tuple[int, ...]:
    return tuple(int(part) for part in re.findall(r"\d+", code))


def expected_base_codes() -> List[str]:
    return [f"A1-1-{index}" for index in range(1, 138)]


def normalize_unit(value: Any) -> str:
    text = compact_text(value)
    replacements = {
        "\ue000": "m3",
        "": "m3",
        "㎡": "m2",
        "m²": "m2",
        "m³": "m3",
        "（": "(",
        "）": ")",
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    return text


def positioned_quota_unit(page: Any) -> str:
    """Read the unit header immediately above a quota-code table by coordinates."""
    words = page.extract_words(use_text_flow=False, keep_blank_chars=False) or []
    code_markers = [word for word in words if "定额编号" in word.get("text", "")]
    if not code_markers:
        return ""
    code_marker = min(code_markers, key=lambda word: word["top"])
    unit_markers = [
        word for word in words
        if "计量单位" in word.get("text", "")
        and word["top"] <= code_marker["top"]
        and code_marker["top"] - word["top"] <= 120
    ]
    if not unit_markers:
        return ""
    marker = max(unit_markers, key=lambda word: word["top"])
    marker_text = compact_text(marker.get("text", ""))
    match = re.search(r"计量单位[:：]?(.+)$", marker_text)
    unit_raw = compact_text(match.group(1)) if match else ""
    if not unit_raw:
        same_line = sorted(
            (
                word for word in words
                if word["x0"] >= marker["x1"] - 1
                and abs(word["top"] - marker["top"]) <= 4
            ),
            key=lambda word: word["x0"],
        )
        unit_raw = compact_text("".join(word.get("text", "") for word in same_line))
    if unit_raw == "见表":
        return unit_raw
    return unit_raw if is_unit_like(normalize_unit(unit_raw)) else ""


def normalize_text(value: Any) -> str:
    return compact_text(value).replace("　", " ")


def is_dash(value: Any) -> bool:
    return compact_text(value) in {"", "-", "—", "–", "－"}


def to_decimal(value: Any, dash_as_zero: bool = False) -> Optional[Decimal]:
    if value is None:
        return Decimal("0") if dash_as_zero else None
    text = compact_text(value).replace(",", "")
    if text in {"", "-", "—", "–", "－"}:
        return Decimal("0") if dash_as_zero else None
    match = NUMBER_RE.search(text)
    if not match:
        return None
    try:
        return Decimal(match.group(0))
    except InvalidOperation:
        return None


def money(value: Optional[Decimal]) -> str:
    if value is None:
        return ""
    quantized = value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return format(quantized, "f")


def number_text(value: Optional[Decimal]) -> str:
    if value is None:
        return ""
    normalized = value.normalize()
    if normalized == normalized.to_integral():
        return format(normalized, "f")
    return format(normalized, "f").rstrip("0").rstrip(".")


def decimal_delta(left: Any, right: Any) -> Optional[Decimal]:
    left_dec = to_decimal(left, dash_as_zero=True)
    right_dec = to_decimal(right, dash_as_zero=True)
    if left_dec is None or right_dec is None:
        return None
    return left_dec - right_dec


def abs_le(value: Optional[Decimal], threshold: str) -> bool:
    return value is not None and abs(value) <= Decimal(threshold)


def write_csv(path: Path, fields: Sequence[str], rows: Iterable[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(fields), extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: "" if row.get(field) is None else row.get(field) for field in fields})


def write_jsonl(path: Path, rows: Iterable[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")


def find_pdf(project_root: Path) -> Path:
    root = project_root / SOURCE_STANDARDS_REL
    matches = [
        path
        for path in root.rglob("*.pdf")
        if "房屋建筑与装饰" in path.name and "上册" in path.name
    ]
    if not matches:
        raise FileNotFoundError(f"Cannot find upper-volume 房屋建筑与装饰 PDF under {root}")
    return sorted(matches, key=lambda path: path.name)[0]


def find_xlsx(project_root: Path) -> Path:
    root = project_root / SOURCE_EXCELS_REL
    matches = [path for path in root.glob("*normalized.xlsx") if "房屋建筑与装饰" in path.name]
    if not matches:
        matches = list(root.glob("*normalized.xlsx"))
    if not matches:
        raise FileNotFoundError(f"Cannot find normalized XLSX under {root}")
    return sorted(matches, key=lambda path: path.name)[0]


def detect_book_page_no(text: str, pdf_page_no: int) -> str:
    lines = [line.strip() for line in (text or "").splitlines() if line.strip()]
    for line in lines[:2] + lines[-2:]:
        if re.fullmatch(r"\d{1,3}", line):
            return line
    if 51 <= pdf_page_no <= 704:
        return str(pdf_page_no - 40)
    return ""


def load_pdf_texts(pdf_path: Path) -> List[str]:
    reader = PdfReader(str(pdf_path))
    return [page.extract_text() or "" for page in reader.pages]


def locate_page_span(texts: Sequence[str]) -> Dict[str, int]:
    desc_start = 0
    rule_start = 0
    table_start = 0
    table_end = 0
    boundary = 0
    for index, text in enumerate(texts, start=1):
        if index > 35 and not desc_start and "本章定额包括土方工程" in text and "回填方及其他" in text:
            desc_start = index
        if desc_start and not rule_start and "工程量计算规则" in text:
            rule_start = index
        if not table_start and "A1-1-1" in text:
            table_start = index
        if "A1-1-137" in text:
            table_end = index
        if table_start and not boundary and index > table_start and "A.1.2" in text:
            boundary = index
            break
    if not desc_start or not rule_start or not table_start or not table_end:
        raise RuntimeError("Could not detect A.1.1 page span from PDF text markers")
    return {
        "desc_start": desc_start,
        "rule_start": rule_start,
        "table_start": table_start,
        "table_end": table_end,
        "boundary": boundary or table_end + 1,
        "scope_start": desc_start,
        "scope_end": (boundary - 1) if boundary else table_end,
    }


def section_for_code(code: str) -> Tuple[str, str]:
    number = code_sort_key(code)[-1]
    if 1 <= number <= 66:
        return "A.1.1.1", "土方工程"
    if 67 <= number <= 125:
        return "A.1.1.2", "石方工程"
    if 126 <= number <= 137:
        return "A.1.1.3", "回填方及其他"
    return "", ""


def section_for_page(pdf_page_no: int, text: str, span: Dict[str, int]) -> Tuple[str, str, str]:
    codes = sorted(set(QUOTA_CODE_RE.findall(text)), key=code_sort_key)
    if codes:
        section_code, section_name = section_for_code(codes[0])
        return section_code, section_name, "quota_code_range"
    if span["desc_start"] <= pdf_page_no < span["rule_start"]:
        return "A.1.1_DESCRIPTION", "说明", "description_marker"
    if span["rule_start"] <= pdf_page_no < span["table_start"]:
        return "A.1.1_RULES", "工程量计算规则", "quantity_rule_marker"
    if "A.1.2" in text:
        return "A.1.2", "围护及支护工程", "next_chapter_boundary"
    return "", "", ""


def build_page_span_profile(pdf_path: Path, texts: Sequence[str], span: Dict[str, int]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    start = max(span["scope_start"] - 1, 1)
    end = min(span["boundary"] + 2, len(texts))
    for pdf_page_no in range(start, end + 1):
        text = texts[pdf_page_no - 1]
        section_code, section_name, basis = section_for_page(pdf_page_no, text, span)
        in_scope = span["scope_start"] <= pdf_page_no <= span["scope_end"]
        rows.append(
            {
                "source_file": str(pdf_path),
                "pdf_page_no": pdf_page_no,
                "book_page_no": detect_book_page_no(text, pdf_page_no),
                "detected_chapter_code": CHAPTER_CODE if in_scope or "A.1.2" not in text else "A.1.2",
                "detected_chapter_name": CHAPTER_NAME if in_scope or "A.1.2" not in text else "围护及支护工程",
                "detected_section_code": section_code,
                "detected_section_name": section_name,
                "quota_code_count": len(QUOTA_CODE_RE.findall(text)),
                "resource_code_count": len(RESOURCE_CODE_RE.findall(text)),
                "work_content_marker_count": text.count("工作内容："),
                "unit_marker_count": text.count("计量单位："),
                "is_in_A111_scope": "yes" if in_scope else "no",
                "detect_basis": basis or ("blank_inside_scope" if in_scope else "outside_scope"),
                "remark": "A.1.1 auto span profile; blank pages retained for boundary audit",
            }
        )
    return rows


def load_xlsx_reference(xlsx_path: Path) -> Dict[str, Dict[str, Any]]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = [compact_text(value).replace("\n", "") for value in rows[0]]
    col = {header: index for index, header in enumerate(headers)}
    result: Dict[str, Dict[str, Any]] = {}
    for row_index, row in enumerate(rows[1:], start=2):
        code = compact_text(row[col.get("项目编码", 1)] if len(row) > 1 else "")
        if not QUOTA_CODE_RE.fullmatch(code):
            continue
        result[code] = {
            "source_row": row_index,
            "code": code,
            "name": compact_text(row[col.get("项目名称", 2)]),
            "unit": normalize_unit(row[col.get("计量单位", 3)]),
            "labor": row[col.get("人工费", 4)],
            "material": row[col.get("材料费", 5)],
            "machine": row[col.get("机具费", 6)],
            "management": row[col.get("管理费", 7)],
            "total": row[col.get("合计", 8)],
        }
    return result


def extract_between(text: str, start: str, end: str) -> str:
    match = re.search(re.escape(start) + r"(.*?)" + re.escape(end), text, flags=re.S)
    return compact_text(match.group(1)) if match else ""


def split_lines(value: Any) -> List[str]:
    if value is None:
        return []
    return [compact_text(part) for part in str(value).splitlines() if compact_text(part)]


def is_unit_like(value: str) -> bool:
    text = normalize_unit(value)
    return bool(re.fullmatch(r"(?:\d+(?:\.\d+)?\s*)?(?:m2|m3|100m2|100m3|1000m3|m|kg|t|台班|工日|元)", text))


def filled_values_by_code_col(rows: Sequence[Sequence[Any]], code_cols: Sequence[int]) -> List[List[str]]:
    result: List[List[str]] = []
    for row in rows:
        values = [clean_cell(cell) for cell in row]
        by_code_col: List[str] = []
        last = ""
        for col in code_cols:
            value = values[col] if col < len(values) else ""
            if value:
                last = value
            by_code_col.append(last)
        result.append(by_code_col)
    return result


def derive_subsection(text: str) -> Tuple[str, str]:
    before_work = compact_text((text or "").split("工作内容：", 1)[0])
    for marker in ["A.1.1 土石方工程", "A.1.1.1 土方工程", "A.1.1.2 石方工程", "A.1.1.3 回填方及其他"]:
        before_work = compact_text(before_work.replace(marker, " "))
    before_work = re.sub(r"^\d+\s+", "", before_work)
    code_match = re.match(r"(\(?\d+\)?|\([一二三四五六七八九十]+\))\s*(.*)", before_work)
    if code_match:
        return code_match.group(1), compact_text(code_match.group(2) or before_work)
    return "", before_work


def split_name_spec(value: str) -> Tuple[str, str]:
    text = compact_text(value)
    if not text:
        return "", ""
    if " " in text:
        left, right = text.split(" ", 1)
        if any(ch.isdigit() for ch in right) or right in {"综合", "其他"} or any(unit in right for unit in ["kW", "kg", "t", "m", "台班"]):
            return left, right
    return text, ""


def category_from_raw(
    raw: str, resource_code: str = "", resource_name: str = "",
) -> Tuple[str, str]:
    text = compact_text(raw)
    if resource_code == "99450760" and compact_text(resource_name) == "其他材料费":
        return text or "材料", "材料"
    if "人工" in text or resource_code == "00010010":
        return text or "人工", "人工"
    if "主材" in text:
        return text, "主材"
    if "设备" in text:
        return text, "设备"
    if "机具" in text or "机械" in text or resource_code.startswith("99"):
        return text or "机具", "机具"
    if "材料" in text or resource_code.startswith(("03", "04", "05")):
        return text or "材料", "材料"
    if "管理" in text or "其他" in text:
        return text, "other"
    return text, "unknown"


def display_category(normalized_detail_category: str) -> str:
    mapping = {
        "人工": "labor",
        "材料": "material",
        "机具": "machine",
        "设备": "equipment",
        "主材": "main_material",
        "other": "other",
        "unknown": "unknown",
    }
    return mapping.get(normalized_detail_category, "unknown")


def parse_resource_entries(
    row: Sequence[Any],
    resource_row_index: int,
    code_cols: Sequence[int],
    current_category: str,
) -> List[Dict[str, Any]]:
    cells = ["" if cell is None else str(cell) for cell in row]
    raw_code_cell = cells[1] if len(cells) > 1 else ""
    resource_codes = RESOURCE_CODE_RE.findall(raw_code_cell)
    if not resource_codes:
        return []

    name_lines = split_lines(cells[2] if len(cells) > 2 else "")
    unit_lines = split_lines(cells[3] if len(cells) > 3 else "")
    price_lines = split_lines(cells[4] if len(cells) > 4 else "")
    consumption_columns = [split_lines(cells[col] if col < len(cells) else "") for col in code_cols]

    parsed: List[Dict[str, Any]] = []
    for index, resource_code in enumerate(resource_codes):
        if len(name_lines) == len(resource_codes) * 2:
            resource_name = name_lines[index * 2]
            resource_spec = name_lines[index * 2 + 1]
        elif len(name_lines) == len(resource_codes):
            resource_name, resource_spec = split_name_spec(name_lines[index])
        else:
            source_name = name_lines[index] if index < len(name_lines) else ""
            resource_name, resource_spec = split_name_spec(source_name)
            if not resource_spec and index * 2 + 1 < len(name_lines):
                maybe_spec = name_lines[index * 2 + 1]
                if maybe_spec != resource_name:
                    resource_spec = maybe_spec
        unit_raw = unit_lines[index] if index < len(unit_lines) else (unit_lines[0] if unit_lines else "")
        price_raw = price_lines[index] if index < len(price_lines) else (price_lines[0] if price_lines else "")
        semantic_name = resource_name
        if resource_code == "99450760" and "其他材料费" in compact_text(cells[2] if len(cells) > 2 else ""):
            semantic_name = "其他材料费"
        category_raw, category_normalized = category_from_raw(
            current_category, resource_code, semantic_name,
        )
        parsed.append(
            {
                "resource_code": resource_code,
                "resource_name": resource_name,
                "resource_spec": resource_spec,
                "resource_unit_raw": unit_raw,
                "resource_unit_normalized": normalize_unit(unit_raw),
                "resource_unit_price": price_raw,
                "consumption_by_column": [
                    parts[index] if index < len(parts) else ""
                    for parts in consumption_columns
                ],
                "resource_category_raw": category_raw,
                "resource_category_normalized": category_normalized,
                "resource_row_index": resource_row_index,
                "raw_row_json": json_dump(list(row)),
                "multi_resource_row": len(resource_codes) > 1,
            }
        )
    return parsed


def resource_fee(category: str, unit_price_raw: str, consumption_raw: str) -> Optional[Decimal]:
    consumption = to_decimal(consumption_raw)
    if consumption is None:
        return None
    unit_price = to_decimal(unit_price_raw)
    if unit_price is None and category == "人工":
        return consumption
    if unit_price is None:
        return None
    return consumption * unit_price


def find_row_index(table: Sequence[Sequence[Any]], marker: str) -> Optional[int]:
    for index, row in enumerate(table):
        if any(marker in clean_cell(cell) for cell in row if cell is not None):
            return index
    return None


def parse_table_page(
    pdf_path: Path,
    pdf_page_no: int,
    text: str,
    table: Sequence[Sequence[Any]],
    unit_context_override: str = "",
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Optional[Dict[str, Any]]]:
    code_row_index = None
    for index, row in enumerate(table):
        if any(QUOTA_CODE_RE.fullmatch(clean_cell(cell)) for cell in row if cell is not None):
            code_row_index = index
            break
    if code_row_index is None:
        return [], [], None
    code_row = table[code_row_index]
    code_cols = [index for index, cell in enumerate(code_row) if QUOTA_CODE_RE.fullmatch(clean_cell(cell))]
    codes = [clean_cell(code_row[col]) for col in code_cols]
    if not codes:
        return [], [], None

    base_row_index = find_row_index(table, "基价")
    resource_header_index = find_row_index(table, "分类")
    if base_row_index is None:
        return [], [], None

    name_rows = table[code_row_index + 1 : base_row_index]
    filled_name_rows = filled_values_by_code_col(name_rows, code_cols)
    unit_context_raw = extract_between(text, "计量单位：", "定额编号")
    if not normalize_unit(unit_context_raw):
        unit_context_raw = compact_text(unit_context_override)
    work_content_raw = extract_between(text, "工作内容：", "计量单位：")
    subsection_code, subsection_name = derive_subsection(text)
    table_header_group_id = f"A111_P{pdf_page_no:03d}_T1"
    source_block_id = f"{table_header_group_id}_TABLE"
    book_page_no = detect_book_page_no(text, pdf_page_no)

    candidates: List[Dict[str, Any]] = []
    resource_rows: List[Dict[str, Any]] = []
    code_to_name: Dict[str, str] = {}
    code_to_unit: Dict[str, str] = {}

    price_rows = table[base_row_index : base_row_index + 5]
    for code_index, code in enumerate(codes):
        raw_parts: List[str] = []
        for row_values in filled_name_rows:
            value = row_values[code_index] if code_index < len(row_values) else ""
            if value and value not in raw_parts:
                raw_parts.append(value)
        unit_raw = unit_context_raw
        if "见表" in unit_context_raw and raw_parts and is_unit_like(raw_parts[-1]):
            unit_raw = raw_parts.pop()
        elif raw_parts and is_unit_like(raw_parts[-1]) and (
            not unit_context_raw
            or normalize_unit(raw_parts[-1]) == normalize_unit(unit_context_raw)
        ):
            embedded_unit = raw_parts.pop()
            if not unit_context_raw:
                unit_raw = embedded_unit
        quota_name = compact_text(" ".join(raw_parts))
        code_to_name[code] = quota_name
        code_to_unit[code] = unit_raw

        def price_at(row_offset: int) -> str:
            row_index = base_row_index + row_offset
            if row_index >= len(table):
                return ""
            row = table[row_index]
            col = code_cols[code_index]
            return clean_cell(row[col] if col < len(row) else "")

        base = price_at(0)
        labor = price_at(1)
        material = price_at(2)
        machine = price_at(3)
        management = price_at(4)
        total_calc = sum(
            (to_decimal(value, dash_as_zero=True) or Decimal("0"))
            for value in [labor, material, machine, management]
        )
        confidence = Decimal("0.88")
        if PRIVATE_USE_RE.search(unit_raw):
            confidence -= Decimal("0.04")
        if "见表" in unit_context_raw:
            confidence -= Decimal("0.03")
        if not quota_name:
            confidence -= Decimal("0.15")
        section_code, section_name = section_for_code(code)
        candidates.append(
            {
                "source_file": str(pdf_path),
                "source_volume": SOURCE_VOLUME,
                "pdf_page_no": pdf_page_no,
                "book_page_no": book_page_no,
                "source_block_id": source_block_id,
                "chapter_code": CHAPTER_CODE,
                "chapter_name": CHAPTER_NAME,
                "section_code": section_code,
                "section_name": section_name,
                "subsection_code": subsection_code,
                "subsection_name": subsection_name,
                "quota_group_title_1": subsection_name,
                "quota_group_title_2": raw_parts[0] if raw_parts else "",
                "quota_group_title_3": raw_parts[1] if len(raw_parts) > 1 else "",
                "quota_source_code": code,
                "quota_name_from_pdf": quota_name,
                "quota_name_full_from_pdf": quota_name,
                "quota_unit_raw": unit_raw,
                "quota_unit_normalized": normalize_unit(unit_raw),
                "base_price": base,
                "labor_fee": labor,
                "material_fee": material,
                "machine_fee": machine,
                "management_fee": management,
                "total_fee_calculated": money(total_calc),
                "price_extract_method": "pdfplumber.extract_tables:code_column_price_rows",
                "table_header_group_id": table_header_group_id,
                "column_index_in_table": code_cols[code_index],
                "is_supplemental_quota_code": "yes" if code in SUPPLEMENTAL_CODES else "no",
                "table_continuation_status": "single_page_table",
                "parse_confidence": number_text(confidence),
                "review_status": REVIEW_STATUS,
                "reviewer": "",
                "review_comment": "",
                "raw_text_sample": compact_text(text, 800),
                "remark": "official PDF structured candidate; pending cost QA",
            }
        )

    if resource_header_index is not None:
        current_category = ""
        for row_index, row in enumerate(table[resource_header_index + 1 :], start=resource_header_index + 1):
            first_cell = clean_cell(row[0] if row else "")
            if first_cell:
                current_category = first_cell
            parsed_resources = parse_resource_entries(row, row_index, code_cols, current_category)
            for resource in parsed_resources:
                for code_index, code in enumerate(codes):
                    consumption = resource["consumption_by_column"][code_index] if code_index < len(resource["consumption_by_column"]) else ""
                    if is_dash(consumption):
                        continue
                    fee = resource_fee(resource["resource_category_normalized"], resource["resource_unit_price"], consumption)
                    confidence = Decimal("0.84")
                    if resource["multi_resource_row"]:
                        confidence -= Decimal("0.06")
                    if fee is None:
                        confidence -= Decimal("0.06")
                    if not resource["resource_name"]:
                        confidence -= Decimal("0.08")
                    section_code, section_name = section_for_code(code)
                    resource_rows.append(
                        {
                            "source_file": str(pdf_path),
                            "source_volume": SOURCE_VOLUME,
                            "pdf_page_no": pdf_page_no,
                            "book_page_no": book_page_no,
                            "source_block_id": source_block_id,
                            "chapter_code": CHAPTER_CODE,
                            "chapter_name": CHAPTER_NAME,
                            "section_code": section_code,
                            "section_name": section_name,
                            "subsection_code": subsection_code,
                            "subsection_name": subsection_name,
                            "quota_source_code": code,
                            "quota_name_from_pdf": code_to_name.get(code, ""),
                            "resource_category_raw": resource["resource_category_raw"],
                            "resource_category_normalized": resource["resource_category_normalized"],
                            "resource_code": resource["resource_code"],
                            "resource_name": resource["resource_name"],
                            "resource_spec": resource["resource_spec"],
                            "resource_unit_raw": resource["resource_unit_raw"],
                            "resource_unit_normalized": resource["resource_unit_normalized"],
                            "resource_unit_price": resource["resource_unit_price"],
                            "resource_consumption": consumption,
                            "resource_fee_calculated": money(fee),
                            "resource_row_index": resource["resource_row_index"],
                            "table_header_group_id": table_header_group_id,
                            "column_index_in_table": code_cols[code_index],
                            "is_cross_page_continuation": "no",
                            "parse_confidence": number_text(confidence),
                            "review_status": REVIEW_STATUS,
                            "reviewer": "",
                            "review_comment": "",
                            "raw_row_json": resource["raw_row_json"],
                            "remark": "resource candidate assigned by code-column consumption cell",
                        }
                    )

    work_content_row = {
        "source_file": str(pdf_path),
        "pdf_page_no": pdf_page_no,
        "book_page_no": book_page_no,
        "source_block_id": f"{table_header_group_id}_WORK_CONTENT",
        "chapter_code": CHAPTER_CODE,
        "chapter_name": CHAPTER_NAME,
        "section_code": section_for_code(codes[0])[0],
        "section_name": section_for_code(codes[0])[1],
        "subsection_code": subsection_code,
        "subsection_name": subsection_name,
        "quota_source_code_start": codes[0],
        "quota_source_code_end": codes[-1],
        "applicable_quota_codes_json": json_dump(codes),
        "work_content_raw": work_content_raw,
        "work_content_normalized": normalize_text(work_content_raw),
        "parse_confidence": "0.82" if work_content_raw else "0.55",
        "review_status": REVIEW_STATUS,
        "remark": "group-level work content; applicability requires pending QA",
    }
    return candidates, resource_rows, work_content_row


def extract_pdf_tables(pdf_path: Path, texts: Sequence[str], span: Dict[str, int]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    candidates: List[Dict[str, Any]] = []
    resources: List[Dict[str, Any]] = []
    work_contents: List[Dict[str, Any]] = []
    inherited_unit = ""
    with pdfplumber.open(str(pdf_path)) as pdf:
        for pdf_page_no in range(span["table_start"], span["table_end"] + 1):
            page = pdf.pages[pdf_page_no - 1]
            tables = page.extract_tables() or []
            if not tables:
                continue
            table = max(tables, key=lambda table_rows: len(table_rows))
            positioned_unit = positioned_quota_unit(page)
            if positioned_unit:
                inherited_unit = positioned_unit
            page_candidates, page_resources, work_content = parse_table_page(
                pdf_path,
                pdf_page_no,
                texts[pdf_page_no - 1],
                table,
                inherited_unit,
            )
            candidates.extend(page_candidates)
            resources.extend(page_resources)
            if work_content:
                work_contents.append(work_content)
    return candidates, resources, work_contents


def build_quantity_rules(pdf_path: Path, texts: Sequence[str], span: Dict[str, int]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for pdf_page_no in range(span["rule_start"], span["table_start"]):
        text = texts[pdf_page_no - 1]
        if not compact_text(text):
            continue
        cleaned = normalize_text(text)
        rows.append(
            {
                "source_file": str(pdf_path),
                "pdf_page_no": pdf_page_no,
                "book_page_no": detect_book_page_no(text, pdf_page_no),
                "source_block_id": f"A111_P{pdf_page_no:03d}_QUANTITY_RULE",
                "chapter_code": CHAPTER_CODE,
                "chapter_name": CHAPTER_NAME,
                "rule_no": f"A111_RULE_P{pdf_page_no:03d}",
                "rule_text_raw": cleaned,
                "rule_text_normalized": cleaned,
                "applicable_section": "A.1.1 土石方工程",
                "applicable_quota_code_range": "A1-1-1..A1-1-137",
                "parse_confidence": "0.72",
                "review_status": REVIEW_STATUS,
                "remark": "page-level广东省定额内部工程量计算规则 block; scope pending manual QA",
            }
        )
    return rows


def build_resource_display(resources: Sequence[Dict[str, Any]], candidates_by_code: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    order_by_quota: DefaultDict[str, int] = defaultdict(int)
    for resource in resources:
        code = resource["quota_source_code"]
        order_by_quota[code] += 1
        candidate = candidates_by_code.get(code, {})
        display_cat = display_category(resource["resource_category_normalized"])
        rows.append(
            {
                "source_file": resource["source_file"],
                "source_volume": SOURCE_VOLUME,
                "pdf_page_no": resource["pdf_page_no"],
                "book_page_no": resource["book_page_no"],
                "chapter_code": CHAPTER_CODE,
                "chapter_name": CHAPTER_NAME,
                "section_code": resource["section_code"],
                "section_name": resource["section_name"],
                "subsection_code": resource["subsection_code"],
                "subsection_name": resource["subsection_name"],
                "quota_source_code": code,
                "quota_name_from_pdf": resource["quota_name_from_pdf"],
                "quota_unit_normalized": candidate.get("quota_unit_normalized", ""),
                "resource_display_order": order_by_quota[code],
                "resource_category_raw": resource["resource_category_raw"],
                "resource_category_normalized": display_cat,
                "resource_code": resource["resource_code"],
                "resource_name": resource["resource_name"],
                "resource_spec": resource["resource_spec"],
                "resource_unit_raw": resource["resource_unit_raw"],
                "resource_unit_normalized": resource["resource_unit_normalized"],
                "resource_consumption": resource["resource_consumption"],
                "resource_unit_price": resource["resource_unit_price"],
                "resource_market_price_ex_tax": "",
                "resource_market_price_tax_included": "",
                "tax_rate": "",
                "resource_fee_calculated": resource["resource_fee_calculated"],
                "is_temporary_price": "",
                "is_main_material": "yes" if display_cat == "main_material" else "no",
                "is_equipment": "yes" if display_cat == "equipment" else "no",
                "is_machine": "yes" if display_cat == "machine" else "no",
                "is_labor": "yes" if display_cat == "labor" else "no",
                "table_header_group_id": resource["table_header_group_id"],
                "column_index_in_table": resource["column_index_in_table"],
                "parse_confidence": resource["parse_confidence"],
                "review_status": REVIEW_STATUS,
                "raw_row_json": resource["raw_row_json"],
                "remark": "future resource display layer only; no Web/database write in this stage",
            }
        )
    return rows


def sum_fee(rows: Sequence[Dict[str, Any]], category: str) -> Decimal:
    total = Decimal("0")
    for row in rows:
        if row["resource_category_normalized"] == category:
            total += to_decimal(row.get("resource_fee_calculated"), dash_as_zero=True) or Decimal("0")
    return total


def build_fee_summary(candidates: Sequence[Dict[str, Any]], resources: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    resources_by_code: DefaultDict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in resources:
        resources_by_code[row["quota_source_code"]].append(row)
    rows: List[Dict[str, Any]] = []
    for candidate in candidates:
        code = candidate["quota_source_code"]
        detail_rows = resources_by_code.get(code, [])
        labor_sum = sum_fee(detail_rows, "人工")
        material_sum = sum_fee(detail_rows, "材料")
        machine_sum = sum_fee(detail_rows, "机具")
        equipment_sum = sum_fee(detail_rows, "设备")
        main_material_sum = sum_fee(detail_rows, "主材")
        other_sum = sum_fee(detail_rows, "other") + sum_fee(detail_rows, "unknown")
        resource_total = labor_sum + material_sum + machine_sum + equipment_sum + main_material_sum + other_sum
        labor_delta = labor_sum - (to_decimal(candidate["labor_fee"], dash_as_zero=True) or Decimal("0"))
        material_delta = material_sum - (to_decimal(candidate["material_fee"], dash_as_zero=True) or Decimal("0"))
        machine_delta = machine_sum - (to_decimal(candidate["machine_fee"], dash_as_zero=True) or Decimal("0"))
        base = to_decimal(candidate["base_price"], dash_as_zero=True) or Decimal("0")
        management = to_decimal(candidate["management_fee"], dash_as_zero=True) or Decimal("0")
        total_delta = resource_total - base
        status = "cannot_reconcile"
        issue_type = ""
        if not detail_rows:
            status = "resource_rows_missing"
            issue_type = "resource_rows_missing"
        elif abs(labor_delta) <= Decimal("0.01") and abs(material_delta) <= Decimal("0.01") and abs(machine_delta) <= Decimal("0.01"):
            if abs((resource_total + management) - base) <= Decimal("0.01"):
                status = "matched"
            elif abs((resource_total + management) - base) <= Decimal("0.05"):
                status = "matched_rounding_delta"
            else:
                status = "partial_match"
                issue_type = "resource_sum_reconciliation_delta"
        elif resource_total > base:
            status = "resource_sum_exceeds_main_price"
            issue_type = "resource_sum_reconciliation_delta"
        else:
            status = "resource_sum_below_main_price"
            issue_type = "resource_sum_reconciliation_delta"
        rows.append(
            {
                "quota_source_code": code,
                "quota_name_from_pdf": candidate["quota_name_from_pdf"],
                "quota_unit_normalized": candidate["quota_unit_normalized"],
                "resource_labor_fee_sum": money(labor_sum),
                "resource_material_fee_sum": money(material_sum),
                "resource_machine_fee_sum": money(machine_sum),
                "resource_equipment_fee_sum": money(equipment_sum),
                "resource_main_material_fee_sum": money(main_material_sum),
                "resource_other_fee_sum": money(other_sum),
                "resource_total_fee_sum": money(resource_total),
                "quota_labor_fee_from_main_table": candidate["labor_fee"],
                "quota_material_fee_from_main_table": candidate["material_fee"],
                "quota_machine_fee_from_main_table": candidate["machine_fee"],
                "quota_management_fee_from_main_table": candidate["management_fee"],
                "quota_base_price_from_main_table": candidate["base_price"],
                "delta_labor": money(labor_delta),
                "delta_material": money(material_delta),
                "delta_machine": money(machine_delta),
                "delta_resource_total_vs_base_price": money(total_delta),
                "resource_row_count": len(detail_rows),
                "resource_reconciliation_status": status,
                "issue_type": issue_type,
                "remark": "resource total excludes management fee; compare category deltas before judging mismatch",
            }
        )
    return rows


def build_reconciliation(candidates_by_code: Dict[str, Dict[str, Any]], xlsx_by_code: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    expected = expected_base_codes() + SUPPLEMENTAL_CODES
    for code in expected:
        pdf_row = candidates_by_code.get(code)
        xlsx_row = xlsx_by_code.get(code)
        if not pdf_row and not xlsx_row:
            match_status = "missing_in_pdf"
            issue_type = "pdf_missing_code;xlsx_missing_code"
        elif not pdf_row:
            match_status = "missing_in_pdf"
            issue_type = "pdf_missing_code"
        elif not xlsx_row:
            match_status = "missing_in_xlsx"
            issue_type = "xlsx_missing_code"
        else:
            deltas = {
                "total": decimal_delta(pdf_row["base_price"], xlsx_row["total"]),
                "labor": decimal_delta(pdf_row["labor_fee"], xlsx_row["labor"]),
                "material": decimal_delta(pdf_row["material_fee"], xlsx_row["material"]),
                "machine": decimal_delta(pdf_row["machine_fee"], xlsx_row["machine"]),
                "management": decimal_delta(pdf_row["management_fee"], xlsx_row["management"]),
            }
            name_diff = compact_text(pdf_row["quota_name_from_pdf"]) != compact_text(xlsx_row["name"])
            unit_diff = normalize_unit(pdf_row["quota_unit_raw"]) != normalize_unit(xlsx_row["unit"])
            if all(abs_le(delta, "0.0001") for delta in deltas.values()):
                if name_diff and not unit_diff:
                    match_status = "matched_name_diff_only"
                elif unit_diff and not name_diff:
                    match_status = "matched_unit_diff_only"
                else:
                    match_status = "matched_exact"
            elif all(abs_le(delta, "0.011") for delta in deltas.values()):
                match_status = "matched_rounding_delta"
            else:
                match_status = "price_mismatch"
            issue_parts: List[str] = []
            if match_status == "price_mismatch":
                issue_parts.append("xlsx_price_delta")
            if unit_diff:
                issue_parts.append("unit_difference")
            if name_diff:
                issue_parts.append("name_difference")
            issue_type = ";".join(issue_parts)
        pdf = pdf_row or {}
        xlsx = xlsx_row or {}
        rows.append(
            {
                "quota_source_code": code,
                "quota_name_from_pdf": pdf.get("quota_name_from_pdf", ""),
                "quota_name_from_xlsx": xlsx.get("name", ""),
                "unit_from_pdf": pdf.get("quota_unit_normalized", ""),
                "unit_from_xlsx": xlsx.get("unit", ""),
                "base_price_from_pdf": pdf.get("base_price", ""),
                "total_from_xlsx": xlsx.get("total", ""),
                "labor_fee_from_pdf": pdf.get("labor_fee", ""),
                "labor_fee_from_xlsx": xlsx.get("labor", ""),
                "material_fee_from_pdf": pdf.get("material_fee", ""),
                "material_fee_from_xlsx": xlsx.get("material", ""),
                "machine_fee_from_pdf": pdf.get("machine_fee", ""),
                "machine_fee_from_xlsx": xlsx.get("machine", ""),
                "management_fee_from_pdf": pdf.get("management_fee", ""),
                "management_fee_from_xlsx": xlsx.get("management", ""),
                "delta_total": money(decimal_delta(pdf.get("base_price", ""), xlsx.get("total", ""))) if pdf and xlsx else "",
                "delta_labor": money(decimal_delta(pdf.get("labor_fee", ""), xlsx.get("labor", ""))) if pdf and xlsx else "",
                "delta_material": money(decimal_delta(pdf.get("material_fee", ""), xlsx.get("material", ""))) if pdf and xlsx else "",
                "delta_machine": money(decimal_delta(pdf.get("machine_fee", ""), xlsx.get("machine", ""))) if pdf and xlsx else "",
                "delta_management": money(decimal_delta(pdf.get("management_fee", ""), xlsx.get("management", ""))) if pdf and xlsx else "",
                "match_status": match_status,
                "issue_type": issue_type,
                "remark": "normalized Excel is checksum reference only; PDF remains source of truth",
            }
        )
    return rows


def add_issue(
    issues: List[Dict[str, Any]],
    issue_type: str,
    severity: str,
    detail: str,
    suggested_action: str,
    pdf_page_no: Any = "",
    book_page_no: Any = "",
    quota_source_code: str = "",
    resource_code: str = "",
    field_name: str = "",
    raw_text_sample: str = "",
) -> None:
    issues.append(
        {
            "issue_id": f"A111_ISSUE_{len(issues) + 1:05d}",
            "issue_type": issue_type,
            "severity": severity,
            "pdf_page_no": pdf_page_no,
            "book_page_no": book_page_no,
            "quota_source_code": quota_source_code,
            "resource_code": resource_code,
            "field_name": field_name,
            "issue_detail": detail,
            "suggested_action": suggested_action,
            "raw_text_sample": compact_text(raw_text_sample, 500),
        }
    )


def build_issues(
    texts: Sequence[str],
    candidates: Sequence[Dict[str, Any]],
    resources: Sequence[Dict[str, Any]],
    display_rows: Sequence[Dict[str, Any]],
    summaries: Sequence[Dict[str, Any]],
    reconciliation: Sequence[Dict[str, Any]],
    work_contents: Sequence[Dict[str, Any]],
    quantity_rules: Sequence[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    issues: List[Dict[str, Any]] = []
    add_issue(
        issues,
        "watermark_visual_interference",
        "low",
        "Rendered A.1.1 table pages contain a diagonal visual watermark, although text extraction is not blocked.",
        "Use PDF source page screenshots during QA when table text looks ambiguous.",
        pdf_page_no="59;99",
        field_name="visual_table_layer",
    )
    for row in candidates:
        if PRIVATE_USE_RE.search(row.get("quota_unit_raw", "")):
            add_issue(
                issues,
                "private_unicode_unit",
                "medium",
                "Quota unit contains private-use glyph normalized to m3.",
                "Verify raw/normalized unit during cost QA.",
                row["pdf_page_no"],
                row["book_page_no"],
                row["quota_source_code"],
                field_name="quota_unit_raw",
                raw_text_sample=row.get("quota_unit_raw", ""),
            )
    for row in work_contents:
        add_issue(
            issues,
            "work_content_scope_uncertain",
            "medium",
            "Work content is extracted at quota-group scope and requires human confirmation of applicable codes.",
            "Confirm applicable_quota_codes_json against PDF table header during QA.",
            row["pdf_page_no"],
            row["book_page_no"],
            row["quota_source_code_start"],
            field_name="applicable_quota_codes_json",
            raw_text_sample=row.get("work_content_raw", ""),
        )
    for row in quantity_rules:
        add_issue(
            issues,
            "quantity_rule_scope_uncertain",
            "medium",
            "Quantity rule is retained as a page-level Guangdong quota rule block.",
            "Cost department should split or tag rule applicability before downstream use.",
            row["pdf_page_no"],
            row["book_page_no"],
            field_name="applicable_section",
            raw_text_sample=row.get("rule_text_raw", ""),
        )
    for row in resources:
        confidence = to_decimal(row.get("parse_confidence")) or Decimal("0")
        if confidence < Decimal("0.80"):
            add_issue(
                issues,
                "low_parse_confidence",
                "medium",
                "Resource row confidence is below 0.80, usually due to merged multi-resource PDF rows.",
                "Manually verify resource name/spec/unit/consumption alignment.",
                row["pdf_page_no"],
                row["book_page_no"],
                row["quota_source_code"],
                row["resource_code"],
                "resource_row",
                row.get("raw_row_json", ""),
            )
            add_issue(
                issues,
                "resource_quota_alignment_uncertain",
                "medium",
                "Merged PDF resource row may affect exact quota-column assignment.",
                "Check the resource consumption cell under the quota column in the official PDF.",
                row["pdf_page_no"],
                row["book_page_no"],
                row["quota_source_code"],
                row["resource_code"],
                "resource_consumption",
                row.get("raw_row_json", ""),
            )
        if row["resource_category_normalized"] == "unknown":
            add_issue(
                issues,
                "resource_category_unknown",
                "medium",
                "Resource category could not be normalized.",
                "Assign 人工/材料/机具/设备/主材/other during QA.",
                row["pdf_page_no"],
                row["book_page_no"],
                row["quota_source_code"],
                row["resource_code"],
                "resource_category_normalized",
                row.get("raw_row_json", ""),
            )
        if not row.get("resource_fee_calculated"):
            if is_dash(row.get("resource_unit_price")):
                add_issue(
                    issues,
                    "resource_fee_calculation_missing_price",
                    "medium",
                    "Resource fee cannot be calculated because unit price is missing/dash.",
                    "Verify whether PDF row is amount-style or requires external price source.",
                    row["pdf_page_no"],
                    row["book_page_no"],
                    row["quota_source_code"],
                    row["resource_code"],
                    "resource_unit_price",
                    row.get("raw_row_json", ""),
                )
            if is_dash(row.get("resource_consumption")):
                add_issue(
                    issues,
                    "resource_fee_calculation_missing_consumption",
                    "medium",
                    "Resource fee cannot be calculated because consumption is missing/dash.",
                    "Verify quota-column resource consumption.",
                    row["pdf_page_no"],
                    row["book_page_no"],
                    row["quota_source_code"],
                    row["resource_code"],
                    "resource_consumption",
                    row.get("raw_row_json", ""),
                )
    for row in display_rows:
        confidence = to_decimal(row.get("parse_confidence")) or Decimal("0")
        if confidence < Decimal("0.80"):
            add_issue(
                issues,
                "resource_display_row_parse_uncertain",
                "medium",
                "Resource display row is generated from a low-confidence resource detail row.",
                "Verify before showing in future Web resource display.",
                row["pdf_page_no"],
                row["book_page_no"],
                row["quota_source_code"],
                row["resource_code"],
                "resource_display_order",
                row.get("raw_row_json", ""),
            )
        add_issue(
            issues,
            "resource_market_price_missing",
            "low",
            "PDF quota resource table does not provide market price excluding tax.",
            "Leave blank until a market price source is explicitly introduced.",
            row["pdf_page_no"],
            row["book_page_no"],
            row["quota_source_code"],
            row["resource_code"],
            "resource_market_price_ex_tax",
        )
        add_issue(
            issues,
            "resource_tax_rate_missing",
            "low",
            "PDF quota resource table does not provide resource tax rate.",
            "Leave blank until a tax/market price source is explicitly introduced.",
            row["pdf_page_no"],
            row["book_page_no"],
            row["quota_source_code"],
            row["resource_code"],
            "tax_rate",
        )
    add_issue(
        issues,
        "resource_main_material_uncertain",
        "low",
        "A.1.1 PDF resource table does not explicitly mark main material flags.",
        "Keep is_main_material=no unless cost QA or a later source identifies main materials.",
        field_name="is_main_material",
    )
    add_issue(
        issues,
        "resource_equipment_uncertain",
        "low",
        "A.1.1 PDF resource table does not explicitly mark equipment resources in this extraction.",
        "Keep is_equipment=no unless cost QA or a later source identifies equipment.",
        field_name="is_equipment",
    )
    add_issue(
        issues,
        "cross_page_continuation_uncertain",
        "low",
        "This stage marks A.1.1 tables as single-page tables; cross-page continuation remains a QA checkpoint for scaling.",
        "Verify long resource tables before extending parser rules to other chapters.",
        field_name="table_continuation_status",
    )
    for row in summaries:
        if row.get("issue_type"):
            add_issue(
                issues,
                "resource_sum_reconciliation_delta",
                "medium",
                "Resource fee summary differs from main table fee components or base price.",
                "Review category sums; management fee and adjustment rows may explain expected differences.",
                quota_source_code=row["quota_source_code"],
                field_name="resource_total_fee_sum",
            )
    for row in reconciliation:
        if row["match_status"] == "missing_in_pdf":
            add_issue(
                issues,
                "pdf_missing_code",
                "high" if row["quota_source_code"] in SUPPLEMENTAL_CODES else "medium",
                "Code exists in checksum set but was not found as an official PDF A.1.1 table code.",
                "Confirm whether the code comes from a supplement outside the supplied PDF before downstream use.",
                quota_source_code=row["quota_source_code"],
                field_name="quota_source_code",
            )
        if row["match_status"] == "missing_in_xlsx":
            add_issue(
                issues,
                "xlsx_missing_code",
                "medium",
                "PDF code is missing from normalized Excel checksum reference.",
                "Confirm Excel source completeness; do not discard PDF candidate.",
                quota_source_code=row["quota_source_code"],
                field_name="quota_source_code",
            )
        if row["match_status"] == "price_mismatch":
            add_issue(
                issues,
                "xlsx_price_delta",
                "high",
                "PDF-vs-XLSX price delta exceeds rounding tolerance.",
                "Manually verify PDF table column, Excel row, and unit normalization.",
                quota_source_code=row["quota_source_code"],
                field_name="base_price_from_pdf",
            )
        if "unit_difference" in row.get("issue_type", ""):
            add_issue(
                issues,
                "unit_normalization_difference",
                "low",
                "PDF normalized unit differs from Excel unit.",
                "Verify whether this is a harmless m2/m3 glyph normalization issue.",
                quota_source_code=row["quota_source_code"],
                field_name="unit_from_pdf",
            )
    missing_base = [code for code in expected_base_codes() if code not in {row["quota_source_code"] for row in candidates}]
    for code in missing_base:
        add_issue(
            issues,
            "quota_code_parse_uncertain",
            "blocking",
            "Expected base A1-1 code was not extracted from PDF candidate table.",
            "Fix parser before using candidate output.",
            quota_source_code=code,
            field_name="quota_source_code",
        )
    return issues


def build_raw_blocks(pdf_path: Path, sample_pages: Sequence[int]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for pdf_page_no in sample_pages:
            page = pdf.pages[pdf_page_no - 1]
            text = page.extract_text() or ""
            lines = page.extract_text_lines(layout=False) if hasattr(page, "extract_text_lines") else []
            for index, line in enumerate(lines, start=1):
                rows.append(
                    {
                        "pdf_page_no": pdf_page_no,
                        "book_page_no": detect_book_page_no(text, pdf_page_no),
                        "block_id": f"P{pdf_page_no:03d}_L{index:03d}",
                        "bbox": [line.get("x0"), line.get("top"), line.get("x1"), line.get("bottom")],
                        "text": line.get("text", ""),
                        "extraction_method": "pdfplumber.extract_text_lines",
                    }
                )
    return rows


def summarize_counts(
    candidates: Sequence[Dict[str, Any]],
    resources: Sequence[Dict[str, Any]],
    display_rows: Sequence[Dict[str, Any]],
    summaries: Sequence[Dict[str, Any]],
    reconciliation: Sequence[Dict[str, Any]],
    issues: Sequence[Dict[str, Any]],
) -> Dict[str, Any]:
    candidate_codes = {row["quota_source_code"] for row in candidates}
    expected = set(expected_base_codes())
    coverage_count = len(expected & candidate_codes)
    supplemental_in_pdf = sum(1 for row in candidates if row["quota_source_code"] in SUPPLEMENTAL_CODES)
    supplemental_in_recon_missing_pdf = sum(1 for row in reconciliation if row["quota_source_code"] in SUPPLEMENTAL_CODES and row["match_status"] == "missing_in_pdf")
    return {
        "candidate_rows": len(candidates),
        "unique_candidate_codes": len(candidate_codes),
        "coverage_count": coverage_count,
        "coverage_rate": coverage_count / len(expected) if expected else 0,
        "missing_expected_codes": sorted(expected - candidate_codes, key=code_sort_key),
        "duplicate_codes": [code for code, count in Counter(row["quota_source_code"] for row in candidates).items() if count > 1],
        "supplemental_in_pdf": supplemental_in_pdf,
        "supplemental_missing_in_pdf": supplemental_in_recon_missing_pdf,
        "resource_rows": len(resources),
        "display_rows": len(display_rows),
        "summary_status_counts": dict(Counter(row["resource_reconciliation_status"] for row in summaries)),
        "recon_status_counts": dict(Counter(row["match_status"] for row in reconciliation)),
        "issues_by_severity": dict(Counter(row["severity"] for row in issues)),
        "high_blocking_issues": sum(1 for row in issues if row["severity"] in {"high", "blocking"}),
    }


def write_report(
    path: Path,
    pdf_path: Path,
    xlsx_path: Path,
    prior_run: Path,
    span: Dict[str, int],
    candidates: Sequence[Dict[str, Any]],
    work_contents: Sequence[Dict[str, Any]],
    quantity_rules: Sequence[Dict[str, Any]],
    resources: Sequence[Dict[str, Any]],
    display_rows: Sequence[Dict[str, Any]],
    summaries: Sequence[Dict[str, Any]],
    reconciliation: Sequence[Dict[str, Any]],
    issues: Sequence[Dict[str, Any]],
) -> None:
    summary = summarize_counts(candidates, resources, display_rows, summaries, reconciliation, issues)
    category_counts = Counter(row["resource_category_normalized"] for row in resources)
    display_category_counts = Counter(row["resource_category_normalized"] for row in display_rows)
    resource_quota_codes = {row["quota_source_code"] for row in resources}
    display_quota_codes = {row["quota_source_code"] for row in display_rows}
    low_conf_resources = sum(1 for row in resources if (to_decimal(row["parse_confidence"]) or Decimal("0")) < Decimal("0.80"))
    low_conf_display = sum(1 for row in display_rows if (to_decimal(row["parse_confidence"]) or Decimal("0")) < Decimal("0.80"))
    resource_alignment_uncertain = sum(1 for row in issues if row["issue_type"] == "resource_quota_alignment_uncertain")
    fee_non_empty = sum(1 for row in display_rows if row.get("resource_fee_calculated"))
    name_diff_only = sum(1 for row in reconciliation if row["match_status"] == "matched_name_diff_only")
    unit_diff = sum(1 for row in reconciliation if "unit_difference" in row.get("issue_type", ""))
    report = f"""# Stage GD2018-PDF-A111-STRUCTURED-CANDIDATE-1 Report

## 1. Task Scope

本轮只做 A.1.1 土石方工程 PDF 官方源结构化候选抽取，不写数据库，不生成 approved，不修改 PDF、normalized Excel、baseline、Web 或现有 pipeline。

## 2. Inputs

- PDF official source: `{pdf_path}`
- normalized Excel checksum reference: `{xlsx_path}`
- prior difficulty assessment run: `{prior_run}`

## 3. A.1.1 Page Span Detection

自动定位依据：正文“本章定额包括土方工程、石方工程、回填方及其他”定位说明起点；“工程量计算规则”定位规则段；`A1-1-1` 定位定额表起点；`A1-1-137` 定位 A.1.1 最后主表；下一章 `A.1.2` 定位停止边界。

- description start pdf page: {span["desc_start"]}
- quantity rule start pdf page: {span["rule_start"]}
- quota table start pdf page: {span["table_start"]}
- quota table end pdf page: {span["table_end"]}
- next chapter boundary pdf page: {span["boundary"]}

## 4. Main Quota Extraction Result

- PDF quota rows: {summary["candidate_rows"]}
- unique quota_source_code: {summary["unique_candidate_codes"]}
- expected A1-1 base code coverage: {summary["coverage_count"]}/137 ({summary["coverage_rate"]:.2%})
- supplemental code count in PDF candidates: {summary["supplemental_in_pdf"]}
- supplemental codes present in XLSX but missing in supplied PDF: {summary["supplemental_missing_in_pdf"]}
- missing expected code: {json_dump(summary["missing_expected_codes"])}
- duplicate code: {json_dump(summary["duplicate_codes"])}
- low confidence count: {sum(1 for row in candidates if (to_decimal(row["parse_confidence"]) or Decimal("0")) < Decimal("0.80"))}

## 5. Work Content Extraction Result

- work content rows: {len(work_contents)}
- applicability scope pending QA rows: {len(work_contents)}

## 6. Quantity Rule Extraction Result

- quantity rule page-level rows: {len(quantity_rules)}
- scope uncertain rows: {len(quantity_rules)}

## 7. Resource Detail Extraction Result

- resource rows: {len(resources)}
- resource rows by category: {json_dump(dict(category_counts))}
- quota codes with resource detail: {len(resource_quota_codes)}
- quota codes without resource detail: {len(set(row["quota_source_code"] for row in candidates) - resource_quota_codes)}
- low confidence resource rows: {low_conf_resources}
- resource alignment uncertain rows: {resource_alignment_uncertain}

## 8. Resource Display Layer Result

- quota_pdf_resource_display_A111_candidate.csv rows: {len(display_rows)}
- quota_source_code with resource detail: {len(display_quota_codes)}
- quota_source_code without resource detail: {len(set(row["quota_source_code"] for row in candidates) - display_quota_codes)}
- labor / material / machine / equipment / main_material / unknown counts: {json_dump(dict(display_category_counts))}
- resource_fee_calculated non-empty rows: {fee_non_empty}
- resource_sum reconciliation status counts: {json_dump(summary["summary_status_counts"])}
- resource_sum reconciliation issue count: {sum(1 for row in summaries if row.get("issue_type"))}
- low confidence display rows: {low_conf_display}

本层数据只用于未来 Web 中在省定额子目下方展示“工料机显示”的数据准备。本轮不进入 Web、不写数据库、不作为 approved。

## 9. PDF-vs-XLSX Reconciliation

- exact match: {summary["recon_status_counts"].get("matched_exact", 0)}
- rounding delta: {summary["recon_status_counts"].get("matched_rounding_delta", 0)}
- price mismatch: {summary["recon_status_counts"].get("price_mismatch", 0)}
- missing in PDF: {summary["recon_status_counts"].get("missing_in_pdf", 0)}
- missing in XLSX: {summary["recon_status_counts"].get("missing_in_xlsx", 0)}
- unit difference: {unit_diff}
- name difference only: {name_diff_only}

## 10. Key Risks

- 横向多子目：PDF 一页内 1-5 个子目横向排列，名称和单位需要按列继承。
- 资源明细归属：资源行存在纵向合并和多资源合并行，必须保留 raw_row_json。
- 私有 Unicode 单位：m3 存在私有区字形，已保留 raw 并 normalized。
- 跨页延续：本轮 A.1.1 表格按单页抽取，扩展到其他章节前仍需 QA。
- 工作内容适用范围：工作内容按 quota group 抽取，适用编码需要人工确认。
- 工程量规则适用范围：规则按页级 block 抽取，不应与 GB/T 50854 混淆。
- PDF 与 Excel 扁平名称差异：Excel 名称是扁平 checksum reference，不作为失败依据。
- resource_sum 与主表四费不完全一致可能正常：主表含管理费、调整项或金额型资源行，不能机械判错。

## 11. Manual QA Recommendation

- 先抽 30 条主项；
- 再抽 30 条资源明细；
- 必查 A1-1-56 系列；
- 必查 A1-1-118 系列；
- 必查 A1-1-134 ~ A1-1-137；
- 必查价格 delta；
- 必查单位转换；
- 必查 resource_display 层的资源归属。

## 12. Go / No-Go Recommendation

go_for_A111_pdf_xlsx_reconciliation_QA_pack
"""
    path.write_text(report, encoding="utf-8")


def run(project_root: Path) -> Dict[str, Path]:
    pdf_path = find_pdf(project_root)
    xlsx_path = find_xlsx(project_root)
    prior_run = project_root / PRIOR_RUN_REL
    output_dir = project_root / OUTPUT_DIR_REL
    output_dir.mkdir(parents=True, exist_ok=True)

    texts = load_pdf_texts(pdf_path)
    span = locate_page_span(texts)
    page_span_rows = build_page_span_profile(pdf_path, texts, span)
    xlsx_by_code = load_xlsx_reference(xlsx_path)
    candidates, resources, work_contents = extract_pdf_tables(pdf_path, texts, span)
    candidates_by_code = {row["quota_source_code"]: row for row in candidates}
    quantity_rules = build_quantity_rules(pdf_path, texts, span)
    display_rows = build_resource_display(resources, candidates_by_code)
    fee_summary = build_fee_summary(candidates, resources)
    reconciliation = build_reconciliation(candidates_by_code, xlsx_by_code)
    issues = build_issues(texts, candidates, resources, display_rows, fee_summary, reconciliation, work_contents, quantity_rules)
    raw_blocks = build_raw_blocks(pdf_path, sample_pages=[59, 60, 76, 92, 99])

    paths = {
        "page_span": output_dir / "pdf_a111_page_span_profile.csv",
        "structured": output_dir / "quota_pdf_structured_A111_candidate.csv",
        "resource_detail": output_dir / "quota_pdf_resource_detail_A111_candidate.csv",
        "resource_display": output_dir / "quota_pdf_resource_display_A111_candidate.csv",
        "resource_fee_summary": output_dir / "quota_pdf_resource_fee_summary_A111.csv",
        "work_content": output_dir / "quota_pdf_work_content_A111_candidate.csv",
        "quantity_rule": output_dir / "quota_pdf_quantity_rule_A111_candidate.csv",
        "reconciliation": output_dir / "quota_pdf_xlsx_reconciliation_A111.csv",
        "issues": output_dir / "quota_pdf_extraction_issues_A111.csv",
        "raw_blocks": output_dir / "quota_pdf_raw_blocks_A111_sample.jsonl",
        "report": output_dir / "stage_gd2018_pdf_a111_structured_candidate_report.md",
    }
    write_csv(paths["page_span"], PAGE_SPAN_FIELDS, page_span_rows)
    write_csv(paths["structured"], STRUCTURED_FIELDS, candidates)
    write_csv(paths["resource_detail"], RESOURCE_DETAIL_FIELDS, resources)
    write_csv(paths["resource_display"], RESOURCE_DISPLAY_FIELDS, display_rows)
    write_csv(paths["resource_fee_summary"], FEE_SUMMARY_FIELDS, fee_summary)
    write_csv(paths["work_content"], WORK_CONTENT_FIELDS, work_contents)
    write_csv(paths["quantity_rule"], QUANTITY_RULE_FIELDS, quantity_rules)
    write_csv(paths["reconciliation"], RECONCILIATION_FIELDS, reconciliation)
    write_csv(paths["issues"], ISSUE_FIELDS, issues)
    write_jsonl(paths["raw_blocks"], raw_blocks)
    write_report(paths["report"], pdf_path, xlsx_path, prior_run, span, candidates, work_contents, quantity_rules, resources, display_rows, fee_summary, reconciliation, issues)
    return paths


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--project-root", type=Path, default=PROJECT_ROOT_DEFAULT)
    return parser.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)
    try:
        paths = run(args.project_root)
    except Exception as exc:  # pragma: no cover - command-line guard
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    print("Generated A.1.1 structured candidate artifacts:")
    for name, path in paths.items():
        print(f"- {name}: {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
