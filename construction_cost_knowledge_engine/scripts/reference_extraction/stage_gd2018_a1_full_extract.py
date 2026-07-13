#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Stage GD2018-A1-FULL extraction.

Extracts all Guangdong 2018 A1 quota reference candidates from the normalized
Excel workbook. Outputs are reference candidates only. This script does not
write databases, approvals, internal_price_library, bill_code mappings, or
enterprise templates, and it does not modify the source workbook.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence, Tuple

from openpyxl import load_workbook


PROJECT_ROOT_DEFAULT = Path(r"E:\workspace\01_Projects\ai-construction-system")
ENGINE_REL = Path("construction_cost_knowledge_engine")
EXCEL_REL = ENGINE_REL / "data" / "private" / "reference_extraction" / "source_excels" / "广东省房屋建筑与装饰工程综合定额2018_normalized.xlsx"
OUTPUT_DIR_REL = ENGINE_REL / "data" / "private" / "reference_extraction" / "runs" / "GD2018_stage2R_A1_full"

SOURCE_TYPE = "reference_excel_normalized_from_third_party"
SOURCE_NAME = "广东省房屋建筑与装饰工程综合定额2018"
PRICING_SOURCE = "广东省房屋建筑与装饰工程综合定额2018_normalized_excel"
PRICING_VERSION = "GD2018_reference_excel_normalized"
SOURCE_TRUST_LEVEL = "L1"
VERIFICATION_STATUS = "structure_checked"
REVIEW_STATUS = "pending"
CHAPTER_CODE = "A.1"
CHAPTER_NAME = "广东省房屋建筑与装饰工程综合定额 A1"
OFFICIAL_SOURCE_FILE = "广东省房屋建筑与装饰工程综合定额2018"
OFFICIAL_SOURCE_PAGE_EVIDENCE_TYPE = "pending"

FIELD_ALIASES = {
    "source_code": ["source_code", "定额编号", "项目编码", "编码"],
    "raw_name": ["raw_name", "项目名称", "定额名称", "名称"],
    "raw_spec_model": ["raw_spec_model", "规格型号", "规格", "型号"],
    "raw_unit": ["raw_unit", "计量单位", "单位"],
    "raw_labor_fee": ["raw_labor_fee", "人工费"],
    "raw_material_fee": ["raw_material_fee", "材料费"],
    "raw_machine_fee": ["raw_machine_fee", "机具费", "机械费"],
    "raw_management_fee": ["raw_management_fee", "管理费"],
    "raw_total_fee": ["raw_total_fee", "合计", "综合单价", "单价"],
}

RAW_FIELDS = [
    "raw_row_id",
    "source_type",
    "source_name",
    "extraction_source_file",
    "extraction_source_file_hash",
    "extraction_source_sheet",
    "extraction_source_row",
    "raw_source_code",
    "raw_name",
    "raw_spec_model",
    "raw_unit",
    "raw_labor_fee",
    "raw_material_fee",
    "raw_machine_fee",
    "raw_management_fee",
    "raw_total_fee",
    "parse_issue",
    "remark",
]

CANDIDATE_FIELDS = [
    "reference_id",
    "source_type",
    "source_name",
    "extraction_source_file",
    "extraction_source_file_hash",
    "extraction_source_sheet",
    "extraction_source_row",
    "official_source_file",
    "official_source_page",
    "official_source_page_evidence_type",
    "chapter_code",
    "chapter_name",
    "section_code",
    "section_name",
    "item_group_name",
    "source_code",
    "raw_name",
    "standard_name_candidate",
    "unit",
    "keywords",
    "aliases",
    "feature_template",
    "extraction_confidence",
    "source_trust_level",
    "verification_status",
    "review_status",
    "reviewer",
    "remark",
]

PRICING_FIELDS = [
    "pricing_snapshot_id",
    "reference_id",
    "source_code",
    "raw_name",
    "unit",
    "labor_fee",
    "material_fee",
    "machine_fee",
    "management_fee",
    "total_fee",
    "pricing_source",
    "pricing_version",
    "source_trust_level",
    "review_status",
    "remark",
]

ISSUE_FIELDS = [
    "issue_id",
    "source_code",
    "raw_name",
    "issue_type",
    "issue_detail",
    "severity",
    "suggested_action",
]

SECTION_INVENTORY_FIELDS = [
    "section_guess",
    "source_code_prefix",
    "first_source_code",
    "last_source_code",
    "row_count",
    "sample_names",
    "confidence",
    "remark",
]


def text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def norm_header(value: Any) -> str:
    return text(value).replace("\ufeff", "")


def header_key(value: Any) -> str:
    return re.sub(r"\s+", "", norm_header(value)).lower()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalize_source_code(value: Any) -> str:
    raw = text(value).upper()
    if not raw:
        return ""
    raw = raw.translate(str.maketrans("０１２３４５６７８９Ａａ", "0123456789Aa"))
    raw = re.sub(r"[－—–‑‐﹣]+", "-", raw)
    raw = re.sub(r"\s*-\s*", "-", raw)
    raw = re.sub(r"\s+", "", raw)
    return raw


def is_a1_code(code: str) -> bool:
    return bool(re.fullmatch(r"A1-\d+(?:-\d+){0,2}", normalize_source_code(code)))


def source_code_prefix(code: str) -> str:
    parts = normalize_source_code(code).split("-")
    return "-".join(parts[:2]) if len(parts) >= 2 else ""


def natural_code_key(code: str) -> Tuple[int, int, int]:
    match = re.fullmatch(r"A1-(\d+)(?:-(\d+))?(?:-(\d+))?", normalize_source_code(code))
    if not match:
        return (10_000, 10_000, 10_000)
    return (int(match.group(1)), int(match.group(2) or 0), int(match.group(3) or 0))


def section_code_for(code: str) -> str:
    prefix = source_code_prefix(code)
    if not prefix:
        return ""
    return "A.1." + prefix.split("-")[1]


def section_name_for(prefix: str) -> str:
    known = {"A1-1": "土石方工程"}
    return known.get(prefix, f"{prefix} 分组（按 source_code prefix 识别）")


def is_supplemental_code(code: str) -> bool:
    return len(normalize_source_code(code).split("-")) >= 4


def light_name(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def item_group_from_name(name: str) -> str:
    cleaned = light_name(name)
    if not cleaned:
        return ""
    split_tokens = [" 一、二类", " 三类", " 四类", " 极软岩", " 软岩", " 较软岩", " 较硬岩", " 坚硬岩", " 每增", " 增运", " 运距"]
    for token in split_tokens:
        if token in cleaned:
            return cleaned.split(token, 1)[0]
    return cleaned


def keywords_from_name(name: str) -> str:
    tokens = [token for token in re.split(r"[\s、，；;（）()《》\"“”]+", name or "") if len(token) >= 2]
    result: List[str] = []
    seen = set()
    for token in tokens:
        if token not in seen:
            result.append(token)
            seen.add(token)
    return ";".join(result[:10])


def build_column_map(headers: Sequence[str]) -> Dict[str, int]:
    alias_lookup: Dict[str, str] = {}
    for normalized_name, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            alias_lookup[header_key(alias)] = normalized_name
    column_map: Dict[str, int] = {}
    for idx, header in enumerate(headers):
        normalized_name = alias_lookup.get(header_key(header))
        if normalized_name and normalized_name not in column_map:
            column_map[normalized_name] = idx
    missing = [field for field in ["source_code", "raw_name", "raw_unit"] if field not in column_map]
    if missing:
        raise SystemExit("blocked_missing_inputs: normalized Excel missing required fields " + ";".join(missing))
    return column_map


def mapped_value(row: Sequence[Any], column_map: Dict[str, int], field: str) -> Any:
    idx = column_map.get(field)
    if idx is None or idx >= len(row):
        return ""
    return row[idx]


def normalized_row(headers: Sequence[str], row: Sequence[Any], row_no: int, column_map: Dict[str, int]) -> Dict[str, Any]:
    data = {headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))}
    for field in FIELD_ALIASES:
        data[field] = mapped_value(row, column_map, field)
    data["_row_no"] = row_no
    data["_source_code"] = normalize_source_code(data.get("source_code"))
    return data


def collect_a1_rows(ws: Any, headers: Sequence[str], column_map: Dict[str, int]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    rows: List[Dict[str, Any]] = []
    parse_issues: List[Dict[str, Any]] = []
    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        data = normalized_row(headers, row, row_no, column_map)
        code = text(data.get("_source_code"))
        raw_name = text(data.get("raw_name"))
        if not code:
            continue
        if is_a1_code(code):
            if "小计" in raw_name or "说明" in raw_name:
                continue
            rows.append(data)
        elif code.startswith("A1-"):
            parse_issues.append(data)
    rows.sort(key=lambda row: natural_code_key(row["_source_code"]))
    return rows, parse_issues


def build_raw_rows(excel_path: Path, excel_hash: str, sheet_name: str, rows: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    output = []
    for idx, row in enumerate(rows, start=1):
        code = text(row.get("_source_code"))
        output.append(
            {
                "raw_row_id": f"GD2018_A1_RAW_{idx:04d}",
                "source_type": SOURCE_TYPE,
                "source_name": SOURCE_NAME,
                "extraction_source_file": excel_path.name,
                "extraction_source_file_hash": excel_hash,
                "extraction_source_sheet": sheet_name,
                "extraction_source_row": row["_row_no"],
                "raw_source_code": code,
                "raw_name": text(row.get("raw_name")),
                "raw_spec_model": text(row.get("raw_spec_model")),
                "raw_unit": text(row.get("raw_unit")),
                "raw_labor_fee": row.get("raw_labor_fee", ""),
                "raw_material_fee": row.get("raw_material_fee", ""),
                "raw_machine_fee": row.get("raw_machine_fee", ""),
                "raw_management_fee": row.get("raw_management_fee", ""),
                "raw_total_fee": row.get("raw_total_fee", ""),
                "parse_issue": "supplemental_source_code" if is_supplemental_code(code) else "",
                "remark": "normalized_excel_raw_reference_row;not_official_truth",
            }
        )
    return output


def build_candidates(excel_path: Path, excel_hash: str, sheet_name: str, rows: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    candidates = []
    for row in rows:
        code = text(row.get("_source_code"))
        prefix = source_code_prefix(code)
        raw_name = text(row.get("raw_name"))
        candidates.append(
            {
                "reference_id": f"GD2018_A1_EXCEL_{code}",
                "source_type": SOURCE_TYPE,
                "source_name": SOURCE_NAME,
                "extraction_source_file": excel_path.name,
                "extraction_source_file_hash": excel_hash,
                "extraction_source_sheet": sheet_name,
                "extraction_source_row": row["_row_no"],
                "official_source_file": OFFICIAL_SOURCE_FILE,
                "official_source_page": "",
                "official_source_page_evidence_type": OFFICIAL_SOURCE_PAGE_EVIDENCE_TYPE,
                "chapter_code": CHAPTER_CODE,
                "chapter_name": CHAPTER_NAME,
                "section_code": section_code_for(code),
                "section_name": section_name_for(prefix),
                "item_group_name": item_group_from_name(raw_name),
                "source_code": code,
                "raw_name": raw_name,
                "standard_name_candidate": light_name(raw_name),
                "unit": text(row.get("raw_unit")),
                "keywords": keywords_from_name(raw_name),
                "aliases": "",
                "feature_template": "",
                "extraction_confidence": "0.84" if not is_supplemental_code(code) else "0.78",
                "source_trust_level": SOURCE_TRUST_LEVEL,
                "verification_status": VERIFICATION_STATUS,
                "review_status": REVIEW_STATUS,
                "reviewer": "",
                "remark": "normalized_excel_reference_candidate;official_pdf_page_pending",
            }
        )
    return candidates


def build_pricing_snapshot(rows: Sequence[Dict[str, Any]], candidates: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    by_code = {text(row.get("_source_code")): row for row in rows}
    output = []
    for candidate in candidates:
        row = by_code[candidate["source_code"]]
        output.append(
            {
                "pricing_snapshot_id": f"GD2018_A1_PRICE_{candidate['source_code']}",
                "reference_id": candidate["reference_id"],
                "source_code": candidate["source_code"],
                "raw_name": candidate["raw_name"],
                "unit": candidate["unit"],
                "labor_fee": row.get("raw_labor_fee", ""),
                "material_fee": row.get("raw_material_fee", ""),
                "machine_fee": row.get("raw_machine_fee", ""),
                "management_fee": row.get("raw_management_fee", ""),
                "total_fee": row.get("raw_total_fee", ""),
                "pricing_source": PRICING_SOURCE,
                "pricing_version": PRICING_VERSION,
                "source_trust_level": SOURCE_TRUST_LEVEL,
                "review_status": REVIEW_STATUS,
                "remark": "reference_pricing_snapshot_only;do_not_generate_internal_price_library",
            }
        )
    return output


def build_issues(rows: Sequence[Dict[str, Any]], parse_issues: Sequence[Dict[str, Any]], candidates: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    issues: List[Dict[str, Any]] = []

    def add(code: str, raw_name: str, issue_type: str, detail: str, severity: str, action: str) -> None:
        issues.append(
            {
                "issue_id": f"ISSUE_GD2018_A1_{len(issues) + 1:04d}",
                "source_code": code,
                "raw_name": raw_name,
                "issue_type": issue_type,
                "issue_detail": detail,
                "severity": severity,
                "suggested_action": action,
            }
        )

    code_counts = Counter(row["source_code"] for row in candidates)
    for row in candidates:
        code = row["source_code"]
        raw_name = row["raw_name"]
        if not row["unit"]:
            add(code, raw_name, "missing_unit", "A1 candidate row has no unit.", "high", "Verify against source Excel.")
        if not raw_name:
            add(code, raw_name, "missing_name", "A1 candidate row has no raw_name.", "high", "Verify against source Excel.")
        if code_counts[code] > 1:
            add(code, raw_name, "duplicated_source_code", "source_code appears more than once.", "high", "Deduplicate before mapping.")
    for row in parse_issues:
        add(text(row.get("_source_code")), text(row.get("raw_name")), "source_code_parse_issue", "Row starts with A1 but does not match accepted source_code format.", "high", "Exclude until classified.")
    add("", "", "non_official_source", "The normalized workbook is derived from third-party Excel and is not official truth.", "medium", "Keep candidates pending and require human confirmation.")
    add("", "", "official_pdf_page_pending", "official_source_page remains pending for A1 candidate rows.", "medium", "Verify source pages before stronger trust status.")
    return issues


def build_section_inventory(candidates: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in candidates:
        grouped[source_code_prefix(row["source_code"])].append(row)
    inventory = []
    for prefix in sorted(grouped, key=lambda value: natural_code_key(value + "-0")):
        rows = sorted(grouped[prefix], key=lambda row: natural_code_key(row["source_code"]))
        samples = [row["raw_name"] for row in rows[:5]]
        inventory.append(
            {
                "section_guess": section_name_for(prefix),
                "source_code_prefix": prefix,
                "first_source_code": rows[0]["source_code"],
                "last_source_code": rows[-1]["source_code"],
                "row_count": len(rows),
                "sample_names": "; ".join(samples),
                "confidence": "0.90" if prefix == "A1-1" else "0.65",
                "remark": "section name confirmed for A1-1" if prefix == "A1-1" else "grouped by source_code prefix; section name requires human confirmation",
            }
        )
    return inventory


def write_csv(path: Path, fields: Sequence[str], rows: Sequence[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def write_report(path: Path, excel_path: Path, rows: Sequence[Dict[str, Any]], candidates: Sequence[Dict[str, Any]], pricing: Sequence[Dict[str, Any]], issues: Sequence[Dict[str, Any]], inventory: Sequence[Dict[str, Any]]) -> None:
    prefix_counts = Counter(source_code_prefix(row["source_code"]) for row in candidates)
    lines = [
        "# Stage GD2018-A1-FULL Report",
        "",
        "## 1. Task Scope",
        "",
        "Extract all GD2018 A1 quota reference candidates from the normalized Excel. This is a reference-candidate extraction only; it does not generate bill_code, approved records, internal_price_library, database rows, or enterprise templates.",
        "",
        "## 2. Input File",
        "",
        f"- normalized_excel: `{excel_path}`",
        f"- sha256: `{sha256_file(excel_path)}`",
        "",
        "## 3. Extraction Summary",
        "",
        f"- raw_reference_rows: {len(rows)}",
        f"- standard_cost_item_reference_A1_candidate_rows: {len(candidates)}",
        f"- pricing_snapshot_rows: {len(pricing)}",
        f"- issue_rows: {len(issues)}",
        f"- source_code_prefix_count: {len(prefix_counts)}",
        "",
        "## 4. Section Inventory",
        "",
        "| Prefix | Rows | First | Last | Section Guess |",
        "|---|---:|---|---|---|",
    ]
    for item in inventory:
        lines.append(f"| {item['source_code_prefix']} | {item['row_count']} | {item['first_source_code']} | {item['last_source_code']} | {item['section_guess']} |")
    lines.extend(
        [
            "",
            "## 5. Quality Checks",
            "",
            f"- invalid_parse_issue_rows: {sum(1 for issue in issues if issue['issue_type'] == 'source_code_parse_issue')}",
            f"- missing_unit_issues: {sum(1 for issue in issues if issue['issue_type'] == 'missing_unit')}",
            f"- missing_name_issues: {sum(1 for issue in issues if issue['issue_type'] == 'missing_name')}",
            f"- duplicate_source_code_issues: {sum(1 for issue in issues if issue['issue_type'] == 'duplicated_source_code')}",
            "- review_status: pending for all generated candidates and pricing snapshots",
            "",
            "## 6. Not Final Statement",
            "",
            "All rows are third-party normalized Excel reference candidates. They are not approved, not final enterprise standard names, and not bill-code mappings.",
            "",
            f"Generated at: {datetime.now().isoformat(timespec='seconds')}",
        ]
    )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Extract GD2018 A1 full reference candidates from normalized Excel.")
    parser.add_argument("--project-root", type=Path, default=PROJECT_ROOT_DEFAULT)
    parser.add_argument("--excel", type=Path, default=None)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    project_root = args.project_root
    excel_path = args.excel if args.excel else project_root / EXCEL_REL
    output_dir = project_root / OUTPUT_DIR_REL
    output_dir.mkdir(parents=True, exist_ok=True)

    if not excel_path.exists():
        raise SystemExit(f"blocked_missing_inputs: Excel file not found: {excel_path}")

    workbook = load_workbook(excel_path, read_only=False, data_only=False)
    visible_sheets = [ws for ws in workbook.worksheets if ws.sheet_state == "visible"]
    if len(visible_sheets) != 1:
        raise SystemExit("blocked_missing_inputs: unable to identify a single normalized data sheet")
    ws = visible_sheets[0]
    headers = [norm_header(cell.value) for cell in ws[1]]
    column_map = build_column_map(headers)
    source_hash = sha256_file(excel_path)
    rows, parse_issues = collect_a1_rows(ws, headers, column_map)
    raw_rows = build_raw_rows(excel_path, source_hash, ws.title, rows)
    candidates = build_candidates(excel_path, source_hash, ws.title, rows)
    pricing = build_pricing_snapshot(rows, candidates)
    issues = build_issues(rows, parse_issues, candidates)
    inventory = build_section_inventory(candidates)

    write_csv(output_dir / "raw_reference_excel_rows_A1.csv", RAW_FIELDS, raw_rows)
    write_csv(output_dir / "standard_cost_item_reference_A1_candidate.csv", CANDIDATE_FIELDS, candidates)
    write_csv(output_dir / "reference_quota_pricing_snapshot_A1.csv", PRICING_FIELDS, pricing)
    write_csv(output_dir / "gd2018_a1_extraction_issues.csv", ISSUE_FIELDS, issues)
    write_csv(output_dir / "gd2018_a1_section_inventory.csv", SECTION_INVENTORY_FIELDS, inventory)
    write_report(output_dir / "stage_gd2018_a1_full_report.md", excel_path, raw_rows, candidates, pricing, issues, inventory)

    print(f"raw_reference_rows={len(raw_rows)}")
    print(f"candidate_rows={len(candidates)}")
    print(f"pricing_snapshot_rows={len(pricing)}")
    print(f"section_inventory_rows={len(inventory)}")
    print(f"issue_rows={len(issues)}")
    print("prefix_counts=" + json.dumps(dict(Counter(source_code_prefix(row["source_code"]) for row in candidates)), ensure_ascii=False, sort_keys=True))
    print(f"output_dir={output_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
