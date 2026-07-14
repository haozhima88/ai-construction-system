#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Stage MAP-APPENDIX-A-TSF-PILOT-1.

Build a small human-review mapping pilot between GB/T 50854 Appendix A
earthwork bill items and GD2018 A1-1 earthwork quota rows. This script only
creates review artifacts; it does not write databases, approvals, enterprise
standard names, internal price library rows, templates, or source-baseline
updates.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


PROJECT_ROOT_DEFAULT = Path(r"E:\workspace\01_Projects\ai-construction-system")
ENGINE_REL = Path("construction_cost_knowledge_engine")
RUNS_REL = ENGINE_REL / "data" / "private" / "reference_extraction" / "runs"
BASELINE_REL = RUNS_REL / "SOURCE_BASELINE_LOCK_1"
GB_BASE_REL = BASELINE_REL / "GB50854_2024_full_standard_parse_review"
GD_BASE_REL = BASELINE_REL / "GD2018_normalized_full_quota_parse_review"
OUTPUT_DIR_REL = RUNS_REL / "MAP_APPENDIX_A_TSF_PILOT_1"
DOCS_REF_REL = ENGINE_REL / "docs" / "reference_extraction"

GB_BILLS_REL = GB_BASE_REL / "gb50854_bill_items_full_review.csv"
GB_RULES_REL = GB_BASE_REL / "gb50854_context_rules_full_review.csv"
GD_QUOTA_REL = GD_BASE_REL / "gd2018_normalized_quota_items_full_review.csv"
GD_PRICING_REL = GD_BASE_REL / "gd2018_normalized_pricing_fields_full_review.csv"

STAGE_NAME = "MAP_APPENDIX_A_TSF_PILOT_1"
REVIEW_STATUS = "pending"
SOURCE_TYPE = "lightweight_rule_mapping_candidate"
MAPPING_SCOPE = "GB50854_appendix_A_to_GD2018_A1_1_tsf_pilot"

APPENDIX_A_CODES = [
    "010101001",
    "010101002",
    "010101003",
    "010102001",
    "010102002",
    "010102003",
    "010102004",
    "010102005",
    "010102006",
    "010102007",
    "010103001",
    "010103002",
]

GB_REVIEW_FIELDS = [
    "bill_reference_id",
    "source_file_hash",
    "source_heading_path",
    "appendix_code",
    "appendix_name",
    "section_code",
    "section_name",
    "table_code",
    "table_name",
    "bill_code_9",
    "bill_name",
    "project_feature_raw",
    "unit",
    "quantity_calculation_rule",
    "work_content_raw",
    "related_context_rules",
    "review_status",
    "remark",
]

GD_UNIT_FIELDS = [
    "reference_id",
    "source_file_hash",
    "source_sheet",
    "source_excel_row",
    "source_code",
    "raw_name",
    "quota_name_candidate",
    "quota_feature_text_candidate",
    "raw_unit",
    "unit_normalized",
    "unit_factor_to_normalized",
    "unit_dimension",
    "unit_normalization_status",
    "raw_spec_model",
    "raw_quantity",
    "raw_labor_fee",
    "raw_material_fee",
    "raw_machine_fee",
    "raw_management_fee",
    "raw_total_fee",
    "review_status",
    "remark",
]

UNIT_DICT_FIELDS = [
    "raw_unit_pattern",
    "unit_normalized",
    "unit_factor_to_normalized",
    "unit_dimension",
    "parse_rule",
    "example_raw_unit",
    "remark",
]

MATRIX_FIELDS = [
    "bill_code_9",
    "bill_name",
    "bill_section_code",
    "bill_section_name",
    "bill_unit",
    "bill_quantity_calculation_rule",
    "bill_work_content_raw",
    "related_context_rules",
    "total_quota_candidate_count",
    "direct_candidate_count",
    "feature_required_count",
    "work_content_component_count",
    "construction_method_only_count",
    "transport_or_disposal_count",
    "no_direct_bill_item_count",
    "manual_review_required_count",
    "top_quota_source_codes",
    "top_quota_names",
    "top_quota_units_raw",
    "top_quota_units_normalized",
    "top_mapping_basis",
    "coverage_status",
    "human_check_priority",
    "human_decision",
    "human_comment",
]

DETAIL_FIELDS = [
    "bill_code_9",
    "bill_name",
    "bill_unit",
    "bill_quantity_calculation_rule",
    "bill_work_content_raw",
    "quota_source_code",
    "quota_raw_name",
    "quota_name_candidate",
    "quota_feature_text_candidate",
    "quota_raw_unit",
    "quota_unit_normalized",
    "quota_unit_factor_to_normalized",
    "quota_unit_dimension",
    "unit_compatibility_status",
    "mapping_status",
    "mapping_type",
    "mapping_basis",
    "mapping_confidence",
    "routing_status",
    "issue_types",
    "review_status",
    "human_decision",
    "human_comment",
]

ISSUE_FIELDS = [
    "issue_id",
    "issue_type",
    "bill_code_9",
    "bill_name",
    "quota_source_code",
    "quota_raw_name",
    "issue_detail",
    "severity",
    "suggested_action",
]

MANIFEST_FIELDS = [
    "stage_name",
    "artifact_name",
    "expected_path",
    "exists",
    "file_size_bytes",
    "row_count",
    "sha256",
    "created_or_modified_time",
    "source_file",
    "can_regenerate",
    "backup_required",
    "backup_path",
    "status",
    "remark",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--project-root", type=Path, default=PROJECT_ROOT_DEFAULT)
    return parser.parse_args()


def read_csv(path: Path) -> List[Dict[str, str]]:
    if not path.exists():
        raise SystemExit(f"blocked_missing_inputs: {path}")
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, fields: Sequence[str], rows: Iterable[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(fields), extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def rel(path: Path, root: Path) -> str:
    return path.relative_to(root).as_posix()


def csv_row_count(path: Path) -> int:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return sum(1 for _ in csv.DictReader(handle))


def artifact_row_count(path: Path) -> str:
    if not path.exists():
        return ""
    if path.suffix.lower() == ".csv":
        return str(csv_row_count(path))
    if path.suffix.lower() == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        return str(sum(max(0, ws.max_row - 1) for ws in workbook.worksheets))
    return ""


def norm(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", "\n")
    text = re.sub(r"[ \t\u3000]+", " ", text)
    return text.strip()


def compact(value: Any) -> str:
    return re.sub(r"\s+", "", norm(value)).lower()


def natural_code_key(value: str) -> Tuple[Any, ...]:
    parts = re.split(r"(\d+)", norm(value).upper())
    key: List[Any] = []
    for part in parts:
        key.append(int(part) if part.isdigit() else part)
    return tuple(key)


def normalize_unit(raw_unit: str) -> Dict[str, str]:
    raw = norm(raw_unit)
    text = compact(raw)
    text = text.replace("m³", "m3").replace("ｍ³", "m3").replace("立方米", "m3")
    text = text.replace("m²", "m2").replace("㎡", "m2").replace("平方米", "m2")
    text = text.replace("米", "m").replace("吨", "t")
    text = text.replace("m３", "m3").replace("m２", "m2")

    m = re.fullmatch(r"(\d+(?:\.\d+)?)?m3", text)
    if m:
        factor = m.group(1) or "1"
        return {"unit_normalized": "m3", "factor": factor, "dimension": "volume", "status": "parsed"}
    m = re.fullmatch(r"(\d+(?:\.\d+)?)?m2", text)
    if m:
        factor = m.group(1) or "1"
        return {"unit_normalized": "m2", "factor": factor, "dimension": "area", "status": "parsed"}
    m = re.fullmatch(r"(\d+(?:\.\d+)?)?m", text)
    if m:
        factor = m.group(1) or "1"
        return {"unit_normalized": "m", "factor": factor, "dimension": "length", "status": "parsed"}
    if text == "t":
        return {"unit_normalized": "t", "factor": "1", "dimension": "weight", "status": "parsed"}
    if text == "台班":
        return {"unit_normalized": "台班", "factor": "1", "dimension": "machine_shift", "status": "parsed"}
    if text == "项":
        return {"unit_normalized": "项", "factor": "1", "dimension": "lump_sum", "status": "parsed"}
    return {"unit_normalized": "", "factor": "", "dimension": "", "status": "unparsed"}


def unit_compatibility(bill_unit: str, quota_unit: Dict[str, str]) -> str:
    bill = normalize_unit(bill_unit)
    if bill["status"] == "unparsed" or quota_unit["status"] == "unparsed":
        return "unparsed"
    if bill["dimension"] != quota_unit["dimension"]:
        return "dimension_mismatch"
    if bill["unit_normalized"] == quota_unit["unit_normalized"] and str(bill["factor"]) == str(quota_unit["factor"]):
        return "compatible"
    return "convertible"


def build_unit_dictionary(raw_units: Sequence[str]) -> List[Dict[str, str]]:
    examples = {unit: normalize_unit(unit) for unit in sorted(set(raw_units), key=str)}
    base = [
        ("m3|m³|立方米", "m3", "1", "volume", "exact_volume_unit", "m³"),
        ("N+m3|N+m³|N+立方米", "m3", "N", "volume", "numeric_prefix_volume_unit", "100m3"),
        ("m2|m²|㎡|平方米", "m2", "1", "area", "exact_area_unit", "m²"),
        ("N+m2|N+m²|N+平方米", "m2", "N", "area", "numeric_prefix_area_unit", "100m2"),
        ("m|米", "m", "1", "length", "exact_length_unit", "m"),
        ("t|吨", "t", "1", "weight", "exact_weight_unit", "t"),
        ("台班", "台班", "1", "machine_shift", "exact_machine_shift_unit", "台班"),
        ("项", "项", "1", "lump_sum", "exact_lump_sum_unit", "项"),
    ]
    rows = [
        {
            "raw_unit_pattern": pattern,
            "unit_normalized": unit,
            "unit_factor_to_normalized": factor,
            "unit_dimension": dimension,
            "parse_rule": rule,
            "example_raw_unit": example,
            "remark": "configured rule for Appendix A pilot unit normalization",
        }
        for pattern, unit, factor, dimension, rule, example in base
    ]
    for raw, parsed in examples.items():
        rows.append(
            {
                "raw_unit_pattern": raw,
                "unit_normalized": parsed["unit_normalized"],
                "unit_factor_to_normalized": parsed["factor"],
                "unit_dimension": parsed["dimension"],
                "parse_rule": "observed_raw_unit_parse",
                "example_raw_unit": raw,
                "remark": f"observed in A1-1 baseline; status={parsed['status']}",
            }
        )
    return rows


def related_context_rules_for_bill(bill: Dict[str, str], rules: Sequence[Dict[str, str]]) -> str:
    code = bill.get("bill_code_9", "")
    appendix = bill.get("appendix_code", "")
    related: List[str] = []
    for rule in rules:
        if rule.get("appendix_code") != appendix:
            continue
        related_codes = rule.get("related_bill_codes", "")
        if related_codes and code not in related_codes.split(";"):
            continue
        text = norm(rule.get("rule_text", ""))
        if text:
            related.append(f"{rule.get('rule_code')}: {text}")
    return "\n".join(related)


def build_gb_appendix_a_rows(bills: Sequence[Dict[str, str]], rules: Sequence[Dict[str, str]]) -> List[Dict[str, Any]]:
    by_code = {row["bill_code_9"]: row for row in bills if row.get("bill_code_9") in APPENDIX_A_CODES}
    output: List[Dict[str, Any]] = []
    for code in APPENDIX_A_CODES:
        bill = by_code.get(code)
        if not bill:
            continue
        output.append(
            {
                "bill_reference_id": bill.get("bill_reference_id", ""),
                "source_file_hash": bill.get("source_file_hash", ""),
                "source_heading_path": bill.get("source_heading_path", ""),
                "appendix_code": bill.get("appendix_code", ""),
                "appendix_name": bill.get("appendix_name", ""),
                "section_code": bill.get("section_code", ""),
                "section_name": bill.get("section_name", ""),
                "table_code": bill.get("table_code", ""),
                "table_name": bill.get("table_name", ""),
                "bill_code_9": bill.get("bill_code_9", ""),
                "bill_name": bill.get("bill_name", ""),
                "project_feature_raw": bill.get("project_feature_raw", ""),
                "unit": bill.get("unit", ""),
                "quantity_calculation_rule": bill.get("quantity_calculation_rule", ""),
                "work_content_raw": bill.get("work_content_raw", ""),
                "related_context_rules": related_context_rules_for_bill(bill, rules),
                "review_status": REVIEW_STATUS,
                "remark": "Appendix A pilot bill baseline; source baseline unchanged; pending human review",
            }
        )
    return output


def build_gd_a111_rows(quotas: Sequence[Dict[str, str]], pricing_by_code: Dict[str, Dict[str, str]]) -> List[Dict[str, Any]]:
    rows = [row for row in quotas if row.get("source_code", "").startswith("A1-1-")]
    rows.sort(key=lambda row: natural_code_key(row.get("source_code", "")))
    output: List[Dict[str, Any]] = []
    for quota in rows:
        raw_unit = quota.get("unit", "")
        parsed = normalize_unit(raw_unit)
        pricing = pricing_by_code.get(quota.get("source_code", ""), {})
        output.append(
            {
                "reference_id": quota.get("reference_id", ""),
                "source_file_hash": quota.get("source_file_hash", ""),
                "source_sheet": quota.get("source_sheet", ""),
                "source_excel_row": quota.get("source_excel_row", ""),
                "source_code": quota.get("source_code", ""),
                "raw_name": quota.get("raw_name", ""),
                "quota_name_candidate": quota.get("quota_name_candidate", ""),
                "quota_feature_text_candidate": quota.get("quota_feature_text_candidate", ""),
                "raw_unit": raw_unit,
                "unit_normalized": parsed["unit_normalized"],
                "unit_factor_to_normalized": parsed["factor"],
                "unit_dimension": parsed["dimension"],
                "unit_normalization_status": parsed["status"],
                "raw_spec_model": quota.get("raw_spec_model", ""),
                "raw_quantity": quota.get("raw_quantity", ""),
                "raw_labor_fee": pricing.get("raw_labor_fee", quota.get("raw_labor_fee", "")),
                "raw_material_fee": pricing.get("raw_material_fee", quota.get("raw_material_fee", "")),
                "raw_machine_fee": pricing.get("raw_machine_fee", quota.get("raw_machine_fee", "")),
                "raw_management_fee": pricing.get("raw_management_fee", quota.get("raw_management_fee", "")),
                "raw_total_fee": pricing.get("raw_total_fee", quota.get("raw_total_fee", "")),
                "review_status": REVIEW_STATUS,
                "remark": "A1-1 pilot quota baseline with derived unit normalization; raw_unit unchanged; no price conversion",
            }
        )
    return output


def contains_any(text: str, terms: Sequence[str]) -> bool:
    compact_text = compact(text)
    return any(compact(term) in compact_text for term in terms)


def quota_candidate_pairs(quota: Dict[str, Any]) -> List[Dict[str, Any]]:
    name = quota.get("quota_name_candidate") or quota.get("raw_name", "")
    pairs: List[Dict[str, Any]] = []

    def add(code: str, status: str, mapping_type: str, routing: str, basis: str, confidence: float) -> None:
        pairs.append(
            {
                "bill_code_9": code,
                "mapping_status": status,
                "mapping_type": mapping_type,
                "routing_status": routing,
                "mapping_basis": basis,
                "mapping_confidence": f"{confidence:.2f}",
            }
        )

    text = compact(name)
    is_transport = contains_any(name, ["运", "运输", "自卸汽车", "人力车", "铲运", "转堆", "垂直运输", "每增加", "每增", "装车"])
    is_loading_only = contains_any(name, ["人工装车", "装载机装", "挖掘机装", "人工装石方", "人工装土方"])
    is_method_only = contains_any(name, ["原土打夯", "夯实", "碾压", "支密板", "支疏板", "挡土板"])

    if "平整场地" in text:
        add("010103001", "feature_required", "feature_required", "routed_to_bill_item", "bill_name_overlap:平整场地;unit_convertible;feature土石类别_required", 0.86)

    if contains_any(name, ["回填"]):
        add("010102007", "feature_required", "feature_required", "routed_to_bill_item", "bill_name_overlap:回填方;填方部位/密实度_features_required", 0.82)
        add("010101003", "feature_required", "feature_required", "routed_to_bill_item", "单独土石方回填_possible_scope;needs_scope_confirmation", 0.62)

    if contains_any(name, ["淤泥", "流砂"]) and contains_any(name, ["挖"]):
        add("010102004", "feature_required", "feature_required", "routed_to_bill_item", "object_overlap:淤泥/流砂;开挖深度_feature_required", 0.82)

    if contains_any(name, ["基坑土方"]):
        add("010102001", "feature_required", "feature_required", "routed_to_bill_item", "object_overlap:基坑土方;土类别/深度_features_required", 0.84)
    if contains_any(name, ["沟槽土方"]):
        add("010102002", "feature_required", "feature_required", "routed_to_bill_item", "object_overlap:沟槽土方;土类别/深度_features_required", 0.84)
    if contains_any(name, ["沟槽、基坑土方", "槽、坑土方"]):
        add("010102001", "feature_required", "one_quota_to_multi_bill", "routed_to_bill_item", "quota combines沟槽/基坑;bill split required", 0.74)
        add("010102002", "feature_required", "one_quota_to_multi_bill", "routed_to_bill_item", "quota combines沟槽/基坑;bill split required", 0.74)

    if contains_any(name, ["一般土方"]) and contains_any(name, ["挖"]):
        add("010101001", "feature_required", "feature_required", "routed_to_bill_item", "object_overlap:一般土方;single/base scope must be confirmed", 0.76)

    if contains_any(name, ["一般石方", "松散石方"]) and contains_any(name, ["凿", "破碎", "爆破", "挖"]):
        add("010101002", "feature_required", "feature_required", "routed_to_bill_item", "object_overlap:一般/松散石方;岩石类别_feature_required", 0.78)
    if contains_any(name, ["槽、坑石方"]):
        add("010102005", "feature_required", "one_quota_to_multi_bill", "routed_to_bill_item", "quota combines槽/坑石方;bill split required", 0.74)
        add("010102006", "feature_required", "one_quota_to_multi_bill", "routed_to_bill_item", "quota combines槽/坑石方;bill split required", 0.74)

    if is_transport or is_loading_only:
        add("010103002", "transport_or_disposal_related", "transport_or_disposal", "routed_to_transport_or_disposal", "transport/loading/disposal risk;not_direct_bill_body", 0.55)

    if is_method_only and not contains_any(name, ["回填"]):
        add("010103001" if "100m2" in quota.get("raw_unit", "") else "010102007", "construction_method_only", "construction_method_only", "routed_to_method_or_measure", "construction_method_or_temporary_support_only;not_direct_bill_body", 0.45)

    # Deduplicate by bill/status/type while retaining the strongest confidence.
    dedup: Dict[Tuple[str, str, str], Dict[str, Any]] = {}
    for pair in pairs:
        key = (pair["bill_code_9"], pair["mapping_status"], pair["mapping_type"])
        existing = dedup.get(key)
        if not existing or float(pair["mapping_confidence"]) > float(existing["mapping_confidence"]):
            dedup[key] = pair
    return list(dedup.values())


def detail_issue_types(detail: Dict[str, Any], quota: Dict[str, Any]) -> str:
    issues: List[str] = []
    if detail["unit_compatibility_status"] == "dimension_mismatch":
        issues.append("unit_dimension_mismatch")
    if detail["unit_compatibility_status"] == "unparsed":
        issues.append("unit_unparsed")
    if detail["mapping_status"] == "feature_required":
        issues.append("feature_required")
    if detail["mapping_status"] == "transport_or_disposal_related":
        issues.append("transport_item_uncertain")
    if detail["mapping_status"] == "construction_method_only":
        issues.append("construction_method_only")
    if re.fullmatch(r"A1-1-\d+-\d+", quota.get("source_code", "")):
        issues.append("supplemental_quota_code")
    if detail["mapping_confidence"] and float(detail["mapping_confidence"]) < 0.6:
        issues.append("possible_wrong_mapping")
    return ";".join(issues)


def build_detail_rows(bills: Sequence[Dict[str, Any]], quotas: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    bill_by_code = {row["bill_code_9"]: row for row in bills}
    details: List[Dict[str, Any]] = []
    for quota in quotas:
        parsed_unit = {
            "unit_normalized": quota.get("unit_normalized", ""),
            "factor": quota.get("unit_factor_to_normalized", ""),
            "dimension": quota.get("unit_dimension", ""),
            "status": quota.get("unit_normalization_status", ""),
        }
        for pair in quota_candidate_pairs(quota):
            bill = bill_by_code.get(pair["bill_code_9"])
            if not bill:
                continue
            compatibility = unit_compatibility(bill.get("unit", ""), parsed_unit)
            confidence = float(pair["mapping_confidence"])
            basis = pair["mapping_basis"]
            if compatibility == "dimension_mismatch":
                confidence = max(0.1, confidence - 0.18)
                basis += ";unit_dimension_mismatch"
            elif compatibility == "convertible":
                basis += ";unit_convertible_by_factor"
            detail = {
                "bill_code_9": bill["bill_code_9"],
                "bill_name": bill["bill_name"],
                "bill_unit": bill["unit"],
                "bill_quantity_calculation_rule": bill["quantity_calculation_rule"],
                "bill_work_content_raw": bill["work_content_raw"],
                "quota_source_code": quota["source_code"],
                "quota_raw_name": quota["raw_name"],
                "quota_name_candidate": quota["quota_name_candidate"],
                "quota_feature_text_candidate": quota["quota_feature_text_candidate"],
                "quota_raw_unit": quota["raw_unit"],
                "quota_unit_normalized": quota["unit_normalized"],
                "quota_unit_factor_to_normalized": quota["unit_factor_to_normalized"],
                "quota_unit_dimension": quota["unit_dimension"],
                "unit_compatibility_status": compatibility,
                "mapping_status": pair["mapping_status"],
                "mapping_type": pair["mapping_type"],
                "mapping_basis": basis,
                "mapping_confidence": f"{confidence:.2f}",
                "routing_status": pair["routing_status"],
                "review_status": REVIEW_STATUS,
                "human_decision": "",
                "human_comment": "",
            }
            detail["issue_types"] = detail_issue_types(detail, quota)
            details.append(detail)
    details.sort(key=lambda row: (APPENDIX_A_CODES.index(row["bill_code_9"]), natural_code_key(row["quota_source_code"]), row["mapping_status"]))
    return details


def top_values(rows: Sequence[Dict[str, Any]], key: str, limit: int = 8) -> str:
    values: List[str] = []
    seen = set()
    for row in rows:
        value = norm(row.get(key, ""))
        if value and value not in seen:
            values.append(value)
            seen.add(value)
        if len(values) >= limit:
            break
    return ";".join(values)


def matrix_status(counts: Counter, total: int) -> Tuple[str, str]:
    if total == 0:
        return "no_quota_candidate", "high"
    if counts.get("transport_or_disposal_related"):
        return "covered_transport_or_disposal", "high"
    if counts.get("feature_required"):
        return "covered_feature_required", "high"
    if counts.get("direct_bill_candidate"):
        return "covered_direct", "medium"
    if counts.get("bill_work_content_component"):
        return "covered_work_content_only", "medium"
    if counts.get("construction_method_only") or counts.get("manual_review_required"):
        return "needs_manual_review", "high"
    return "covered_weak", "medium"


def build_matrix_rows(bills: Sequence[Dict[str, Any]], details: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    detail_by_bill: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in details:
        detail_by_bill[row["bill_code_9"]].append(row)
    matrix: List[Dict[str, Any]] = []
    for bill in bills:
        rows = sorted(detail_by_bill.get(bill["bill_code_9"], []), key=lambda row: float(row.get("mapping_confidence", "0")), reverse=True)
        counts = Counter(row["mapping_status"] for row in rows)
        total = len(rows)
        status, priority = matrix_status(counts, total)
        matrix.append(
            {
                "bill_code_9": bill["bill_code_9"],
                "bill_name": bill["bill_name"],
                "bill_section_code": bill["section_code"],
                "bill_section_name": bill["section_name"],
                "bill_unit": bill["unit"],
                "bill_quantity_calculation_rule": bill["quantity_calculation_rule"],
                "bill_work_content_raw": bill["work_content_raw"],
                "related_context_rules": bill["related_context_rules"],
                "total_quota_candidate_count": total,
                "direct_candidate_count": counts.get("direct_bill_candidate", 0),
                "feature_required_count": counts.get("feature_required", 0),
                "work_content_component_count": counts.get("bill_work_content_component", 0),
                "construction_method_only_count": counts.get("construction_method_only", 0),
                "transport_or_disposal_count": counts.get("transport_or_disposal_related", 0),
                "no_direct_bill_item_count": counts.get("no_direct_bill_item", 0),
                "manual_review_required_count": counts.get("manual_review_required", 0),
                "top_quota_source_codes": top_values(rows, "quota_source_code"),
                "top_quota_names": top_values(rows, "quota_raw_name"),
                "top_quota_units_raw": top_values(rows, "quota_raw_unit"),
                "top_quota_units_normalized": top_values(rows, "quota_unit_normalized"),
                "top_mapping_basis": top_values(rows, "mapping_basis", limit=5),
                "coverage_status": status,
                "human_check_priority": priority,
                "human_decision": "",
                "human_comment": "",
            }
        )
    return matrix


def add_issue(issues: List[Dict[str, Any]], issue_type: str, bill: Optional[Dict[str, Any]], quota: Optional[Dict[str, Any]], detail: str, severity: str, action: str) -> None:
    issues.append(
        {
            "issue_id": f"ISSUE_APPENDIX_A_TSF_{len(issues) + 1:05d}",
            "issue_type": issue_type,
            "bill_code_9": bill.get("bill_code_9", "") if bill else "",
            "bill_name": bill.get("bill_name", "") if bill else "",
            "quota_source_code": quota.get("source_code", quota.get("quota_source_code", "")) if quota else "",
            "quota_raw_name": quota.get("raw_name", quota.get("quota_raw_name", "")) if quota else "",
            "issue_detail": detail,
            "severity": severity,
            "suggested_action": action,
        }
    )


def build_issues(bills: Sequence[Dict[str, Any]], quotas: Sequence[Dict[str, Any]], matrix: Sequence[Dict[str, Any]], details: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    issues: List[Dict[str, Any]] = []
    bill_by_code = {row["bill_code_9"]: row for row in bills}
    quota_by_code = {row["source_code"]: row for row in quotas}
    for bill in bills:
        if not bill.get("quantity_calculation_rule"):
            add_issue(issues, "missing_quantity_calculation_rule", bill, None, "GB/T bill item quantity calculation rule is blank.", "high", "Verify baseline before mapping review.")
        if not bill.get("work_content_raw"):
            add_issue(issues, "missing_work_content", bill, None, "GB/T bill item work content is blank.", "high", "Verify baseline before mapping review.")
        if bill.get("related_context_rules"):
            add_issue(issues, "context_rule_required", bill, None, "Appendix A context rules apply to this bill item and must be checked with candidate mapping.", "low", "Review A.4 context rules during human confirmation.")
    for quota in quotas:
        if quota.get("unit_normalization_status") == "unparsed":
            add_issue(issues, "unit_unparsed", None, quota, f"Could not parse raw_unit={quota.get('raw_unit')}.", "high", "Add dictionary rule before using unit compatibility.")
        if re.fullmatch(r"A1-1-\d+-\d+", quota.get("source_code", "")):
            add_issue(issues, "supplemental_quota_code", None, quota, "Supplemental quota code detected.", "medium", "Keep separate in human review.")
        if contains_any(quota.get("raw_name", ""), ["原土打夯", "支密板", "支疏板", "挡土板", "碾压"]):
            add_issue(issues, "construction_method_only", None, quota, "Quota appears to describe construction method, compaction, or temporary support rather than bill item body.", "medium", "Do not confirm as direct bill mapping without review.")
    for row in matrix:
        bill = bill_by_code.get(row["bill_code_9"], row)
        if int(row["total_quota_candidate_count"]) == 0:
            add_issue(issues, "no_quota_candidate", bill, None, "No GD2018 A1-1 quota candidate was routed to this bill item by lightweight rules.", "high", "Human review should decide if missing by scope or baseline gap.")
        elif int(row["total_quota_candidate_count"]) > 20:
            add_issue(issues, "multiple_candidate_quota", bill, None, f"{row['total_quota_candidate_count']} quota candidates routed to this bill item.", "medium", "Review feature grouping before any enterprise template draft.")
    for detail in details:
        bill = bill_by_code.get(detail["bill_code_9"], detail)
        quota = quota_by_code.get(detail["quota_source_code"], detail)
        for issue_type in [item for item in detail.get("issue_types", "").split(";") if item]:
            severity = "high" if issue_type in {"unit_dimension_mismatch", "transport_item_uncertain"} else "medium"
            add_issue(issues, issue_type, bill, quota, f"Detail row issue: {issue_type}; basis={detail.get('mapping_basis')}", severity, "Review before any mapping decision.")
    return issues


def write_checklist(path: Path) -> None:
    lines = [
        "# Appendix A 土石方 Mapping Pilot Human Review Checklist",
        "",
        "## 1. Scope",
        "",
        "- GB/T 50854 Appendix A bill items: 010101001 to 010103002.",
        "- GD2018 quota rows: source_code prefix A1-1.",
        "- This is a pilot review package only; no approved mapping is produced.",
        "",
        "## 2. Required Review Files",
        "",
        "- `gb_appendix_A_bill_12_review.csv`",
        "- `gd_a111_quota_143_unit_normalized.csv`",
        "- `bill_to_quota_matrix_appendix_A_12.csv`",
        "- `bill_to_quota_detail_appendix_A.csv`",
        "- `mapping_issues_appendix_A.csv`",
        "",
        "## 3. Bill Item Checks",
        "",
        "- Confirm bill code, name, feature, unit, quantity calculation rule, and work content against GB/T baseline.",
        "- Confirm A.4 context rules before deciding whether a quota is bill body or work content.",
        "",
        "## 4. Quota Item Checks",
        "",
        "- Confirm source_code, raw_name, raw_unit, and pricing fields against GD2018 normalized baseline.",
        "- Keep supplemental codes separate during review.",
        "",
        "## 5. Unit Normalization Checks",
        "",
        "- Verify raw_unit is unchanged.",
        "- Verify `100m2`, `100m3`, and `1000m3` conversion factors before comparing bill units.",
        "",
        "## 6. Mapping Logic Checks",
        "",
        "- Treat 土类、岩石类别、深度、运距、施工方法 as features unless cost department says otherwise.",
        "- 装车、运输、转堆、垂直运输 should not be direct bill-body mappings by default.",
        "- 支挡土板 and compaction rows should be reviewed as method/measure/work content.",
        "",
        "## 7. High-Risk Rows",
        "",
        "- transport_or_disposal_related",
        "- construction_method_only",
        "- unit_dimension_mismatch",
        "- one_quota_to_multi_bill",
        "- supplemental_quota_code",
        "",
        "## 8. Human Decision Levels",
        "",
        "- P0_not_mapping",
        "- P1_work_content_only",
        "- P2_candidate_feature_required",
        "- P3_enterprise_template_candidate",
        "- P4_confirmed_by_cost_department",
        "",
        "Current stage may suggest at most P3. P4 is not allowed in this pilot output.",
        "",
        "## 9. Pass / Fail Criteria",
        "",
        "- Pass if all 12 bill items and all 143 quota rows are retained.",
        "- Pass if unit normalization is auditable and raw units are unchanged.",
        "- Fail if any row is approved, any source baseline is modified, or bill_code is written back to quota rows.",
        "",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")


def worksheet_safe(value: Any) -> Any:
    if isinstance(value, str) and len(value) > 32767:
        return value[:32740] + "\n...[TRUNCATED_FOR_XLSX_CELL_LIMIT]"
    return value


def write_xlsx(path: Path, sheet_specs: Sequence[Tuple[str, Sequence[str], Sequence[Dict[str, Any]]]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for sheet_name, fields, rows in sheet_specs:
        ws = wb.create_sheet(sheet_name[:31])
        ws.append(list(fields))
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for row in rows:
            ws.append([worksheet_safe(row.get(field, "")) for field in fields])
        ws.freeze_panes = "A2"
        for idx, field in enumerate(fields, start=1):
            width = min(max(len(field) + 2, 12), 45)
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
        for row_cells in ws.iter_rows(min_row=2):
            for cell in row_cells:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(path)


def summary_rows(gb_rows: Sequence[Dict[str, Any]], quota_rows: Sequence[Dict[str, Any]], matrix: Sequence[Dict[str, Any]], details: Sequence[Dict[str, Any]], issues: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    parsed = sum(1 for row in quota_rows if row.get("unit_normalization_status") == "parsed")
    unparsed = sum(1 for row in quota_rows if row.get("unit_normalization_status") == "unparsed")
    return [
        {"metric_name": "gb_appendix_A_bill_rows", "metric_value": len(gb_rows), "remark": "expected 12"},
        {"metric_name": "gd_a111_quota_rows", "metric_value": len(quota_rows), "remark": "expected 143"},
        {"metric_name": "unit_parsed_count", "metric_value": parsed, "remark": ""},
        {"metric_name": "unit_unparsed_count", "metric_value": unparsed, "remark": ""},
        {"metric_name": "matrix_rows", "metric_value": len(matrix), "remark": "expected 12"},
        {"metric_name": "detail_rows", "metric_value": len(details), "remark": ""},
        {"metric_name": "issue_rows", "metric_value": len(issues), "remark": ""},
        {"metric_name": "approved_count", "metric_value": 0, "remark": "no approved generated"},
        {"metric_name": "recommendation", "metric_value": recommendation(gb_rows, quota_rows, matrix, issues), "remark": ""},
    ]


def recommendation(gb_rows: Sequence[Dict[str, Any]], quota_rows: Sequence[Dict[str, Any]], matrix: Sequence[Dict[str, Any]], issues: Sequence[Dict[str, Any]]) -> str:
    if len(gb_rows) != 12 or len(quota_rows) != 143 or len(matrix) != 12:
        return "appendix_A_pilot_partial_manual_intervention_required"
    if any(row.get("unit_normalization_status") == "unparsed" for row in quota_rows):
        return "appendix_A_pilot_partial_manual_intervention_required"
    return "appendix_A_pilot_ready_for_human_review"


def write_report(path: Path, gb_rows: Sequence[Dict[str, Any]], quota_rows: Sequence[Dict[str, Any]], unit_dict: Sequence[Dict[str, Any]], matrix: Sequence[Dict[str, Any]], details: Sequence[Dict[str, Any]], issues: Sequence[Dict[str, Any]]) -> str:
    parsed = sum(1 for row in quota_rows if row.get("unit_normalization_status") == "parsed")
    unparsed = sum(1 for row in quota_rows if row.get("unit_normalization_status") == "unparsed")
    rec = recommendation(gb_rows, quota_rows, matrix, issues)
    status_counts = Counter(row.get("mapping_status", "") for row in details)
    lines = [
        "# Stage MAP-APPENDIX-A-TSF-PILOT-1 Report",
        "",
        "## 1. Task Scope",
        "",
        "Build a small human-review mapping pilot between GB/T Appendix A earthwork bill items and GD2018 A1-1 quota rows.",
        "",
        "## 2. Input Baselines",
        "",
        f"- `{GB_BILLS_REL.as_posix()}`",
        f"- `{GB_RULES_REL.as_posix()}`",
        f"- `{GD_QUOTA_REL.as_posix()}`",
        f"- `{GD_PRICING_REL.as_posix()}`",
        "",
        "## 3. GB/T Appendix A Bill Baseline",
        "",
        f"- bill_rows: {len(gb_rows)}",
        f"- missing_quantity_calculation_rule: {sum(1 for row in gb_rows if not row.get('quantity_calculation_rule'))}",
        f"- missing_work_content: {sum(1 for row in gb_rows if not row.get('work_content_raw'))}",
        "",
        "## 4. GD2018 A1-1 Quota Baseline",
        "",
        f"- quota_rows: {len(quota_rows)}",
        f"- supplemental_quota_codes: {sum(1 for row in quota_rows if re.fullmatch(r'A1-1-\\d+-\\d+', row.get('source_code', '')))}",
        "",
        "## 5. Unit Normalization Result",
        "",
        f"- dictionary_rows: {len(unit_dict)}",
        f"- parsed: {parsed}",
        f"- unparsed: {unparsed}",
        "Raw units are preserved; normalized unit fields are derived only for this pilot review.",
        "- XLSX sheet name uses `gd_a111_quota_143_unit_norm` because Excel limits sheet names to 31 characters; the CSV artifact keeps the full `gd_a111_quota_143_unit_normalized.csv` name.",
        "",
        "## 6. Mapping Strategy",
        "",
        "The pilot uses lightweight rules: object keywords, unit dimension compatibility, bill work content, and project-feature fit. Transport, loading, compaction, vertical transport, transfer piles, and support-board rows are treated as high-risk review rows rather than direct mappings.",
        "",
        "## 7. 12 Bill Item Matrix Summary",
        "",
        f"- matrix_rows: {len(matrix)}",
        f"- bill_items_with_candidates: {sum(1 for row in matrix if int(row.get('total_quota_candidate_count', 0)) > 0)}",
        f"- bill_items_without_candidates: {sum(1 for row in matrix if int(row.get('total_quota_candidate_count', 0)) == 0)}",
        "",
        "## 8. Detail Mapping Summary",
        "",
        f"- detail_rows: {len(details)}",
        f"- mapping_status_counts: {json.dumps(dict(status_counts), ensure_ascii=False, sort_keys=True)}",
        "",
        "## 9. Issues and Risk Groups",
        "",
        f"- issue_rows: {len(issues)}",
        f"- issue_type_counts: {json.dumps(dict(Counter(row.get('issue_type', '') for row in issues)), ensure_ascii=False, sort_keys=True)}",
        "",
        "## 10. Human Review Guidance",
        "",
        "- Confirm context rules before deciding bill body versus work content.",
        "- Confirm unit factors before comparing 100m3/1000m3 quota pricing to m3 bill items.",
        "- Treat all feature_required rows as candidates only.",
        "",
        "## 11. Not Approved / Not Final Statement",
        "",
        "All rows remain pending. This stage does not approve mappings, does not write databases, does not write bill_code back to quota rows, and does not generate enterprise standard names or templates.",
        "",
        "## 12. Next Step Recommendation",
        "",
        rec,
        "",
        f"Generated at: {datetime.now().isoformat(timespec='seconds')}",
        "",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")
    return rec


def manifest_row(stage: str, artifact: str, path: Path, source_file: str, project_root: Path) -> Dict[str, str]:
    exists = path.exists()
    return {
        "stage_name": stage,
        "artifact_name": artifact,
        "expected_path": rel(path, project_root),
        "exists": str(exists).lower(),
        "file_size_bytes": str(path.stat().st_size) if exists else "",
        "row_count": artifact_row_count(path) if exists else "",
        "sha256": sha256_file(path) if exists else "",
        "created_or_modified_time": datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds") if exists else "",
        "source_file": source_file,
        "can_regenerate": "true",
        "backup_required": "true",
        "backup_path": "construction_cost_knowledge_engine/data/private/reference_extraction/backups/runs_backup_after_MAP_APPENDIX_A_TSF_PILOT_1",
        "status": "generated" if exists else "missing",
        "remark": "Appendix A TSF pilot review artifact; private; pending human review; no approved mapping",
    }


def update_manifest(project_root: Path, output_dir: Path, artifacts: Sequence[str]) -> None:
    manifest_path = project_root / DOCS_REF_REL / "reference_artifact_manifest.csv"
    existing = read_csv(manifest_path) if manifest_path.exists() else []
    source_file = ";".join([rel(project_root / GB_BILLS_REL, project_root), rel(project_root / GD_QUOTA_REL, project_root)])
    replacement = {
        (STAGE_NAME, artifact): manifest_row(STAGE_NAME, artifact, output_dir / artifact, source_file, project_root)
        for artifact in artifacts
    }
    filtered = [row for row in existing if (row.get("stage_name"), row.get("artifact_name")) not in replacement]
    filtered.extend(replacement.values())
    write_csv(manifest_path, MANIFEST_FIELDS, filtered)
    write_manifest_md(project_root, filtered)


def write_manifest_md(project_root: Path, rows: Sequence[Dict[str, str]]) -> None:
    latest = [row for row in rows if row.get("stage_name") == STAGE_NAME]
    registered = len(rows)
    existing = sum(1 for row in rows if row.get("exists") == "true")
    lines = [
        "# Reference Artifact Manifest",
        "",
        "## Governance",
        "",
        "- `construction_cost_knowledge_engine/data/private/` is intentionally not tracked by Git.",
        "- Every private artifact must be registered with `row_count` and `sha256` after generation.",
        "- Each completed stage must back up its `runs` output directory after validation.",
        "- Mock SQLite and seed CSV artifacts must not be used as source of truth.",
        "- Appendix A pilot outputs are pending review artifacts only and do not approve mappings.",
        "",
        "## Current Manifest Summary",
        "",
        f"- registered_artifacts: {registered}",
        f"- existing_artifacts: {existing}",
        f"- missing_artifacts: {registered - existing}",
        "",
        "## Manifest CSV",
        "",
        "`construction_cost_knowledge_engine/docs/reference_extraction/reference_artifact_manifest.csv`",
        "",
        "## Latest Appendix A TSF Pilot Outputs",
        "",
    ]
    for row in latest:
        lines.append(f"- `{row.get('expected_path')}` ({row.get('row_count') or 'n/a'} rows)")
    lines.extend(
        [
            "",
            "## Backup Requirement",
            "",
            "`construction_cost_knowledge_engine/data/private/reference_extraction/backups/runs_backup_after_MAP_APPENDIX_A_TSF_PILOT_1/`",
            "",
        ]
    )
    (project_root / DOCS_REF_REL / "REFERENCE_ARTIFACT_MANIFEST.md").write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    args = parse_args()
    project_root = args.project_root
    output_dir = project_root / OUTPUT_DIR_REL
    output_dir.mkdir(parents=True, exist_ok=True)
    gb_bills = read_csv(project_root / GB_BILLS_REL)
    gb_rules = read_csv(project_root / GB_RULES_REL)
    gd_quotas = read_csv(project_root / GD_QUOTA_REL)
    gd_pricing = read_csv(project_root / GD_PRICING_REL)
    pricing_by_code = {row.get("source_code", ""): row for row in gd_pricing}

    gb_rows = build_gb_appendix_a_rows(gb_bills, gb_rules)
    gd_rows = build_gd_a111_rows(gd_quotas, pricing_by_code)
    unit_dict = build_unit_dictionary([row.get("raw_unit", "") for row in gd_rows] + [row.get("unit", "") for row in gb_rows])
    details = build_detail_rows(gb_rows, gd_rows)
    matrix = build_matrix_rows(gb_rows, details)
    issues = build_issues(gb_rows, gd_rows, matrix, details)
    summary = summary_rows(gb_rows, gd_rows, matrix, details, issues)

    write_csv(output_dir / "gb_appendix_A_bill_12_review.csv", GB_REVIEW_FIELDS, gb_rows)
    write_csv(output_dir / "gd_a111_quota_143_unit_normalized.csv", GD_UNIT_FIELDS, gd_rows)
    write_csv(output_dir / "unit_normalization_dictionary_A111.csv", UNIT_DICT_FIELDS, unit_dict)
    write_csv(output_dir / "bill_to_quota_matrix_appendix_A_12.csv", MATRIX_FIELDS, matrix)
    write_csv(output_dir / "bill_to_quota_detail_appendix_A.csv", DETAIL_FIELDS, details)
    write_csv(output_dir / "mapping_issues_appendix_A.csv", ISSUE_FIELDS, issues)
    write_checklist(output_dir / "human_review_checklist_appendix_A.md")
    rec = write_report(output_dir / "stage_map_appendix_A_tsf_pilot_report.md", gb_rows, gd_rows, unit_dict, matrix, details, issues)
    try:
        write_xlsx(
            output_dir / "Appendix_A_TSF_mapping_pilot_review.xlsx",
            [
                ("gb_appendix_A_bill_12", GB_REVIEW_FIELDS, gb_rows),
                ("gd_a111_quota_143_unit_norm", GD_UNIT_FIELDS, gd_rows),
                ("unit_normalization_dictionary", UNIT_DICT_FIELDS, unit_dict),
                ("bill_to_quota_matrix_12", MATRIX_FIELDS, matrix),
                ("bill_to_quota_detail", DETAIL_FIELDS, details),
                ("mapping_issues", ISSUE_FIELDS, issues),
                ("summary", ["metric_name", "metric_value", "remark"], summary),
            ],
        )
    except Exception as exc:
        raise SystemExit(f"blocked_xlsx_generation_failed: {exc}") from exc

    artifacts = [
        "gb_appendix_A_bill_12_review.csv",
        "gd_a111_quota_143_unit_normalized.csv",
        "unit_normalization_dictionary_A111.csv",
        "bill_to_quota_matrix_appendix_A_12.csv",
        "bill_to_quota_detail_appendix_A.csv",
        "mapping_issues_appendix_A.csv",
        "human_review_checklist_appendix_A.md",
        "stage_map_appendix_A_tsf_pilot_report.md",
        "Appendix_A_TSF_mapping_pilot_review.xlsx",
    ]
    update_manifest(project_root, output_dir, artifacts)

    print(f"recommendation={rec}")
    print(f"gb_appendix_A_bill_rows={len(gb_rows)}")
    print(f"gd_a111_quota_rows={len(gd_rows)}")
    print(f"unit_parsed={sum(1 for row in gd_rows if row.get('unit_normalization_status') == 'parsed')}")
    print(f"unit_unparsed={sum(1 for row in gd_rows if row.get('unit_normalization_status') == 'unparsed')}")
    print(f"matrix_rows={len(matrix)}")
    print(f"detail_rows={len(details)}")
    print(f"issue_rows={len(issues)}")
    print(f"xlsx_exists={(output_dir / 'Appendix_A_TSF_mapping_pilot_review.xlsx').exists()}")
    print(f"output_dir={output_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
