#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Stage MAP-REPRESENTATIVE-GOVERNANCE-CALIBRATION-1.

Create a multi-appendix representative mapping governance calibration package.
Outputs are review artifacts only: no database writes, no approvals, no
enterprise standard names, no internal price library, and no bill_code write
back into quota references.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


PROJECT_ROOT_DEFAULT = Path(r"E:\workspace\01_Projects\ai-construction-system")
ENGINE_REL = Path("construction_cost_knowledge_engine")
RUNS_REL = ENGINE_REL / "data" / "private" / "reference_extraction" / "runs"
BASELINE_REL = RUNS_REL / "SOURCE_BASELINE_LOCK_1"
GB_BASE_REL = BASELINE_REL / "GB50854_2024_full_standard_parse_review"
GD_BASE_REL = BASELINE_REL / "GD2018_normalized_full_quota_parse_review"
OUTPUT_DIR_REL = RUNS_REL / "MAP_REPRESENTATIVE_GOVERNANCE_CALIBRATION_1"
DOCS_REF_REL = ENGINE_REL / "docs" / "reference_extraction"

GB_BILLS_REL = GB_BASE_REL / "gb50854_bill_items_full_review.csv"
GB_RULES_REL = GB_BASE_REL / "gb50854_context_rules_full_review.csv"
GD_QUOTA_REL = GD_BASE_REL / "gd2018_normalized_quota_items_full_review.csv"
GD_PRICING_REL = GD_BASE_REL / "gd2018_normalized_pricing_fields_full_review.csv"

STAGE_NAME = "MAP_REPRESENTATIVE_GOVERNANCE_CALIBRATION_1"
REVIEW_STATUS = "pending"

SCOPE_FIELDS = [
    "scope_id",
    "selection_bucket",
    "bill_reference_id",
    "bill_code_9",
    "bill_name",
    "appendix_code",
    "appendix_name",
    "section_code",
    "section_name",
    "unit",
    "quantity_calculation_rule",
    "work_content_raw",
    "project_feature_raw",
    "selection_reason",
    "expected_mapping_challenge",
    "human_scope_confirmed",
    "human_comment",
]

RATIONALE_FIELDS = [
    "selection_bucket",
    "target_count",
    "actual_count",
    "business_reason",
    "mapping_challenge_covered",
    "cost_department_review_value",
    "remark",
]

MATRIX_FIELDS = [
    "bill_code_9",
    "bill_name",
    "appendix_code",
    "appendix_name",
    "unit",
    "quantity_calculation_rule",
    "work_content_raw",
    "project_feature_raw",
    "total_quota_candidate_count",
    "R1_direct_bill_body_count",
    "R2_feature_variant_count",
    "R3_work_content_component_count",
    "R4_method_or_measure_component_count",
    "R5_not_applicable_or_unrouted_count",
    "top_quota_source_codes",
    "top_quota_names",
    "top_quota_units",
    "recommended_bill_level_decision",
    "template_readiness",
    "required_human_confirmation",
    "cost_engineer_decision",
    "cost_engineer_comment",
]

DETAIL_FIELDS = [
    "bill_code_9",
    "bill_name",
    "appendix_code",
    "appendix_name",
    "bill_unit",
    "bill_quantity_calculation_rule",
    "bill_work_content_raw",
    "bill_project_feature_raw",
    "quota_source_code",
    "quota_raw_name",
    "quota_name_candidate",
    "quota_feature_text_candidate",
    "quota_unit",
    "quota_raw_total_fee",
    "source_code_prefix",
    "governance_role",
    "relationship_basis",
    "mapping_confidence",
    "selection_condition",
    "forbidden_use",
    "issue_types",
    "review_status",
    "cost_engineer_decision",
    "cost_engineer_comment",
]

SHARED_FIELDS = [
    "quota_source_code",
    "quota_raw_name",
    "quota_name_candidate",
    "candidate_bill_codes",
    "candidate_bill_names",
    "candidate_count",
    "governance_role",
    "shared_component_type",
    "selection_condition",
    "allowed_use",
    "forbidden_use",
    "review_priority",
    "cost_engineer_decision",
    "cost_engineer_comment",
]

ISSUE_FIELDS = [
    "issue_id",
    "issue_type",
    "bill_code_9",
    "bill_name",
    "quota_source_code",
    "quota_raw_name",
    "governance_role",
    "issue_detail",
    "severity",
    "suggested_action",
]

REVIEW_FIELDS = [
    "bill_code_9",
    "bill_name",
    "appendix_name",
    "recommended_bill_level_decision",
    "template_readiness",
    "top_quota_source_codes",
    "top_quota_names",
    "main_risk",
    "cost_engineer_decision",
    "cost_engineer_selected_quota_codes",
    "cost_engineer_comment",
    "follow_up_required",
    "remark",
]

MANIFEST_FIELDS = [
    "stage_name",
    "artifact_name",
    "expected_path",
    "exists",
    "file_size_bytes",
    "row_count",
    "sha256",
    "created_or_modified_time",
    "source_file",
    "can_regenerate",
    "backup_required",
    "backup_path",
    "status",
    "remark",
]

ALL_TERMS = [
    "土方",
    "石方",
    "基坑",
    "沟槽",
    "回填",
    "平整场地",
    "余方弃置",
    "淤泥",
    "流砂",
    "地基",
    "换填",
    "强夯",
    "搅拌桩",
    "钢板桩",
    "锚杆",
    "喷射混凝土",
    "边坡",
    "桩",
    "预制桩",
    "灌注桩",
    "截桩",
    "砖",
    "砌块",
    "墙",
    "基础",
    "混凝土",
    "钢筋",
    "梁",
    "板",
    "柱",
    "楼梯",
    "门",
    "窗",
    "屋面",
    "防水",
    "保温",
    "隔热",
    "楼地面",
    "抹灰",
    "天棚",
    "油漆",
    "涂料",
    "模板",
    "脚手架",
    "垂直运输",
    "超高",
    "成品保护",
    "施工排水",
    "降水",
]

FEATURE_TERMS = ["土类别", "岩石类别", "一、二类", "三类", "四类", "深度", "厚度", "高度", "强度", "等级", "直径", "运距", "材料", "部位", "规格"]
TRANSPORT_TERMS = ["运输", "运距", "自卸汽车", "人力车", "装车", "装载", "垂直运输", "转堆", "每增加", "每增"]
METHOD_TERMS = ["模板", "脚手架", "打夯", "碾压", "支护", "支撑", "挡土板", "爆破", "安装", "拆除", "成品保护", "降水", "排水", "机具", "机械"]

PREFIX_HINTS = {
    "A": {"A1-1"},
    "B": {"A1-2"},
    "C": {"A1-3"},
    "D": {"A1-4"},
    "E": {"A1-5", "A1-6"},
    "F": {"A1-7"},
    "G": {"A1-8"},
    "H": {"A1-9"},
    "J": {"A1-10"},
    "K": {"A1-11"},
    "L": {"A1-12"},
    "M": {"A1-13"},
    "N": {"A1-14"},
    "P": {"A1-15"},
    "Q": {"A1-16", "A1-17", "A1-18", "A1-19"},
    "R": {"A1-20", "A1-21", "A1-22", "A1-23", "A1-24", "A1-25"},
}

BUCKET_CONFIG = [
    {
        "bucket": "A_tsf_all",
        "target": 12,
        "appendices": ["A"],
        "mode": "all",
        "keywords": [],
        "business_reason": "土石方已完成试点，需要作为治理对照基准。",
        "mapping_challenge": "运输、装车、回填、施工方法、特征变体混杂。",
        "review_value": "校准 R2/R3/R4/R5 的边界。",
    },
    {
        "bucket": "B_foundation_support",
        "target": 6,
        "appendices": ["B"],
        "keywords": ["换填", "强夯", "水泥土搅拌桩", "钢板桩", "锚杆", "喷射混凝土", "地下连续墙", "边坡"],
        "business_reason": "地基处理和支护常与施工方法、措施、主体项目混杂。",
        "mapping_challenge": "方法项、支护措施、材料/工艺特征。",
        "review_value": "确认哪些可以进入模板种子，哪些只能作为工作内容或措施。",
    },
    {
        "bucket": "C_pile",
        "target": 4,
        "appendices": ["C"],
        "keywords": ["预制桩", "灌注桩", "钢管桩", "截桩头", "桩尖", "声测管"],
        "business_reason": "桩基存在成孔、浇筑、截桩、材料构件等关系。",
        "mapping_challenge": "清单本体与施工工序/材料构件拆分。",
        "review_value": "校准桩基多工序组合的治理角色。",
    },
    {
        "bucket": "D_masonry",
        "target": 4,
        "appendices": ["D"],
        "keywords": ["砖基础", "砖墙", "砌块墙", "石基础", "零星砌砖", "垫层"],
        "business_reason": "砌筑项目与材料、厚度、墙体部位强相关。",
        "mapping_challenge": "材料/部位 feature variant 与主体清单关系。",
        "review_value": "帮助成本部定义 feature condition。",
    },
    {
        "bucket": "E_concrete_rebar",
        "target": 10,
        "appendices": ["E"],
        "keywords": ["混凝土基础", "矩形柱", "构造柱", "基础梁", "矩形梁", "有梁板", "楼梯", "后浇带", "钢筋", "预埋铁件", "预制混凝土"],
        "business_reason": "混凝土和钢筋是企业模板最核心高频对象。",
        "mapping_challenge": "混凝土构件、钢筋、模板、预制安装之间边界。",
        "review_value": "形成后续企业组价模板种子的优先规则。",
    },
    {
        "bucket": "HJKLMN_finish_envelope",
        "target": 8,
        "appendices": ["H", "J", "K", "L", "M", "N"],
        "per_appendix": {"H": 2, "J": 2, "K": 1, "L": 1, "M": 1, "N": 1},
        "keywords": ["门", "窗", "屋面", "防水", "保温", "隔热", "楼地面", "抹灰", "幕墙", "天棚"],
        "business_reason": "装饰、防水、保温、门窗、天棚类覆盖材料和做法差异。",
        "mapping_challenge": "材料层次、做法厚度、构造层、部位关系。",
        "review_value": "校准装饰类 feature 与工作内容边界。",
    },
    {
        "bucket": "R_measures",
        "target": 4,
        "appendices": ["R"],
        "keywords": ["脚手架", "模板", "垂直运输", "超高", "施工排水", "降水", "大型机械"],
        "business_reason": "措施项目不能默认等同主体清单，需要治理红线。",
        "mapping_challenge": "措施、方法、运输类不得自动 direct。",
        "review_value": "明确 forbidden auto-mapping 规则。",
    },
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--project-root", type=Path, default=PROJECT_ROOT_DEFAULT)
    return parser.parse_args()


def norm(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"[ \t\u3000]+", " ", str(value).replace("\r", "\n")).strip()


def compact(value: Any) -> str:
    return re.sub(r"\s+", "", norm(value)).lower()


def contains_any(text: str, terms: Sequence[str]) -> bool:
    c = compact(text)
    return any(compact(term) in c for term in terms)


def read_csv(path: Path) -> List[Dict[str, str]]:
    if not path.exists():
        raise SystemExit(f"blocked_missing_inputs: {path}")
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, fields: Sequence[str], rows: Iterable[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(fields), extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def rel(path: Path, root: Path) -> str:
    return path.relative_to(root).as_posix()


def csv_row_count(path: Path) -> int:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return sum(1 for _ in csv.DictReader(handle))


def artifact_row_count(path: Path) -> str:
    if not path.exists():
        return ""
    if path.suffix.lower() == ".csv":
        return str(csv_row_count(path))
    if path.suffix.lower() == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        return str(sum(max(0, ws.max_row - 1) for ws in workbook.worksheets))
    return ""


def natural_code_key(value: str) -> Tuple[Any, ...]:
    parts = re.split(r"(\d+)", norm(value).upper())
    return tuple(int(part) if part.isdigit() else part for part in parts)


def source_prefix(code: str) -> str:
    parts = norm(code).split("-")
    return "-".join(parts[:2]) if len(parts) >= 2 else norm(code)


def normalize_unit_dimension(unit: str) -> str:
    text = compact(unit).replace("m³", "m3").replace("m²", "m2").replace("㎡", "m2")
    text = text.replace("立方米", "m3").replace("平方米", "m2").replace("米", "m").replace("吨", "t")
    if re.fullmatch(r"(\d+(?:\.\d+)?)?m3", text):
        return "volume"
    if re.fullmatch(r"(\d+(?:\.\d+)?)?m2", text):
        return "area"
    if re.fullmatch(r"(\d+(?:\.\d+)?)?m", text):
        return "length"
    if text == "t":
        return "weight"
    if text == "台班":
        return "machine_shift"
    if text in {"项", "宗"}:
        return "lump_sum"
    return "unknown"


def bill_text(row: Dict[str, str]) -> str:
    return " ".join([row.get("bill_name", ""), row.get("project_feature_raw", ""), row.get("work_content_raw", ""), row.get("section_name", ""), row.get("appendix_name", "")])


def quota_text(row: Dict[str, str]) -> str:
    return " ".join([row.get("quota_name_candidate", ""), row.get("quota_feature_text_candidate", ""), row.get("raw_name", "")])


def terms_in(text: str) -> List[str]:
    return [term for term in ALL_TERMS if contains_any(text, [term])]


def select_rows_by_keywords(candidates: List[Dict[str, str]], keywords: Sequence[str], target: int) -> List[Dict[str, str]]:
    selected: List[Dict[str, str]] = []
    seen = set()
    for keyword in keywords:
        for row in candidates:
            if row["bill_code_9"] in seen:
                continue
            if contains_any(bill_text(row), [keyword]):
                selected.append(row)
                seen.add(row["bill_code_9"])
                break
        if len(selected) >= target:
            break
    for row in candidates:
        if len(selected) >= target:
            break
        if row["bill_code_9"] not in seen:
            selected.append(row)
            seen.add(row["bill_code_9"])
    return selected[:target]


def select_scope(bills: Sequence[Dict[str, str]]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    bills_by_appendix: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    for row in sorted(bills, key=lambda r: r.get("bill_code_9", "")):
        bills_by_appendix[row.get("appendix_code", "")].append(row)

    scope_rows: List[Dict[str, Any]] = []
    rationale: List[Dict[str, Any]] = []
    selected_codes = set()
    scope_id = 1
    for config in BUCKET_CONFIG:
        bucket = config["bucket"]
        target = int(config["target"])
        selected: List[Dict[str, str]] = []
        if config.get("mode") == "all":
            selected = [row for app in config["appendices"] for row in bills_by_appendix.get(app, [])]
        elif "per_appendix" in config:
            for app, count in config["per_appendix"].items():
                selected.extend(select_rows_by_keywords(bills_by_appendix.get(app, []), config["keywords"], int(count)))
            if len(selected) < target:
                pool = [row for app in config["appendices"] for row in bills_by_appendix.get(app, []) if row["bill_code_9"] not in {r["bill_code_9"] for r in selected}]
                selected.extend(select_rows_by_keywords(pool, config["keywords"], target - len(selected)))
        else:
            pool = [row for app in config["appendices"] for row in bills_by_appendix.get(app, [])]
            selected = select_rows_by_keywords(pool, config["keywords"], target)
        selected = [row for row in selected if row.get("bill_code_9") not in selected_codes]
        for row in selected[:target]:
            selected_codes.add(row["bill_code_9"])
            scope_rows.append(
                {
                    "scope_id": f"REP_SCOPE_{scope_id:03d}",
                    "selection_bucket": bucket,
                    "bill_reference_id": row.get("bill_reference_id", ""),
                    "bill_code_9": row.get("bill_code_9", ""),
                    "bill_name": row.get("bill_name", ""),
                    "appendix_code": row.get("appendix_code", ""),
                    "appendix_name": row.get("appendix_name", ""),
                    "section_code": row.get("section_code", ""),
                    "section_name": row.get("section_name", ""),
                    "unit": row.get("unit", ""),
                    "quantity_calculation_rule": row.get("quantity_calculation_rule", ""),
                    "work_content_raw": row.get("work_content_raw", ""),
                    "project_feature_raw": row.get("project_feature_raw", ""),
                    "selection_reason": f"bucket={bucket}; selected from source baseline by appendix and representative keywords",
                    "expected_mapping_challenge": config["mapping_challenge"],
                    "human_scope_confirmed": "",
                    "human_comment": "",
                }
            )
            scope_id += 1
        actual = sum(1 for row in scope_rows if row["selection_bucket"] == bucket)
        rationale.append(
            {
                "selection_bucket": bucket,
                "target_count": target,
                "actual_count": actual,
                "business_reason": config["business_reason"],
                "mapping_challenge_covered": config["mapping_challenge"],
                "cost_department_review_value": config["review_value"],
                "remark": "Selected only from real GB/T bill baseline rows; no synthetic bill codes.",
            }
        )
    return scope_rows, rationale


def score_pair(bill: Dict[str, Any], quota: Dict[str, str]) -> Tuple[float, List[str]]:
    btext = bill_text(bill)
    qtext = quota_text(quota)
    bterms = set(terms_in(btext))
    qterms = set(terms_in(qtext))
    overlap = sorted(bterms.intersection(qterms))
    score = 0.0
    basis: List[str] = []
    if overlap:
        score += len(overlap) * 1.25
        basis.append("keyword_overlap:" + "/".join(overlap[:8]))
    bill_name = compact(bill.get("bill_name", ""))
    if bill_name and bill_name in compact(qtext):
        score += 5.0
        basis.append("bill_name_in_quota_name")
    for term in terms_in(bill.get("work_content_raw", "")):
        if contains_any(qtext, [term]):
            score += 0.7
    if source_prefix(quota.get("source_code", "")) in PREFIX_HINTS.get(bill.get("appendix_code", ""), set()):
        score += 2.0
        basis.append("source_code_prefix_hint")
    bill_dim = normalize_unit_dimension(bill.get("unit", ""))
    quota_dim = normalize_unit_dimension(quota.get("unit", ""))
    if bill_dim != "unknown" and quota_dim != "unknown":
        if bill_dim == quota_dim:
            score += 1.0
            basis.append("unit_dimension_match")
        else:
            score -= 0.8
            basis.append("unit_dimension_mismatch")
    if contains_any(qtext, TRANSPORT_TERMS):
        score -= 0.25
        basis.append("transport_or_loading_risk")
    if contains_any(qtext, METHOD_TERMS):
        basis.append("method_or_measure_risk")
    return score, basis


def classify_role(bill: Dict[str, Any], quota: Dict[str, str], score: float, basis: Sequence[str]) -> Tuple[str, str, str, str]:
    qtext = quota_text(quota)
    btext = bill_text(bill)
    prefix_match = source_prefix(quota.get("source_code", "")) in PREFIX_HINTS.get(bill.get("appendix_code", ""), set())
    transport = contains_any(qtext, TRANSPORT_TERMS)
    method = contains_any(qtext, METHOD_TERMS)
    feature = contains_any(qtext + " " + btext, FEATURE_TERMS)
    dim_mismatch = "unit_dimension_mismatch" in basis
    if dim_mismatch and score < 4.5:
        return "R5_not_applicable_or_unrouted", "candidate rejected by unit dimension or weak evidence", "blocked_by_unit_or_low_score", "low"
    if transport:
        return "R4_method_or_measure_component", "transport/loading/vertical movement should not be direct bill body", "transport_or_loading_review_required", "medium"
    if method and bill.get("appendix_code") != "R":
        return "R4_method_or_measure_component", "method/measure component, not bill body", "method_or_measure_review_required", "medium"
    if score >= 7.0 and prefix_match and not feature:
        return "R1_direct_bill_body", "strong name/object/prefix/unit evidence", "direct after cost review only", "high"
    if score >= 5.0 and prefix_match:
        return "R2_feature_variant", "bill body likely but feature/condition split required", "feature condition required", "medium"
    if score >= 4.0:
        return "R3_work_content_component", "candidate appears related as work content/component", "work content only unless reviewer promotes", "medium"
    return "R5_not_applicable_or_unrouted", "weak evidence, do not use automatically", "unrouted candidate", "low"


def forbidden_use_for(role: str) -> str:
    base = ["must_not_auto_approve", "must_not_write_back_bill_code", "must_not_generate_enterprise_standard_name"]
    if role in {"R3_work_content_component", "R4_method_or_measure_component", "R5_not_applicable_or_unrouted"}:
        base.append("must_not_be_direct_bill_body")
    return ";".join(base)


def issue_types_for(role: str, basis: Sequence[str], confidence: float, quota: Dict[str, str]) -> str:
    issues: List[str] = []
    if role == "R2_feature_variant":
        issues.append("feature_required")
    if role == "R3_work_content_component":
        issues.append("work_content_only")
    if "method_or_measure_risk" in basis:
        issues.append("construction_method_only")
    if "transport_or_loading_risk" in basis:
        issues.append("transport_item_uncertain")
    if "unit_dimension_mismatch" in basis:
        issues.append("unit_dimension_mismatch")
    if confidence < 0.55:
        issues.append("possible_over_mapping")
    if role == "R5_not_applicable_or_unrouted":
        issues.append("possible_over_mapping")
    return ";".join(dict.fromkeys(issues))


def prepare_quota_profiles(quotas: Sequence[Dict[str, str]]) -> List[Dict[str, Any]]:
    profiles: List[Dict[str, Any]] = []
    for quota in quotas:
        qtext = quota_text(quota)
        profiles.append(
            {
                "quota": quota,
                "qtext": qtext,
                "qcompact": compact(qtext),
                "qterms": set(terms_in(qtext)),
                "prefix": source_prefix(quota.get("source_code", "")),
                "unit_dim": normalize_unit_dimension(quota.get("unit", "")),
                "transport": contains_any(qtext, TRANSPORT_TERMS),
                "method": contains_any(qtext, METHOD_TERMS),
            }
        )
    return profiles


def score_pair_prepared(bill_profile: Dict[str, Any], quota_profile: Dict[str, Any]) -> Tuple[float, List[str]]:
    bterms = bill_profile["bterms"]
    qterms = quota_profile["qterms"]
    prefix_match = quota_profile["prefix"] in bill_profile["prefix_hints"]
    overlap = sorted(bterms.intersection(qterms))
    bill_name = bill_profile["bill_name_compact"]

    score = 0.0
    basis: List[str] = []
    if overlap:
        score += len(overlap) * 1.25
        basis.append("keyword_overlap:" + "/".join(overlap[:8]))
    if bill_name and bill_name in quota_profile["qcompact"]:
        score += 5.0
        basis.append("bill_name_in_quota_name")
    for term in bill_profile["work_terms"]:
        if term in quota_profile["qtext"]:
            score += 0.7
    if prefix_match:
        score += 2.0
        basis.append("source_code_prefix_hint")

    bill_dim = bill_profile["unit_dim"]
    quota_dim = quota_profile["unit_dim"]
    if bill_dim != "unknown" and quota_dim != "unknown":
        if bill_dim == quota_dim:
            score += 1.0
            basis.append("unit_dimension_match")
        else:
            score -= 0.8
            basis.append("unit_dimension_mismatch")
    if quota_profile["transport"]:
        score -= 0.25
        basis.append("transport_or_loading_risk")
    if quota_profile["method"]:
        basis.append("method_or_measure_risk")
    return score, basis


def build_details(scope_rows: Sequence[Dict[str, Any]], quotas: Sequence[Dict[str, str]], pricing_by_code: Dict[str, Dict[str, str]]) -> List[Dict[str, Any]]:
    details: List[Dict[str, Any]] = []
    quota_profiles = prepare_quota_profiles(quotas)
    for bill in scope_rows:
        btext = bill_text(bill)
        bill_profile = {
            "btext": btext,
            "bterms": set(terms_in(btext)),
            "bill_name_compact": compact(bill.get("bill_name", "")),
            "work_terms": terms_in(bill.get("work_content_raw", "")),
            "prefix_hints": PREFIX_HINTS.get(bill.get("appendix_code", ""), set()),
            "unit_dim": normalize_unit_dimension(bill.get("unit", "")),
        }
        scored: List[Tuple[float, Dict[str, str], List[str]]] = []
        for quota_profile in quota_profiles:
            score, basis = score_pair_prepared(bill_profile, quota_profile)
            if score >= 3.2:
                scored.append((score, quota_profile["quota"], basis))
        scored.sort(key=lambda item: item[0], reverse=True)
        for score, quota, basis in scored[:12]:
            role, role_basis, condition, readiness = classify_role(bill, quota, score, basis)
            confidence = max(0.1, min(0.96, score / 10.0))
            pricing = pricing_by_code.get(quota.get("source_code", ""), {})
            detail = {
                "bill_code_9": bill.get("bill_code_9", ""),
                "bill_name": bill.get("bill_name", ""),
                "appendix_code": bill.get("appendix_code", ""),
                "appendix_name": bill.get("appendix_name", ""),
                "bill_unit": bill.get("unit", ""),
                "bill_quantity_calculation_rule": bill.get("quantity_calculation_rule", ""),
                "bill_work_content_raw": bill.get("work_content_raw", ""),
                "bill_project_feature_raw": bill.get("project_feature_raw", ""),
                "quota_source_code": quota.get("source_code", ""),
                "quota_raw_name": quota.get("raw_name", ""),
                "quota_name_candidate": quota.get("quota_name_candidate", ""),
                "quota_feature_text_candidate": quota.get("quota_feature_text_candidate", ""),
                "quota_unit": quota.get("unit", ""),
                "quota_raw_total_fee": pricing.get("raw_total_fee", quota.get("raw_total_fee", "")),
                "source_code_prefix": source_prefix(quota.get("source_code", "")),
                "governance_role": role,
                "relationship_basis": role_basis + ";" + ";".join(basis[:8]),
                "mapping_confidence": f"{confidence:.2f}",
                "selection_condition": condition,
                "forbidden_use": forbidden_use_for(role),
                "issue_types": "",
                "review_status": REVIEW_STATUS,
                "cost_engineer_decision": "",
                "cost_engineer_comment": "",
                "_readiness": readiness,
            }
            detail["issue_types"] = issue_types_for(role, basis, confidence, quota)
            details.append(detail)
    details.sort(key=lambda row: (row["bill_code_9"], row["governance_role"], natural_code_key(row["quota_source_code"])))
    return details


def top_values(rows: Sequence[Dict[str, Any]], key: str, limit: int = 8) -> str:
    out: List[str] = []
    seen = set()
    for row in rows:
        value = norm(row.get(key, ""))
        if value and value not in seen:
            out.append(value)
            seen.add(value)
        if len(out) >= limit:
            break
    return ";".join(out)


def decision_from_counts(counts: Counter, total: int) -> Tuple[str, str, str]:
    if total == 0:
        return "no_reliable_quota_candidate", "blocked", "No reliable candidate. Cost department must decide whether to defer."
    if counts.get("R1_direct_bill_body", 0) > 0:
        return "ready_for_template_seed_after_review", "high", "Confirm R1 rows and feature boundaries."
    if counts.get("R2_feature_variant", 0) > 0:
        return "usable_with_feature_conditions", "medium", "Confirm feature conditions before template seed."
    if counts.get("R3_work_content_component", 0) > 0 and counts.get("R4_method_or_measure_component", 0) == 0:
        return "only_work_content_components_available", "low", "Review whether components support bill work content only."
    if counts.get("R4_method_or_measure_component", 0) > 0:
        return "requires_cost_department_review", "low", "Method/measure/transport candidates require manual classification."
    return "defer", "blocked", "Weak or unrouted evidence."


def build_matrix(scope_rows: Sequence[Dict[str, Any]], details: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    by_bill: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in details:
        by_bill[row["bill_code_9"]].append(row)
    matrix: List[Dict[str, Any]] = []
    for bill in scope_rows:
        rows = sorted(by_bill.get(bill["bill_code_9"], []), key=lambda row: float(row.get("mapping_confidence", "0")), reverse=True)
        counts = Counter(row["governance_role"] for row in rows)
        decision, readiness, confirmation = decision_from_counts(counts, len(rows))
        matrix.append(
            {
                "bill_code_9": bill["bill_code_9"],
                "bill_name": bill["bill_name"],
                "appendix_code": bill["appendix_code"],
                "appendix_name": bill["appendix_name"],
                "unit": bill["unit"],
                "quantity_calculation_rule": bill["quantity_calculation_rule"],
                "work_content_raw": bill["work_content_raw"],
                "project_feature_raw": bill["project_feature_raw"],
                "total_quota_candidate_count": len(rows),
                "R1_direct_bill_body_count": counts.get("R1_direct_bill_body", 0),
                "R2_feature_variant_count": counts.get("R2_feature_variant", 0),
                "R3_work_content_component_count": counts.get("R3_work_content_component", 0),
                "R4_method_or_measure_component_count": counts.get("R4_method_or_measure_component", 0),
                "R5_not_applicable_or_unrouted_count": counts.get("R5_not_applicable_or_unrouted", 0),
                "top_quota_source_codes": top_values(rows, "quota_source_code"),
                "top_quota_names": top_values(rows, "quota_raw_name"),
                "top_quota_units": top_values(rows, "quota_unit"),
                "recommended_bill_level_decision": decision,
                "template_readiness": readiness,
                "required_human_confirmation": confirmation,
                "cost_engineer_decision": "",
                "cost_engineer_comment": "",
            }
        )
    return matrix


def shared_type_for(role: str, rows: Sequence[Dict[str, Any]]) -> str:
    text = " ".join(row.get("quota_raw_name", "") for row in rows)
    if "transport_item_uncertain" in ";".join(row.get("issue_types", "") for row in rows) or contains_any(text, TRANSPORT_TERMS):
        return "transport_component_shared"
    if role == "R2_feature_variant":
        return "feature_variant_shared"
    if role == "R3_work_content_component":
        return "work_content_component_shared"
    if role == "R4_method_or_measure_component":
        return "method_component_shared"
    return "possible_over_mapping"


def build_shared_components(details: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in details:
        grouped[row["quota_source_code"]].append(row)
    rows: List[Dict[str, Any]] = []
    for quota_code, group in sorted(grouped.items(), key=lambda item: natural_code_key(item[0])):
        bill_codes = sorted({row["bill_code_9"] for row in group})
        if len(bill_codes) <= 1:
            continue
        roles = Counter(row["governance_role"] for row in group)
        role = roles.most_common(1)[0][0]
        rows.append(
            {
                "quota_source_code": quota_code,
                "quota_raw_name": group[0]["quota_raw_name"],
                "quota_name_candidate": group[0]["quota_name_candidate"],
                "candidate_bill_codes": ";".join(bill_codes),
                "candidate_bill_names": ";".join(sorted({row["bill_name"] for row in group})),
                "candidate_count": len(bill_codes),
                "governance_role": role,
                "shared_component_type": shared_type_for(role, group),
                "selection_condition": "shared candidate must have bill-specific feature/work-content condition",
                "allowed_use": "evidence candidate for cost engineer review only",
                "forbidden_use": forbidden_use_for(role),
                "review_priority": "high" if len(bill_codes) >= 3 or role in {"R4_method_or_measure_component", "R5_not_applicable_or_unrouted"} else "medium",
                "cost_engineer_decision": "",
                "cost_engineer_comment": "",
            }
        )
    return rows


def add_issue(issues: List[Dict[str, Any]], issue_type: str, matrix_row: Optional[Dict[str, Any]], detail: Optional[Dict[str, Any]], issue_detail: str, severity: str, action: str) -> None:
    issues.append(
        {
            "issue_id": f"ISSUE_REP_GOV_{len(issues) + 1:05d}",
            "issue_type": issue_type,
            "bill_code_9": (detail or matrix_row or {}).get("bill_code_9", ""),
            "bill_name": (detail or matrix_row or {}).get("bill_name", ""),
            "quota_source_code": (detail or {}).get("quota_source_code", ""),
            "quota_raw_name": (detail or {}).get("quota_raw_name", ""),
            "governance_role": (detail or {}).get("governance_role", ""),
            "issue_detail": issue_detail,
            "severity": severity,
            "suggested_action": action,
        }
    )


def build_issues(matrix: Sequence[Dict[str, Any]], details: Sequence[Dict[str, Any]], shared: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    issues: List[Dict[str, Any]] = []
    for row in matrix:
        total = int(row["total_quota_candidate_count"])
        if total == 0:
            add_issue(issues, "no_candidate_quota", row, None, "No reliable quota candidate found by lightweight rules.", "high", "Cost department should decide defer/reject or refine rules.")
        if total > 10:
            add_issue(issues, "too_many_candidate_quota", row, None, f"{total} candidate quota rows selected.", "medium", "Review feature grouping and selection conditions.")
        if row["recommended_bill_level_decision"] in {"requires_cost_department_review", "defer", "no_reliable_quota_candidate"}:
            add_issue(issues, "cost_department_review_required", row, None, "Bill-level decision requires manual confirmation.", "high", "Review matrix first, then detail rows.")
    for detail in details:
        for issue_type in [item for item in detail.get("issue_types", "").split(";") if item]:
            severity = "high" if issue_type in {"transport_item_uncertain", "unit_dimension_mismatch", "possible_over_mapping"} else "medium"
            add_issue(issues, issue_type, None, detail, f"Detail issue: {issue_type}; basis={detail.get('relationship_basis')}", severity, "Do not approve automatically.")
    for row in shared:
        add_issue(issues, "shared_quota_component", None, {"bill_code_9": row["candidate_bill_codes"], "bill_name": row["candidate_bill_names"], "quota_source_code": row["quota_source_code"], "quota_raw_name": row["quota_raw_name"], "governance_role": row["governance_role"]}, f"Quota appears under {row['candidate_count']} representative bill items.", "high" if row["review_priority"] == "high" else "medium", "Require selection_condition before any template seed.")
    add_issue(issues, "candidate_name_not_final_standard", None, None, "quota_name_candidate is a candidate description only, not an enterprise standard name.", "high", "Enterprise standard names must wait for reviewed template stage.")
    return issues


def build_review_sheet(matrix: Sequence[Dict[str, Any]], issues: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    issue_by_bill: Dict[str, List[str]] = defaultdict(list)
    for issue in issues:
        for code in norm(issue.get("bill_code_9", "")).split(";"):
            if code:
                issue_by_bill[code].append(issue.get("issue_type", ""))
    rows: List[Dict[str, Any]] = []
    for row in matrix:
        issue_types = list(dict.fromkeys(issue_by_bill.get(row["bill_code_9"], [])))
        rows.append(
            {
                "bill_code_9": row["bill_code_9"],
                "bill_name": row["bill_name"],
                "appendix_name": row["appendix_name"],
                "recommended_bill_level_decision": row["recommended_bill_level_decision"],
                "template_readiness": row["template_readiness"],
                "top_quota_source_codes": row["top_quota_source_codes"],
                "top_quota_names": row["top_quota_names"],
                "main_risk": ";".join(issue_types[:5]) if issue_types else "feature confirmation required",
                "cost_engineer_decision": "",
                "cost_engineer_selected_quota_codes": "",
                "cost_engineer_comment": "",
                "follow_up_required": "yes" if issue_types else "review",
                "remark": "decision options: accept_for_template_seed; accept_with_feature_conditions; work_content_only; requires_cost_department_review; defer; reject_current_mapping",
            }
        )
    return rows


def write_governance_rules(path: Path) -> None:
    lines = [
        "# Representative Mapping Governance Rules V0.2",
        "",
        "## 1. Governance Objective",
        "",
        "This is a multi-appendix representative governance calibration package. It is not a full approved mapping and must not be used as enterprise production mapping.",
        "",
        "## 2. Evidence Hierarchy",
        "",
        "1. GB/T 50854 and GD2018 source baselines are the highest evidence.",
        "2. Company historical projects, Glodon pricing files, and cost-department experience are enterprise evidence.",
        "3. Internet experience and AI inference are candidate hints only and cannot directly confirm mappings.",
        "",
        "## 3. Relationship Roles",
        "",
        "- `R1_direct_bill_body`: quota appears to describe the bill item body itself, still pending review.",
        "- `R2_feature_variant`: quota is a feature/material/spec/depth/condition variant under a bill item.",
        "- `R3_work_content_component`: quota supports a work-content component but is not bill body.",
        "- `R4_method_or_measure_component`: quota is method, measure, transport, support, or temporary process.",
        "- `R5_not_applicable_or_unrouted`: evidence is weak, mismatched, or intentionally unrouted.",
        "",
        "## 4. Multi-Candidate Governance",
        "",
        "A quota may appear under multiple bill items, but every reuse must carry a `selection_condition` explaining when it can be used.",
        "",
        "## 5. Forbidden Auto-Mapping Rules",
        "",
        "- 不允许为了提高覆盖率强行匹配。",
        "- 不允许运输、装车、垂直运输默认 direct。",
        "- 不允许把施工方法直接等同于 bill item。",
        "- 不允许把 feature variant 当作最终标准名称。",
        "- 不允许 bill_code 回写到 quota reference。",
        "- 不允许生成 approved。",
        "",
        "## 6. Cost Department Review Workflow",
        "",
        "Cost engineers should review the matrix first to decide bill-level feasibility, then inspect detail rows and shared components.",
        "",
        "## 7. Stage Gate",
        "",
        "Only after cost-department confirmation may selected rows move to enterprise template seed draft. This stage itself may not create enterprise templates.",
        "",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")


def worksheet_safe(value: Any) -> Any:
    if isinstance(value, str) and len(value) > 32767:
        return value[:32740] + "\n...[TRUNCATED_FOR_XLSX_CELL_LIMIT]"
    return value


def write_xlsx(path: Path, specs: Sequence[Tuple[str, Sequence[str], Sequence[Dict[str, Any]]]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for sheet_name, fields, rows in specs:
        ws = wb.create_sheet(sheet_name[:31])
        ws.append(list(fields))
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for row in rows:
            ws.append([worksheet_safe(row.get(field, "")) for field in fields])
        ws.freeze_panes = "A2"
        for idx, field in enumerate(fields, start=1):
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = min(max(len(field) + 2, 12), 46)
        for row_cells in ws.iter_rows(min_row=2):
            for cell in row_cells:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def summary_rows(scope: Sequence[Dict[str, Any]], matrix: Sequence[Dict[str, Any]], details: Sequence[Dict[str, Any]], shared: Sequence[Dict[str, Any]], issues: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return [
        {"metric_name": "representative_bill_items", "metric_value": len(scope), "remark": "target around 48"},
        {"metric_name": "detail_rows", "metric_value": len(details), "remark": ""},
        {"metric_name": "shared_quota_components", "metric_value": len(shared), "remark": ""},
        {"metric_name": "issue_rows", "metric_value": len(issues), "remark": ""},
        {"metric_name": "high_risk_issue_rows", "metric_value": sum(1 for issue in issues if issue.get("severity") == "high"), "remark": ""},
        {"metric_name": "approved_count", "metric_value": 0, "remark": "no approved generated"},
        {"metric_name": "recommendation", "metric_value": recommendation(scope, matrix, details), "remark": ""},
    ]


def recommendation(scope: Sequence[Dict[str, Any]], matrix: Sequence[Dict[str, Any]], details: Sequence[Dict[str, Any]]) -> str:
    if len(scope) < 44 or len(scope) > 52 or len(matrix) != len(scope) or not details:
        return "representative_governance_partial_manual_intervention_required"
    return "representative_governance_ready_for_cost_engineer_review"


def write_report(path: Path, scope: Sequence[Dict[str, Any]], rationale: Sequence[Dict[str, Any]], matrix: Sequence[Dict[str, Any]], details: Sequence[Dict[str, Any]], shared: Sequence[Dict[str, Any]], issues: Sequence[Dict[str, Any]]) -> str:
    rec = recommendation(scope, matrix, details)
    bucket_counts = Counter(row["selection_bucket"] for row in scope)
    role_counts = Counter(row["governance_role"] for row in details)
    issue_counts = Counter(row["issue_type"] for row in issues)
    lines = [
        "# Stage MAP-REPRESENTATIVE-GOVERNANCE-CALIBRATION-1 Report",
        "",
        "## 1. Task Scope",
        "",
        "Generate a multi-appendix representative mapping governance calibration package for cost-engineer review. This is not a full mapping and not an approved mapping.",
        "",
        "## 2. Why More Bill Items Are Needed",
        "",
        "Appendix A earthwork alone cannot calibrate governance for piles, masonry, concrete, decoration, waterproofing, insulation, doors/windows, ceilings, or measures. Representative cross-section sampling exposes direct, feature, work-content, method, measure, and unrouted relationships.",
        "",
        "## 3. Selected Representative Bill Items",
        "",
        f"- representative_bill_item_count: {len(scope)}",
        f"- bucket_counts: {json.dumps(dict(bucket_counts), ensure_ascii=False, sort_keys=True)}",
        "",
        "## 4. GD2018 Candidate Source",
        "",
        "Candidate quota rows are read from `gd2018_normalized_quota_items_full_review.csv` and pricing fields are preserved only as reference context.",
        "",
        "## 5. Mapping Governance Strategy",
        "",
        "Lightweight rules use object keywords, part/material keywords, unit dimension, work-content text, project-feature text, quota names, source-code prefix hints, and pricing fields as non-confirming context. No candidate is automatically confirmed.",
        "",
        "## 6. Matrix Summary",
        "",
        f"- matrix_rows: {len(matrix)}",
        f"- detail_rows: {len(details)}",
        f"- role_counts: {json.dumps(dict(role_counts), ensure_ascii=False, sort_keys=True)}",
        "",
        "## 7. Shared Component Findings",
        "",
        f"- shared_quota_component_rows: {len(shared)}",
        "- Shared quota components require bill-specific selection conditions before any template seed.",
        "",
        "## 8. High-Risk Categories",
        "",
        f"- issue_rows: {len(issues)}",
        f"- high_risk_issue_rows: {sum(1 for issue in issues if issue.get('severity') == 'high')}",
        f"- issue_type_counts: {json.dumps(dict(issue_counts), ensure_ascii=False, sort_keys=True)}",
        "",
        "## 9. Cost Department Review Guidance",
        "",
        "- Review `representative_bill_to_quota_matrix.csv` first.",
        "- Then review `representative_bill_to_quota_detail.csv` for each bill item.",
        "- Use `representative_shared_quota_components.csv` to prevent over-mapping shared method/transport/work-content rows.",
        "",
        "## 10. Not Approved / Not Final Statement",
        "",
        "All rows remain pending. This stage does not write databases, approve mappings, write bill_code back to quota references, generate internal_price_library, generate enterprise standard names, or create enterprise templates.",
        "",
        "## 11. Next Step Recommendation",
        "",
        rec,
        "",
        f"Generated at: {datetime.now().isoformat(timespec='seconds')}",
        "",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")
    return rec


def manifest_row(stage: str, artifact: str, path: Path, source_file: str, project_root: Path) -> Dict[str, str]:
    exists = path.exists()
    return {
        "stage_name": stage,
        "artifact_name": artifact,
        "expected_path": rel(path, project_root),
        "exists": str(exists).lower(),
        "file_size_bytes": str(path.stat().st_size) if exists else "",
        "row_count": artifact_row_count(path) if exists else "",
        "sha256": sha256_file(path) if exists else "",
        "created_or_modified_time": datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds") if exists else "",
        "source_file": source_file,
        "can_regenerate": "true",
        "backup_required": "true",
        "backup_path": "construction_cost_knowledge_engine/data/private/reference_extraction/backups/runs_backup_after_MAP_REPRESENTATIVE_GOVERNANCE_CALIBRATION_1",
        "status": "generated" if exists else "missing",
        "remark": "representative governance calibration artifact; pending review; no approved mapping",
    }


def update_manifest(project_root: Path, output_dir: Path, artifacts: Sequence[str]) -> None:
    manifest_path = project_root / DOCS_REF_REL / "reference_artifact_manifest.csv"
    existing = read_csv(manifest_path) if manifest_path.exists() else []
    source_file = ";".join([rel(project_root / GB_BILLS_REL, project_root), rel(project_root / GD_QUOTA_REL, project_root)])
    replacement = {
        (STAGE_NAME, artifact): manifest_row(STAGE_NAME, artifact, output_dir / artifact, source_file, project_root)
        for artifact in artifacts
    }
    filtered = [row for row in existing if (row.get("stage_name"), row.get("artifact_name")) not in replacement]
    filtered.extend(replacement.values())
    write_csv(manifest_path, MANIFEST_FIELDS, filtered)
    write_manifest_md(project_root, filtered)


def write_manifest_md(project_root: Path, rows: Sequence[Dict[str, str]]) -> None:
    latest = [row for row in rows if row.get("stage_name") == STAGE_NAME]
    registered = len(rows)
    existing = sum(1 for row in rows if row.get("exists") == "true")
    lines = [
        "# Reference Artifact Manifest",
        "",
        "## Governance",
        "",
        "- `construction_cost_knowledge_engine/data/private/` is intentionally not tracked by Git.",
        "- Every private artifact must be registered with `row_count` and `sha256` after generation.",
        "- Each completed stage must back up its `runs` output directory after validation.",
        "- Mock SQLite and seed CSV artifacts must not be used as source of truth.",
        "- Representative governance outputs are pending review artifacts only and do not approve mappings.",
        "",
        "## Current Manifest Summary",
        "",
        f"- registered_artifacts: {registered}",
        f"- existing_artifacts: {existing}",
        f"- missing_artifacts: {registered - existing}",
        "",
        "## Manifest CSV",
        "",
        "`construction_cost_knowledge_engine/docs/reference_extraction/reference_artifact_manifest.csv`",
        "",
        "## Latest Representative Governance Calibration Outputs",
        "",
    ]
    for row in latest:
        lines.append(f"- `{row.get('expected_path')}` ({row.get('row_count') or 'n/a'} rows)")
    lines.extend(
        [
            "",
            "## Backup Requirement",
            "",
            "`construction_cost_knowledge_engine/data/private/reference_extraction/backups/runs_backup_after_MAP_REPRESENTATIVE_GOVERNANCE_CALIBRATION_1/`",
            "",
        ]
    )
    (project_root / DOCS_REF_REL / "REFERENCE_ARTIFACT_MANIFEST.md").write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    args = parse_args()
    project_root = args.project_root
    output_dir = project_root / OUTPUT_DIR_REL
    output_dir.mkdir(parents=True, exist_ok=True)

    bills = read_csv(project_root / GB_BILLS_REL)
    _rules = read_csv(project_root / GB_RULES_REL)
    quotas = read_csv(project_root / GD_QUOTA_REL)
    pricing = read_csv(project_root / GD_PRICING_REL)
    pricing_by_code = {row.get("source_code", ""): row for row in pricing}

    scope, rationale = select_scope(bills)
    details = build_details(scope, quotas, pricing_by_code)
    matrix = build_matrix(scope, details)
    shared = build_shared_components(details)
    issues = build_issues(matrix, details, shared)
    review_rows = build_review_sheet(matrix, issues)
    summary = summary_rows(scope, matrix, details, shared, issues)

    write_csv(output_dir / "representative_bill_scope_48.csv", SCOPE_FIELDS, scope)
    write_csv(output_dir / "representative_bill_selection_rationale.csv", RATIONALE_FIELDS, rationale)
    write_csv(output_dir / "representative_bill_to_quota_matrix.csv", MATRIX_FIELDS, matrix)
    write_csv(output_dir / "representative_bill_to_quota_detail.csv", DETAIL_FIELDS, details)
    write_csv(output_dir / "representative_shared_quota_components.csv", SHARED_FIELDS, shared)
    write_csv(output_dir / "representative_mapping_issues.csv", ISSUE_FIELDS, issues)
    write_governance_rules(output_dir / "representative_governance_rules_v0_2.md")
    write_csv(output_dir / "cost_engineer_review_sheet.csv", REVIEW_FIELDS, review_rows)
    rec = write_report(output_dir / "stage_map_representative_governance_calibration_report.md", scope, rationale, matrix, details, shared, issues)
    try:
        write_xlsx(
            output_dir / "Representative_Mapping_Governance_Calibration_Review.xlsx",
            [
                ("representative_bill_scope", SCOPE_FIELDS, scope),
                ("bill_to_quota_matrix", MATRIX_FIELDS, matrix),
                ("bill_to_quota_detail", DETAIL_FIELDS, details),
                ("shared_quota_components", SHARED_FIELDS, shared),
                ("mapping_issues", ISSUE_FIELDS, issues),
                ("cost_engineer_review_sheet", REVIEW_FIELDS, review_rows),
                ("selection_rationale", RATIONALE_FIELDS, rationale),
                ("summary", ["metric_name", "metric_value", "remark"], summary),
            ],
        )
    except Exception as exc:
        raise SystemExit(f"blocked_xlsx_generation_failed: {exc}") from exc

    artifacts = [
        "representative_bill_scope_48.csv",
        "representative_bill_selection_rationale.csv",
        "representative_bill_to_quota_matrix.csv",
        "representative_bill_to_quota_detail.csv",
        "representative_shared_quota_components.csv",
        "representative_mapping_issues.csv",
        "representative_governance_rules_v0_2.md",
        "cost_engineer_review_sheet.csv",
        "Representative_Mapping_Governance_Calibration_Review.xlsx",
        "stage_map_representative_governance_calibration_report.md",
    ]
    update_manifest(project_root, output_dir, artifacts)

    print(f"recommendation={rec}")
    print(f"representative_bill_items={len(scope)}")
    print("selection_bucket_counts=" + json.dumps(dict(Counter(row['selection_bucket'] for row in scope)), ensure_ascii=False, sort_keys=True))
    print(f"detail_rows={len(details)}")
    print(f"shared_quota_components={len(shared)}")
    print(f"high_risk_issue_rows={sum(1 for issue in issues if issue.get('severity') == 'high')}")
    print(f"xlsx_exists={(output_dir / 'Representative_Mapping_Governance_Calibration_Review.xlsx').exists()}")
    print(f"output_dir={output_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
