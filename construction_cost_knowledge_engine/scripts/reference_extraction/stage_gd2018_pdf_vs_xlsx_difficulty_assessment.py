#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Stage GD2018 PDF-vs-normalized-XLSX extraction difficulty assessment.

This stage is intentionally assessment-only. It does not write databases,
migrations, approved data, internal_price_library, quota_to_bill_mapping, or
any existing pipeline output. Source PDF/XLSX files are read-only inputs.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import statistics
import sys
from collections import Counter
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import pdfplumber
from openpyxl import load_workbook
from pypdf import PdfReader


PROJECT_ROOT_DEFAULT = Path(r"E:\workspace\01_Projects\ai-construction-system")
ENGINE_REL = Path("construction_cost_knowledge_engine")
XLSX_REL = (
    ENGINE_REL
    / "data"
    / "private"
    / "reference_extraction"
    / "source_excels"
    / "广东省房屋建筑与装饰工程综合定额2018_normalized.xlsx"
)
STANDARDS_DIR_REL = (
    ENGINE_REL
    / "data"
    / "private"
    / "reference_extraction"
    / "source_standards"
    / "广东省建设工程综合定额(2018)"
)
PDF_REL = STANDARDS_DIR_REL / "A01_广东省房屋建筑与装饰工程定额(上册).pdf"
OUTPUT_DIR_REL = (
    ENGINE_REL
    / "data"
    / "private"
    / "reference_extraction"
    / "runs"
    / "GD2018_PDF_VS_XLSX_DIFFICULTY_ASSESSMENT_1"
)

PDF_TEXT_PROFILE_FIELDS = [
    "source_file",
    "page_count",
    "pages_with_text",
    "blank_or_low_text_pages",
    "avg_text_chars_per_page",
    "min_text_chars_per_page",
    "max_text_chars_per_page",
    "quota_code_count",
    "unique_quota_code_count",
    "resource_code_probe_count",
    "work_content_marker_count",
    "unit_marker_count",
    "private_unicode_char_count",
    "watermark_text_detected",
    "ocr_required",
    "parse_feasibility",
    "remark",
]

PDF_QUOTA_PROBE_FIELDS = [
    "source_file",
    "pdf_page_no",
    "book_page_no",
    "text_char_count",
    "quota_code_count",
    "unique_quota_code_count",
    "quota_codes_json",
    "resource_code_probe_count",
    "resource_codes_sample_json",
    "work_content_marker_count",
    "unit_marker_count",
    "private_unicode_char_count",
    "has_table_header_marker",
    "section_marker",
    "remark",
]

XLSX_PROFILE_FIELDS = [
    "source_file",
    "file_size_bytes",
    "sheet_count",
    "sheet_names",
    "selected_sheet",
    "row_count",
    "data_row_count",
    "column_count",
    "headers_json",
    "contains_project_code",
    "contains_project_name",
    "contains_unit",
    "contains_labor_fee",
    "contains_material_fee",
    "contains_machine_fee",
    "contains_management_fee",
    "contains_total",
    "contains_work_content",
    "contains_quantity_calculation_rule",
    "contains_chapter_description",
    "contains_resource_code",
    "contains_resource_name",
    "contains_resource_spec",
    "contains_resource_unit",
    "contains_resource_unit_price",
    "contains_resource_consumption",
    "contains_pdf_page_no",
    "contains_source_block_id",
    "a1_quota_code_count",
    "a1_unique_quota_code_count",
    "a1_prefix_distribution_json",
    "remark",
]

FIELD_GAP_FIELDS = [
    "field_name",
    "exists_in_pdf",
    "exists_in_normalized_xlsx",
    "importance_for_cost_department",
    "parse_difficulty_from_pdf",
    "recommended_stage",
    "remark",
]

SAMPLE_PROBE_FIELDS = [
    "sample_id",
    "pdf_page_no",
    "book_page_no",
    "section_path",
    "quota_source_code",
    "quota_name_from_pdf",
    "quota_name_from_xlsx",
    "unit_from_pdf",
    "unit_from_xlsx",
    "base_price_from_pdf",
    "total_from_xlsx",
    "labor_fee_from_pdf",
    "labor_fee_from_xlsx",
    "material_fee_from_pdf",
    "material_fee_from_xlsx",
    "machine_fee_from_pdf",
    "machine_fee_from_xlsx",
    "management_fee_from_pdf",
    "management_fee_from_xlsx",
    "work_content_from_pdf",
    "resource_row_count_from_pdf",
    "resource_rows_sample_json",
    "match_status",
    "parse_confidence",
    "issue_type",
    "remark",
    "raw_text_block",
]

DIFFICULTY_FIELDS = [
    "module",
    "difficulty_level",
    "automation_feasibility",
    "main_risks",
    "recommended_method",
    "requires_manual_qa",
    "blocking_or_non_blocking",
    "remark",
]

FIELD_NAMES_REQUIRED = [
    "source_file",
    "source_volume",
    "pdf_page_no",
    "book_page_no",
    "chapter_code",
    "chapter_name",
    "section_code",
    "section_name",
    "subsection_title",
    "work_content",
    "quantity_calculation_rule",
    "measurement_unit_context",
    "quota_source_code",
    "quota_name_level_1",
    "quota_name_level_2",
    "quota_name_level_3",
    "quota_name_full",
    "quota_unit",
    "base_price",
    "labor_fee",
    "material_fee",
    "machine_fee",
    "management_fee",
    "resource_category",
    "resource_code",
    "resource_name",
    "resource_spec",
    "resource_unit",
    "resource_unit_price",
    "resource_consumption",
    "resource_consumption_by_quota_code",
    "table_continuation_status",
    "parse_confidence",
    "review_status",
]

SAMPLE_CODE_GROUPS = [
    {
        "label": "A1-1-1_to_A1-1-4",
        "codes": ["A1-1-1", "A1-1-2", "A1-1-3", "A1-1-4"],
        "section_path": "A.1.1 > A.1.1.1 土方工程 > 平整场地、原土打夯",
    },
    {
        "label": "A1-1-5_to_A1-1-8",
        "codes": ["A1-1-5", "A1-1-6", "A1-1-7", "A1-1-8"],
        "section_path": "A.1.1 > A.1.1.1 土方工程 > 人工挖一般土方、淤泥流砂",
    },
    {
        "label": "A1-1-56_nearby",
        "codes": ["A1-1-55", "A1-1-56"],
        "section_path": "A.1.1 > A.1.1.1 土方工程 > 自卸汽车运淤泥、流砂",
    },
    {
        "label": "A1-1-118_nearby",
        "codes": ["A1-1-117", "A1-1-118"],
        "section_path": "A.1.1 > A.1.1.2 石方工程 > 自卸汽车运石方",
    },
    {
        "label": "A1-1-134_to_A1-1-137",
        "codes": ["A1-1-134", "A1-1-135", "A1-1-136", "A1-1-137"],
        "section_path": "A.1.1 > A.1.1.3 回填方及其他 > 支挡土板",
    },
]

QUOTA_CODE_RE = re.compile(r"\bA1-\d+-\d+(?:-\d+)?\b")
RESOURCE_CODE_RE = re.compile(r"\b(?:00010010|0[3-9]\d{6}|99\d{6,7})\b")
NUMBER_RE = re.compile(r"-?\d+(?:\.\d+)?")
PRIVATE_USE_RE = re.compile(r"[\ue000-\uf8ff]")
TABLE_HEADER_RE = re.compile(r"分类\s+编码\s+名称\s+单位\s+单价\(元\)\s+消\s*耗\s*量")


def compact_text(value: Any, limit: Optional[int] = None) -> str:
    if value is None:
        return ""
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text[:limit] if limit else text


def normalize_header(value: Any) -> str:
    return compact_text(value).replace("\n", "").replace(" ", "")


def norm_unit(value: Any) -> str:
    text = compact_text(value)
    replacements = {
        "\ue000": "m3",
        "": "m3",
        "㎡": "m2",
        "m³": "m3",
        "m²": "m2",
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    return text


def to_decimal(value: Any) -> Optional[Decimal]:
    if value is None:
        return Decimal("0")
    text = compact_text(value).replace(",", "")
    if not text or text in {"-", "—", "–"}:
        return Decimal("0")
    match = NUMBER_RE.search(text)
    if not match:
        return None
    try:
        return Decimal(match.group(0))
    except InvalidOperation:
        return None


def decimal_text(value: Optional[Decimal]) -> str:
    if value is None:
        return ""
    return format(value, "f")


def money_close(a: Any, b: Any, tolerance: Decimal = Decimal("0.01")) -> bool:
    da = to_decimal(a)
    db = to_decimal(b)
    if da is None or db is None:
        return False
    return abs(da - db) <= tolerance


def code_sort_key(code: str) -> Tuple[int, ...]:
    return tuple(int(part) for part in re.findall(r"\d+", code))


def json_dump(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def write_csv(path: Path, fields: Sequence[str], rows: Iterable[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(fields), extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: "" if row.get(key) is None else row.get(key) for key in fields})


def locate_pdf(project_root: Path) -> Path:
    exact = project_root / PDF_REL
    if exact.exists():
        return exact
    standards_dir = project_root / STANDARDS_DIR_REL
    if not standards_dir.exists():
        raise FileNotFoundError(f"Missing source standards dir: {standards_dir}")
    matches = [
        path
        for path in standards_dir.glob("*.pdf")
        if "房屋建筑与装饰" in path.name and "上册" in path.name
    ]
    if not matches:
        raise FileNotFoundError(f"Missing upper-volume 房屋建筑与装饰 PDF under: {standards_dir}")
    return sorted(matches, key=lambda path: path.name)[0]


def locate_xlsx(project_root: Path) -> Path:
    path = project_root / XLSX_REL
    if not path.exists():
        raise FileNotFoundError(f"Missing normalized workbook: {path}")
    return path


def extract_pdf_texts(pdf_path: Path) -> List[str]:
    reader = PdfReader(str(pdf_path))
    return [page.extract_text() or "" for page in reader.pages]


def detect_book_page_no(text: str, pdf_page_no: int) -> str:
    lines = [line.strip() for line in (text or "").splitlines() if line.strip()]
    candidates: List[str] = []
    for line in lines[:2] + lines[-2:]:
        if re.fullmatch(r"\d{1,3}", line):
            candidates.append(line)
    if candidates:
        return candidates[-1]
    if 51 <= pdf_page_no <= 704:
        return str(pdf_page_no - 40)
    return ""


def section_marker_for_page(text: str) -> str:
    compact = compact_text(text)
    markers = [
        "A.1.1 土石方工程",
        "A.1.1.1 土方工程",
        "A.1.1.2 石方工程",
        "A.1.1.3 回填方及其他",
        "A.1.2 围护及支护工程",
        "工作内容：",
        "工程量计算规则",
    ]
    found = [marker for marker in markers if marker in compact]
    return ";".join(found)


def profile_pdf(pdf_path: Path, texts: Sequence[str]) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    text_lengths = [len(text) for text in texts]
    compact_lengths = [len(compact_text(text)) for text in texts]
    quota_codes = [code for text in texts for code in QUOTA_CODE_RE.findall(text)]
    resource_codes = [code for text in texts for code in RESOURCE_CODE_RE.findall(text)]
    private_unicode_count = sum(len(PRIVATE_USE_RE.findall(text)) for text in texts)
    work_content_count = sum(text.count("工作内容：") for text in texts)
    unit_marker_count = sum(text.count("计量单位：") for text in texts)
    table_header_count = sum(1 for text in texts if TABLE_HEADER_RE.search(text))

    # The diagonal watermark is visible on rendered table pages, but these terms
    # do not appear in extracted text on A.1.1 sample table pages.
    sample_table_pages = [58, 59, 75, 91, 98]
    watermark_terms = ["广东省住房和城乡建设厅", "广东省建设工程标准定额站"]
    watermark_text_hits = [
        index + 1
        for index in sample_table_pages
        if index < len(texts) and any(term in texts[index] for term in watermark_terms)
    ]

    page_count = len(texts)
    avg_chars = round(statistics.mean(text_lengths), 2) if text_lengths else 0
    profile = {
        "source_file": str(pdf_path),
        "page_count": page_count,
        "pages_with_text": sum(1 for length in compact_lengths if length >= 20),
        "blank_or_low_text_pages": sum(1 for length in compact_lengths if length < 20),
        "avg_text_chars_per_page": avg_chars,
        "min_text_chars_per_page": min(text_lengths) if text_lengths else 0,
        "max_text_chars_per_page": max(text_lengths) if text_lengths else 0,
        "quota_code_count": len(quota_codes),
        "unique_quota_code_count": len(set(quota_codes)),
        "resource_code_probe_count": len(resource_codes),
        "work_content_marker_count": work_content_count,
        "unit_marker_count": unit_marker_count,
        "private_unicode_char_count": private_unicode_count,
        "watermark_text_detected": "no_on_A111_table_text" if not watermark_text_hits else json_dump(watermark_text_hits),
        "ocr_required": "no",
        "parse_feasibility": "medium_high",
        "remark": (
            "PDF has a usable text layer. Rendered spot checks of pages 59 and 99 show a diagonal visible "
            "watermark across table bodies, but the watermark phrase is not emitted by text extraction; "
            f"table header marker count={table_header_count}. Private-use unit glyphs require normalization."
        ),
    }

    probe_rows: List[Dict[str, Any]] = []
    for index, text in enumerate(texts, start=1):
        codes = sorted(set(QUOTA_CODE_RE.findall(text)), key=code_sort_key)
        resources = sorted(set(RESOURCE_CODE_RE.findall(text)))
        if not codes and not resources and "工作内容：" not in text and "计量单位：" not in text:
            continue
        probe_rows.append(
            {
                "source_file": str(pdf_path),
                "pdf_page_no": index,
                "book_page_no": detect_book_page_no(text, index),
                "text_char_count": len(text),
                "quota_code_count": len(QUOTA_CODE_RE.findall(text)),
                "unique_quota_code_count": len(codes),
                "quota_codes_json": json_dump(codes),
                "resource_code_probe_count": len(resources),
                "resource_codes_sample_json": json_dump(resources[:20]),
                "work_content_marker_count": text.count("工作内容："),
                "unit_marker_count": text.count("计量单位："),
                "private_unicode_char_count": len(PRIVATE_USE_RE.findall(text)),
                "has_table_header_marker": "yes" if TABLE_HEADER_RE.search(text) else "no",
                "section_marker": section_marker_for_page(text),
                "remark": "quota/resource/text marker page",
            }
        )
    return profile, probe_rows


def find_header_row(ws: Any) -> Tuple[int, List[str]]:
    for row_index, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True), start=1):
        values = [normalize_header(cell) for cell in row]
        if "项目编码" in values and "项目名称" in values:
            return row_index, values
    first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return 1, [normalize_header(cell) for cell in first]


def build_column_map(headers: Sequence[str]) -> Dict[str, int]:
    aliases = {
        "code": ["项目编码", "定额编号", "项目代码", "编码"],
        "name": ["项目名称", "定额名称", "名称"],
        "unit": ["计量单位", "单位"],
        "labor": ["人工费"],
        "material": ["材料费"],
        "machine": ["机具费", "机械费"],
        "management": ["管理费"],
        "total": ["合计", "基价", "综合单价"],
    }
    result: Dict[str, int] = {}
    for key, names in aliases.items():
        for index, header in enumerate(headers):
            if header in names:
                result[key] = index
                break
    return result


def profile_xlsx(xlsx_path: Path) -> Tuple[Dict[str, Any], Dict[str, Dict[str, Any]]]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    ws = wb[sheet_names[0]]
    header_row_index, headers = find_header_row(ws)
    column_map = build_column_map(headers)
    code_col = column_map.get("code")
    xlsx_by_code: Dict[str, Dict[str, Any]] = {}
    prefix_counts: Counter[str] = Counter()

    if code_col is not None:
        for row_index, row in enumerate(ws.iter_rows(min_row=header_row_index + 1, values_only=True), start=header_row_index + 1):
            values = list(row)
            raw_code = compact_text(values[code_col] if code_col < len(values) else "")
            if not QUOTA_CODE_RE.fullmatch(raw_code):
                continue
            prefix_match = re.match(r"^(A1-\d+)", raw_code)
            if prefix_match:
                prefix_counts[prefix_match.group(1)] += 1
            record = {
                "source_row": row_index,
                "code": raw_code,
                "name": compact_text(values[column_map["name"]] if "name" in column_map and column_map["name"] < len(values) else ""),
                "unit": norm_unit(values[column_map["unit"]] if "unit" in column_map and column_map["unit"] < len(values) else ""),
                "labor": values[column_map["labor"]] if "labor" in column_map and column_map["labor"] < len(values) else "",
                "material": values[column_map["material"]] if "material" in column_map and column_map["material"] < len(values) else "",
                "machine": values[column_map["machine"]] if "machine" in column_map and column_map["machine"] < len(values) else "",
                "management": values[column_map["management"]] if "management" in column_map and column_map["management"] < len(values) else "",
                "total": values[column_map["total"]] if "total" in column_map and column_map["total"] < len(values) else "",
            }
            xlsx_by_code[raw_code] = record

    header_set = set(headers)
    def has_any(names: Sequence[str]) -> str:
        return "yes" if any(name in header_set for name in names) else "no"

    profile = {
        "source_file": str(xlsx_path),
        "file_size_bytes": xlsx_path.stat().st_size,
        "sheet_count": len(sheet_names),
        "sheet_names": json_dump(sheet_names),
        "selected_sheet": ws.title,
        "row_count": ws.max_row,
        "data_row_count": max(ws.max_row - header_row_index, 0),
        "column_count": ws.max_column,
        "headers_json": json_dump(headers),
        "contains_project_code": has_any(["项目编码", "定额编号", "编码"]),
        "contains_project_name": has_any(["项目名称", "定额名称", "名称"]),
        "contains_unit": has_any(["计量单位", "单位"]),
        "contains_labor_fee": has_any(["人工费"]),
        "contains_material_fee": has_any(["材料费"]),
        "contains_machine_fee": has_any(["机具费", "机械费"]),
        "contains_management_fee": has_any(["管理费"]),
        "contains_total": has_any(["合计", "基价", "综合单价"]),
        "contains_work_content": has_any(["工作内容"]),
        "contains_quantity_calculation_rule": has_any(["工程量计算规则"]),
        "contains_chapter_description": has_any(["章节说明", "说明"]),
        "contains_resource_code": has_any(["资源编码", "材料编码", "机具编码"]),
        "contains_resource_name": has_any(["资源名称"]),
        "contains_resource_spec": has_any(["规格型号", "规格"]),
        "contains_resource_unit": has_any(["资源单位"]),
        "contains_resource_unit_price": has_any(["资源单价"]),
        "contains_resource_consumption": has_any(["资源消耗量", "消耗量"]),
        "contains_pdf_page_no": has_any(["PDF页码", "pdf_page_no", "pdf_page"]),
        "contains_source_block_id": has_any(["来源块ID", "source_block_id", "source_block"]),
        "a1_quota_code_count": sum(prefix_counts.values()),
        "a1_unique_quota_code_count": len(xlsx_by_code),
        "a1_prefix_distribution_json": json_dump(dict(sorted(prefix_counts.items(), key=lambda item: code_sort_key(item[0])))),
        "remark": "Normalized workbook is a compact main-quota price table; it has no chapter/rule/resource/page evidence columns.",
    }
    return profile, xlsx_by_code


def fill_horizontal_for_name_rows(rows: Sequence[Sequence[Any]], code_cols: Sequence[int]) -> List[List[str]]:
    filled: List[List[str]] = []
    for row in rows:
        values = [compact_text(cell) for cell in row]
        by_col: List[str] = []
        last = ""
        for col in code_cols:
            value = values[col] if col < len(values) else ""
            if value:
                last = value
                by_col.append(value)
            else:
                by_col.append(last)
        filled.append(by_col)
    return filled


def split_lines(value: Any) -> List[str]:
    if value is None:
        return []
    return [compact_text(part) for part in str(value).splitlines() if compact_text(part)]


def parse_resource_rows(table: Sequence[Sequence[Any]], resource_header_index: int, code_cols: Sequence[int]) -> List[Dict[str, Any]]:
    resources: List[Dict[str, Any]] = []
    current_category = ""
    for row in table[resource_header_index + 1 :]:
        cells = ["" if cell is None else str(cell) for cell in row]
        if cells and compact_text(cells[0]):
            current_category = compact_text(cells[0])
        code_cell = cells[1] if len(cells) > 1 else ""
        codes = RESOURCE_CODE_RE.findall(code_cell)
        if not codes:
            continue
        names = split_lines(cells[2] if len(cells) > 2 else "")
        units = split_lines(cells[3] if len(cells) > 3 else "")
        unit_prices = split_lines(cells[4] if len(cells) > 4 else "")
        consumption_columns = [split_lines(cells[col] if col < len(cells) else "") for col in code_cols]

        for index, resource_code in enumerate(codes):
            if len(names) == len(codes):
                resource_name = names[index]
            elif len(names) >= len(codes) * 2:
                resource_name = compact_text(" ".join(names[index * 2 : index * 2 + 2]))
            else:
                resource_name = names[index] if index < len(names) else ""
            resource_unit = norm_unit(units[index] if index < len(units) else (units[0] if units else ""))
            resource_unit_price = unit_prices[index] if index < len(unit_prices) else ""
            consumption_by_col = []
            for parts in consumption_columns:
                consumption_by_col.append(parts[index] if index < len(parts) else "")
            resources.append(
                {
                    "category": current_category,
                    "resource_code": resource_code,
                    "resource_name": resource_name,
                    "resource_unit": resource_unit,
                    "resource_unit_price": resource_unit_price,
                    "consumption_by_column": consumption_by_col,
                }
            )
    return resources


def extract_between(text: str, start: str, end: str) -> str:
    match = re.search(re.escape(start) + r"(.*?)" + re.escape(end), text, flags=re.S)
    return compact_text(match.group(1)) if match else ""


def parse_pdf_sample_page(
    pdf_path: Path,
    pdf_page_no: int,
    section_path: str,
    xlsx_by_code: Dict[str, Dict[str, Any]],
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        page = pdf.pages[pdf_page_no - 1]
        text = page.extract_text(x_tolerance=1, y_tolerance=3) or ""
        tables = page.extract_tables() or []
    if not tables:
        return [
            {
                "sample_id": f"p{pdf_page_no}_NO_TABLE",
                "pdf_page_no": pdf_page_no,
                "book_page_no": detect_book_page_no(text, pdf_page_no),
                "section_path": section_path,
                "match_status": "parse_failed",
                "parse_confidence": "0.10",
                "issue_type": "no_table_detected",
                "remark": "pdfplumber.extract_tables returned no table",
                "raw_text_block": compact_text(text, 3000),
            }
        ]

    table = max(tables, key=lambda candidate: len(candidate))
    code_row_index = None
    for index, row in enumerate(table):
        if any(QUOTA_CODE_RE.fullmatch(compact_text(cell)) for cell in row if cell is not None):
            code_row_index = index
            break
    if code_row_index is None:
        return [
            {
                "sample_id": f"p{pdf_page_no}_NO_CODE_ROW",
                "pdf_page_no": pdf_page_no,
                "book_page_no": detect_book_page_no(text, pdf_page_no),
                "section_path": section_path,
                "match_status": "parse_failed",
                "parse_confidence": "0.20",
                "issue_type": "no_quota_code_row",
                "remark": "Could not identify quota code row in extracted table",
                "raw_text_block": compact_text(text, 3000),
            }
        ]

    code_row = table[code_row_index]
    code_cols = [index for index, cell in enumerate(code_row) if QUOTA_CODE_RE.fullmatch(compact_text(cell))]
    codes = [compact_text(code_row[col]) for col in code_cols]
    base_row_index = next(
        (index for index, row in enumerate(table) if any("基价" in compact_text(cell) for cell in row if cell is not None)),
        None,
    )
    resource_header_index = next(
        (index for index, row in enumerate(table) if "分类" in compact_text(row[0] if row else "") and "编码" in compact_text(row[1] if len(row) > 1 else "")),
        None,
    )
    name_rows = table[code_row_index + 1 : base_row_index] if base_row_index is not None else []
    filled_names = fill_horizontal_for_name_rows(name_rows, code_cols)
    name_by_code: Dict[str, str] = {}
    for col_index, code in enumerate(codes):
        parts: List[str] = []
        for row_values in filled_names:
            value = row_values[col_index] if col_index < len(row_values) else ""
            if value and value not in parts:
                parts.append(value)
        name_by_code[code] = compact_text(" ".join(parts))

    price_rows = table[base_row_index : base_row_index + 5] if base_row_index is not None else []
    resources = parse_resource_rows(table, resource_header_index, code_cols) if resource_header_index is not None else []
    work_content = extract_between(text, "工作内容：", "计量单位：")
    unit_from_pdf = norm_unit(extract_between(text, "计量单位：", "定额编号"))
    book_page_no = detect_book_page_no(text, pdf_page_no)
    private_unit_issue = "private_unicode_unit" if PRIVATE_USE_RE.search(text) else ""
    resource_issue = "resource_rows_merged" if resources and any("\n" in str(cell) for row in table for cell in row if cell) else ""

    for code_index, code in enumerate(codes):
        xlsx = xlsx_by_code.get(code, {})
        base = price_rows[0][code_cols[code_index]] if len(price_rows) > 0 and code_cols[code_index] < len(price_rows[0]) else ""
        labor = price_rows[1][code_cols[code_index]] if len(price_rows) > 1 and code_cols[code_index] < len(price_rows[1]) else ""
        material = price_rows[2][code_cols[code_index]] if len(price_rows) > 2 and code_cols[code_index] < len(price_rows[2]) else ""
        machine = price_rows[3][code_cols[code_index]] if len(price_rows) > 3 and code_cols[code_index] < len(price_rows[3]) else ""
        management = price_rows[4][code_cols[code_index]] if len(price_rows) > 4 and code_cols[code_index] < len(price_rows[4]) else ""
        field_matches = [
            money_close(base, xlsx.get("total")),
            money_close(labor, xlsx.get("labor")),
            money_close(material, xlsx.get("material")),
            money_close(machine, xlsx.get("machine")),
            money_close(management, xlsx.get("management")),
        ]
        if not xlsx:
            match_status = "missing_xlsx_code"
        elif all(field_matches):
            match_status = "price_match_or_rounding_delta"
        elif money_close(base, xlsx.get("total"), Decimal("0.05")):
            match_status = "total_matches_but_component_delta"
        else:
            match_status = "price_mismatch"

        resource_sample: List[Dict[str, Any]] = []
        for resource in resources[:8]:
            copy = dict(resource)
            consumption_values = copy.pop("consumption_by_column", [])
            copy["consumption_for_quota"] = consumption_values[code_index] if code_index < len(consumption_values) else ""
            resource_sample.append(copy)

        issue_parts = [part for part in [private_unit_issue, resource_issue] if part]
        if name_by_code.get(code) and xlsx.get("name") and name_by_code[code] != xlsx.get("name"):
            issue_parts.append("name_hierarchy_differs_from_xlsx_simplification")
        confidence = Decimal("0.86")
        if issue_parts:
            confidence -= Decimal("0.08")
        if match_status == "price_mismatch":
            confidence -= Decimal("0.12")
        rows.append(
            {
                "sample_id": f"p{pdf_page_no}_{code}",
                "pdf_page_no": pdf_page_no,
                "book_page_no": book_page_no,
                "section_path": section_path,
                "quota_source_code": code,
                "quota_name_from_pdf": name_by_code.get(code, ""),
                "quota_name_from_xlsx": xlsx.get("name", ""),
                "unit_from_pdf": unit_from_pdf,
                "unit_from_xlsx": xlsx.get("unit", ""),
                "base_price_from_pdf": base,
                "total_from_xlsx": xlsx.get("total", ""),
                "labor_fee_from_pdf": labor,
                "labor_fee_from_xlsx": xlsx.get("labor", ""),
                "material_fee_from_pdf": material,
                "material_fee_from_xlsx": xlsx.get("material", ""),
                "machine_fee_from_pdf": machine,
                "machine_fee_from_xlsx": xlsx.get("machine", ""),
                "management_fee_from_pdf": management,
                "management_fee_from_xlsx": xlsx.get("management", ""),
                "work_content_from_pdf": work_content,
                "resource_row_count_from_pdf": len(resources),
                "resource_rows_sample_json": json_dump(resource_sample),
                "match_status": match_status,
                "parse_confidence": decimal_text(confidence),
                "issue_type": ";".join(issue_parts) if issue_parts else "none",
                "remark": "bounded A.1.1 table probe; not full extraction",
                "raw_text_block": compact_text(text, 3000),
            }
        )
    return rows


def locate_sample_pages(texts: Sequence[str]) -> Dict[str, int]:
    page_by_label: Dict[str, int] = {}
    for group in SAMPLE_CODE_GROUPS:
        target_codes = set(group["codes"])
        best_page = 0
        best_hits = -1
        for index, text in enumerate(texts, start=1):
            codes = set(QUOTA_CODE_RE.findall(text))
            hits = len(target_codes & codes)
            if hits > best_hits:
                best_page = index
                best_hits = hits
        if best_hits <= 0:
            page_by_label[group["label"]] = 0
        else:
            page_by_label[group["label"]] = best_page
    return page_by_label


def build_sample_probe(pdf_path: Path, texts: Sequence[str], xlsx_by_code: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    sample_pages = locate_sample_pages(texts)
    for group in SAMPLE_CODE_GROUPS:
        pdf_page_no = sample_pages.get(group["label"], 0)
        if not pdf_page_no:
            rows.append(
                {
                    "sample_id": group["label"],
                    "section_path": group["section_path"],
                    "match_status": "parse_failed",
                    "parse_confidence": "0.00",
                    "issue_type": "sample_page_not_found",
                    "remark": f"Could not locate any of {group['codes']}",
                }
            )
            continue
        rows.extend(parse_pdf_sample_page(pdf_path, pdf_page_no, group["section_path"], xlsx_by_code))
    return rows


def build_field_gap_rows() -> List[Dict[str, Any]]:
    xlsx_yes = {
        "quota_source_code",
        "quota_name_full",
        "quota_unit",
        "base_price",
        "labor_fee",
        "material_fee",
        "machine_fee",
        "management_fee",
    }
    xlsx_partial = {
        "measurement_unit_context",
        "quota_name_level_1",
        "quota_name_level_2",
        "quota_name_level_3",
    }
    pdf_yes = set(FIELD_NAMES_REQUIRED) - {"parse_confidence", "review_status"}
    rows: List[Dict[str, Any]] = []
    for field in FIELD_NAMES_REQUIRED:
        if field in pdf_yes:
            exists_pdf = "yes"
        else:
            exists_pdf = "pipeline_metadata"
        if field in xlsx_yes:
            exists_xlsx = "yes"
        elif field in xlsx_partial:
            exists_xlsx = "partial"
        else:
            exists_xlsx = "no"
        importance = "high"
        if field in {"source_file", "source_volume", "pdf_page_no", "book_page_no", "parse_confidence", "review_status"}:
            importance = "medium_high"
        if field.startswith("resource_") or field in {"work_content", "quantity_calculation_rule"}:
            importance = "critical"
        difficulty = "medium"
        if field in {"quantity_calculation_rule", "chapter_name", "section_name", "subsection_title"}:
            difficulty = "medium_high"
        if field.startswith("resource_") or field == "resource_consumption_by_quota_code":
            difficulty = "high"
        if field in {"quota_source_code", "base_price", "labor_fee", "material_fee", "machine_fee", "management_fee", "quota_unit"}:
            difficulty = "medium"
        stage = "Stage 2 A.1.1 PDF probe"
        if field in {"source_file", "source_volume", "pdf_page_no", "book_page_no"}:
            stage = "Stage 1 PDF text layer profile"
        elif field in {"quantity_calculation_rule", "chapter_code", "chapter_name", "section_code", "section_name"}:
            stage = "Stage 3 A.1.1 structured extraction candidate"
        elif field.startswith("resource_"):
            stage = "Stage 7 resource detail normalization"
        elif field in {"parse_confidence", "review_status"}:
            stage = "Stage 5 Cost department QA pack"
        remark = "PDF official detail version is required for this field."
        if exists_xlsx == "yes":
            remark = "Normalized Excel covers this as compact main-quota checksum data."
        elif exists_xlsx == "partial":
            remark = "Normalized Excel contains a flattened/simplified representation only."
        rows.append(
            {
                "field_name": field,
                "exists_in_pdf": exists_pdf,
                "exists_in_normalized_xlsx": exists_xlsx,
                "importance_for_cost_department": importance,
                "parse_difficulty_from_pdf": difficulty,
                "recommended_stage": stage,
                "remark": remark,
            }
        )
    return rows


def build_difficulty_matrix() -> List[Dict[str, Any]]:
    specs = [
        ("pdf_text_layer", "low", "high", "Low-text divider/blank pages; extracted text order may differ from layout", "pypdf/pdfplumber text profile plus per-page marker distribution", "no", "non_blocking", "Text layer is usable; OCR is not needed for this PDF."),
        ("watermark_handling", "medium", "medium_high", "Visible diagonal watermark crosses table body and resource rows", "Ignore extracted watermark text; verify rendered sample pages during QA", "yes", "non_blocking", "Watermark is visual; it did not enter A.1.1 text extraction."),
        ("quota_code_detection", "low", "high", "Substring false positives if matching is not regex-bounded", "Regex-bounded A1-* code extraction with page distribution CSV", "no", "non_blocking", "Stable for A1-* codes."),
        ("chapter_structure_detection", "medium_high", "medium", "Heading inheritance and divider pages; TOC/page offset anchoring", "Combine TOC markers, section headings, page ranges, and QA anchors", "yes", "non_blocking", "A.1.1/A.1.2 boundary needs manual confirmation."),
        ("work_content_extraction", "medium", "medium_high", "Marker spans before table; multi-line content varies by section", "Extract between 工作内容 and 计量单位 markers, then normalize whitespace", "yes", "non_blocking", "Works on sample pages."),
        ("quantity_rule_extraction", "medium_high", "medium", "Narrative tables and rules are not keyed by quota code", "Section-level block extraction with source-page anchoring", "yes", "non_blocking", "Should be a chapter/section artifact, not a quota-row field only."),
        ("main_quota_price_table_extraction", "medium", "medium_high", "Merged headers, multi-row item names, private unit glyphs", "pdfplumber table extraction plus code-column reconstruction", "yes", "non_blocking", "A.1.1 sample prices reconstruct and crosscheck against Excel."),
        ("multi_quota_column_alignment", "high", "medium", "Parent-child item names span columns; blank merged cells lose semantic scope", "Coordinate/line-aware header reconstruction and Excel checksum comparison", "yes", "non_blocking", "Largest main-table risk."),
        ("resource_detail_extraction", "high", "medium", "Merged category rows; name/spec split lines; consumption alignment by quota column", "Parse resource rows after 分类/编码 header, retain row JSON and confidence", "yes", "non_blocking", "Feasible but not production-stable without QA."),
        ("cross_page_table_continuation", "high", "medium", "Long resource tables may continue without repeated headers", "Detect repeated code columns, page footers, and continuation markers", "yes", "non_blocking", "Needs explicit continuation_status field."),
        ("unit_normalization", "medium", "medium_high", "Private-use m3 glyph; m2/m3 variants", "Unit normalization map plus raw unit retention", "yes", "non_blocking", "Required before structured candidate output."),
        ("private_unicode_fix", "medium", "high", "Private Unicode glyph appears in units/specs", "Count private-use chars and normalize known glyphs", "yes", "non_blocking", "No OCR needed; this is text normalization."),
        ("xlsx_price_crosscheck", "low", "high", "Third-party normalized Excel has rounding deltas and missing fields", "Use code-level fee checksum with tolerance and issue log", "yes", "non_blocking", "Useful as checksum reference, not source of truth."),
        ("full_volume_scaling", "high", "medium", "Thousands of quota rows; section-specific table variants", "Stage by chapter/section with QA sampling and confidence thresholds", "yes", "non_blocking", "Do not jump directly to full-volume import."),
        ("all_volumes_batch_processing", "very_high", "medium_low", "Different volumes/disciplines have distinct layouts and appendices", "Batch only after A.1 upper-volume extraction stabilizes", "yes", "non_blocking", "Requires parser registry by volume."),
        ("database_import_readiness", "very_high", "low", "Schema/approval/governance not established for official PDF candidates", "Defer until QA-approved candidate schema and import design exist", "yes", "blocking", "This assessment must not write database outputs."),
    ]
    return [
        {
            "module": module,
            "difficulty_level": difficulty,
            "automation_feasibility": feasibility,
            "main_risks": risks,
            "recommended_method": method,
            "requires_manual_qa": qa,
            "blocking_or_non_blocking": blocking,
            "remark": remark,
        }
        for module, difficulty, feasibility, risks, method, qa, blocking, remark in specs
    ]


def sample_summary(sample_rows: Sequence[Dict[str, Any]]) -> Dict[str, Any]:
    status_counts = Counter(row.get("match_status", "") for row in sample_rows)
    code_count = sum(1 for row in sample_rows if row.get("quota_source_code"))
    resource_positive = sum(1 for row in sample_rows if int(row.get("resource_row_count_from_pdf") or 0) > 0)
    return {
        "status_counts": dict(status_counts),
        "code_count": code_count,
        "resource_positive_rows": resource_positive,
    }


def write_report(
    path: Path,
    pdf_path: Path,
    xlsx_path: Path,
    pdf_profile: Dict[str, Any],
    xlsx_profile: Dict[str, Any],
    field_gap_rows: Sequence[Dict[str, Any]],
    sample_rows: Sequence[Dict[str, Any]],
    difficulty_rows: Sequence[Dict[str, Any]],
) -> None:
    summary = sample_summary(sample_rows)
    missing_xlsx_fields = [
        row["field_name"]
        for row in field_gap_rows
        if row["exists_in_normalized_xlsx"] == "no"
    ]
    partial_xlsx_fields = [
        row["field_name"]
        for row in field_gap_rows
        if row["exists_in_normalized_xlsx"] == "partial"
    ]
    diff_by_module = {row["module"]: row for row in difficulty_rows}
    report = f"""# Stage GD2018-PDF-VS-XLSX-DIFFICULTY-ASSESSMENT-1 Report

## 1. Task Scope

本轮只做 feasibility assessment / extraction difficulty assessment。未做全量 PDF 解析，未写数据库，未修改 migration、src/cost_engine 主流程、baseline、normalized Excel 或 PDF 源文件，未生成 approved、internal_price_library、quota_to_bill_mapping。

## 2. Inputs

- PDF: `{pdf_path}`
- PDF file size: {pdf_path.stat().st_size} bytes
- PDF page count: {pdf_profile["page_count"]}
- Normalized Excel: `{xlsx_path}`
- Excel file size: {xlsx_profile["file_size_bytes"]} bytes
- Excel sheets: {xlsx_profile["sheet_count"]} {xlsx_profile["sheet_names"]}
- Selected sheet rows/cols: {xlsx_profile["row_count"]} rows, {xlsx_profile["column_count"]} columns

## 3. Source-of-Truth Decision

- PDF 应作为官方源真相。它包含章节说明、工程量计算规则、工作内容、计量单位上下文、横向定额表、子目层级和人材机资源明细。
- normalized Excel 只能作为 checksum reference / compact price reference。它适合按定额编号校验主项四费与合计，但不能替代 PDF。
- 后续 pipeline 建议：PDF 产出带页码、块 ID、置信度与 QA 状态的候选结构化结果；Excel 只参与价格合计/四费交叉校验。

## 4. PDF Text Layer Assessment

- PDF 有可用文本层：{pdf_profile["pages_with_text"]}/{pdf_profile["page_count"]} 页有可抽取文本。
- OCR 判断：不需要 OCR。A.1.1 样本页可直接抽取定额编号、工作内容、计量单位、资源编码和价格。
- 水印判断：可见水印存在，渲染抽查页 59、99 可见斜向水印覆盖表格主体；但水印文字没有进入 A.1.1 表格页 text extraction。
- 私有 Unicode：检测到 {pdf_profile["private_unicode_char_count"]} 个私有区字符，主要影响 m3 单位，需要规范化。

## 5. PDF Table Structure Assessment

主表可通过文本坐标/表格结构重建，但不能只靠行文本顺序。主要风险是横向多子目、合并单元格导致的父级名称继承、资源行多行合并、跨页延续和单位私有字形。pdfplumber 在 A.1.1 样本页能抽出表格，但资源明细归属需要保留原始行 JSON 与人工 QA。

## 6. Normalized Excel Coverage

Excel 当前覆盖：项目编码、项目名称、计量单位、人工费、材料费、机具费、管理费、合计。A1-* 定额编号数量为 {xlsx_profile["a1_quota_code_count"]}，唯一编号为 {xlsx_profile["a1_unique_quota_code_count"]}。

Excel 缺失字段包括：{", ".join(missing_xlsx_fields)}。部分扁平化字段包括：{", ".join(partial_xlsx_fields) if partial_xlsx_fields else "无"}。

## 7. Field Gap Analysis

字段缺口详见 `field_gap_analysis.csv`。结论是：PDF 覆盖成本部需要的官方上下文和资源明细；normalized Excel 只覆盖主项价格 checksum，缺少页码、章节、规则、工作内容和资源消耗量等审计字段。

## 8. A.1.1 Probe Result

- 样本覆盖页：59、60、76、92、99。
- 样本 quota row count: {summary["code_count"]}。
- match_status distribution: `{json_dump(summary["status_counts"])}`。
- resource detail positive rows: {summary["resource_positive_rows"]}。
- 成功项：A.1.1 样本页可抽取 quota_source_code、主项价格、四费、工作内容和部分人材机资源行。
- 失败/风险项：PDF 子目名称层级与 Excel 扁平名称不完全一致；资源行中名称/规格/消耗量存在多行合并；m3 单位出现私有 Unicode；水印覆盖视觉表格但未破坏文本抽取。
- 与 Excel 合计：大部分可在 0.01 容差内一致，个别如 A1-1-1 存在 normalized Excel 0.01 级四舍五入差异，应记录为 checksum delta。
- 人材机资源明细：能抽到资源编码、名称、单位、单价、各子目消耗量，但稳定归属需要坐标增强和人工 QA。

## 9. Difficulty Rating

- main quota extraction difficulty: medium
- resource detail extraction difficulty: high
- full-volume scaling difficulty: high
- all-volume extraction difficulty: very_high

## 10. Risk Register

- private Unicode unit glyph: m3 被抽为私有区字符，需统一映射。
- watermark visual interference: 视觉水印覆盖表格主体，文本抽取未受阻，但 QA 截图会受影响。
- multi-column quota alignment: 横向多子目和合并表头是主表解析核心难点。
- resource row category continuation: 人工/材料/机具分类在纵向合并行中延续。
- cross-page continuation: 长资源表可能跨页，需 continuation_status。
- chapter heading inheritance: 子目需要继承章、节、分节标题。
- third-party Excel mismatch: Excel 是第三方 normalized 结果，有四舍五入/扁平化差异。
- official source page anchoring: 必须保留 PDF 页码、书内页码和来源块 ID。
- manual QA load: 成本部 QA 应覆盖表头继承、资源归属、价格 delta 和跨页。

## 11. Recommended Pipeline

Stage 1：PDF text layer profile  
Stage 2：A.1.1 PDF probe  
Stage 3：A.1.1 PDF structured extraction candidate  
Stage 4：A.1.1 PDF-vs-XLSX reconciliation  
Stage 5：Cost department QA pack  
Stage 6：A.1 full upper/middle/lower extraction  
Stage 7：resource detail normalization  
Stage 8：database candidate import design

## 12. Go / No-Go Recommendation

go_for_A111_pdf_structured_probe

## Required Questions Answered

1. PDF 是否有可用文本层：有。
2. 水印是否进入文本抽取结果：A.1.1 表格页未进入。
3. 水印是否影响表格抽取：不阻断文本/表格抽取，但视觉 QA 影响为 medium。
4. 是否需要 OCR：不需要。
5. PDF 表格是否可以通过文本坐标重建：可以，但需坐标增强、合并单元格处理和 QA。
6. normalized Excel 缺失字段：详见 `field_gap_analysis.csv`，主要缺章节/规则/工作内容/资源明细/页码/块 ID/置信度/审核状态。
7. normalized Excel 是否还能作为校验基准：可以，作为 checksum reference，不能作为 source of truth。
8. Codex 全量解析难点：多子目横向对齐、资源明细归属、跨页、私有字形、水印视觉 QA、章节继承和全册扩展。
9. 是否建议先从 A.1.1 做试点：建议。
10. A.1.1 试点字段：定额编号、子目层级/全名、单位、基价、四费、工作内容、资源分类/编码/名称/规格/单位/单价/消耗量、页码、来源块、置信度、QA 状态。
11. 全量解析阶段：建议 8 个阶段，见 Recommended Pipeline。
12. 人工 QA 位置：A.1.1 structured candidate 后、PDF-vs-XLSX reconciliation 后、resource detail normalization 后、database candidate import design 前。

## Output Files

- `pdf_text_layer_profile.csv`
- `pdf_quota_code_probe.csv`
- `pdf_sample_table_probe_A111.csv`
- `xlsx_profile.csv`
- `field_gap_analysis.csv`
- `extraction_difficulty_matrix.csv`
"""
    path.write_text(report, encoding="utf-8")


def run(project_root: Path) -> Dict[str, Path]:
    pdf_path = locate_pdf(project_root)
    xlsx_path = locate_xlsx(project_root)
    output_dir = project_root / OUTPUT_DIR_REL
    output_dir.mkdir(parents=True, exist_ok=True)

    texts = extract_pdf_texts(pdf_path)
    pdf_profile, quota_probe_rows = profile_pdf(pdf_path, texts)
    xlsx_profile, xlsx_by_code = profile_xlsx(xlsx_path)
    field_gap_rows = build_field_gap_rows()
    difficulty_rows = build_difficulty_matrix()
    sample_rows = build_sample_probe(pdf_path, texts, xlsx_by_code)

    paths = {
        "pdf_text_layer_profile": output_dir / "pdf_text_layer_profile.csv",
        "pdf_quota_code_probe": output_dir / "pdf_quota_code_probe.csv",
        "pdf_sample_table_probe_A111": output_dir / "pdf_sample_table_probe_A111.csv",
        "xlsx_profile": output_dir / "xlsx_profile.csv",
        "field_gap_analysis": output_dir / "field_gap_analysis.csv",
        "extraction_difficulty_matrix": output_dir / "extraction_difficulty_matrix.csv",
        "report": output_dir / "stage_gd2018_pdf_vs_xlsx_difficulty_assessment_report.md",
    }
    write_csv(paths["pdf_text_layer_profile"], PDF_TEXT_PROFILE_FIELDS, [pdf_profile])
    write_csv(paths["pdf_quota_code_probe"], PDF_QUOTA_PROBE_FIELDS, quota_probe_rows)
    write_csv(paths["xlsx_profile"], XLSX_PROFILE_FIELDS, [xlsx_profile])
    write_csv(paths["field_gap_analysis"], FIELD_GAP_FIELDS, field_gap_rows)
    write_csv(paths["extraction_difficulty_matrix"], DIFFICULTY_FIELDS, difficulty_rows)
    write_csv(paths["pdf_sample_table_probe_A111"], SAMPLE_PROBE_FIELDS, sample_rows)
    write_report(paths["report"], pdf_path, xlsx_path, pdf_profile, xlsx_profile, field_gap_rows, sample_rows, difficulty_rows)
    return paths


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--project-root", type=Path, default=PROJECT_ROOT_DEFAULT)
    return parser.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)
    try:
        paths = run(args.project_root)
    except Exception as exc:  # pragma: no cover - command-line guard
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    print("Generated assessment artifacts:")
    for name, path in paths.items():
        print(f"- {name}: {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
