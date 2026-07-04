#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Stage GD2018-A111-FULL extraction.

Extracts A.1.1 quota reference candidates from the third-party Guangdong 2018
Excel workbook. The workbook is treated as structured reference input only.

This script does not write databases, migrations, cost_items,
knowledge_review_records, internal_price_library, quota_to_bill_mapping,
approved data, or bill_code mappings. It does not modify the source Excel.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook


PROJECT_ROOT_DEFAULT = Path(r"E:\workspace\01_Projects\ai-construction-system")
ENGINE_REL = Path("construction_cost_knowledge_engine")
EXCEL_REL = ENGINE_REL / "data" / "private" / "reference_extraction" / "source_excels" / "广东省房屋建筑与装饰工程综合定额（2018 ）.xlsx"
OUTPUT_DIR_REL = ENGINE_REL / "data" / "private" / "reference_extraction" / "runs" / "GD2018_stage2R_A111_full"

SOURCE_TYPE = "reference_excel_third_party"
SOURCE_NAME = "广东省房屋建筑与装饰工程综合定额2018"
PRICING_SOURCE = "广东省房屋建筑与装饰工程综合定额2018第三方Excel"
PRICING_VERSION = "GD2018_reference_excel"
SOURCE_TRUST_LEVEL = "L1"
VERIFICATION_STATUS = "structure_checked"
REVIEW_STATUS = "pending"
CHAPTER_CODE = "A.1.1"
CHAPTER_NAME = "土石方工程"
OFFICIAL_SOURCE_FILE = "1. 广东省房屋建筑与装饰工程定额20190112(上册).pdf"
OFFICIAL_SOURCE_PAGE_EVIDENCE_TYPE = "pending"

EXPECTED_HEADERS = [
    "序号",
    "定额编号",
    "项目名称",
    "规格型号",
    "计量单位",
    "主材系数",
    "工程数量",
    "主材单价（元）",
    "人工费",
    "材料费",
    "机具费",
    "管理费",
    "合计",
    "主材合价（元）",
]

REQUIRED_CODES = ["A1-1-1", "A1-1-67", "A1-1-126", "A1-1-137"]
REQUIRED_SUPPLEMENTAL_CODES = [
    "A1-1-56-1",
    "A1-1-56-2",
    "A1-1-56-3",
    "A1-1-56-4",
    "A1-1-118-1",
    "A1-1-118-2",
]

RAW_FIELDS = [
    "raw_row_id",
    "source_type",
    "source_name",
    "extraction_source_file",
    "extraction_source_file_hash",
    "extraction_source_sheet",
    "extraction_source_row",
    "raw_source_code",
    "raw_name",
    "raw_spec_model",
    "raw_unit",
    "raw_main_material_factor",
    "raw_quantity",
    "raw_main_material_price",
    "raw_labor_fee",
    "raw_material_fee",
    "raw_machine_fee",
    "raw_management_fee",
    "raw_total_fee",
    "raw_main_material_total",
    "parse_issue",
    "remark",
]

CANDIDATE_FIELDS = [
    "reference_id",
    "source_type",
    "source_name",
    "extraction_source_file",
    "extraction_source_file_hash",
    "extraction_source_sheet",
    "extraction_source_row",
    "official_source_file",
    "official_source_page",
    "official_source_page_evidence_type",
    "chapter_code",
    "chapter_name",
    "section_code",
    "section_name",
    "item_group_name",
    "source_code",
    "raw_name",
    "standard_name_candidate",
    "unit",
    "keywords",
    "aliases",
    "feature_template",
    "extraction_confidence",
    "source_trust_level",
    "verification_status",
    "review_status",
    "reviewer",
    "remark",
]

PRICING_FIELDS = [
    "pricing_snapshot_id",
    "reference_id",
    "source_code",
    "raw_name",
    "unit",
    "labor_fee",
    "material_fee",
    "machine_fee",
    "management_fee",
    "total_fee",
    "pricing_source",
    "pricing_version",
    "source_trust_level",
    "review_status",
    "remark",
]

ISSUE_FIELDS = [
    "issue_id",
    "source_code",
    "raw_name",
    "issue_type",
    "issue_detail",
    "severity",
    "suggested_action",
]


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def norm_header(value: Any) -> str:
    return text(value).replace("\ufeff", "")


def is_source_code_format(code: str) -> bool:
    return bool(re.fullmatch(r"A1-1-\d+(?:-\d+)?", text(code)))


def code_base_number(code: str) -> int:
    match = re.fullmatch(r"A1-1-(\d+)(?:-\d+)?", text(code))
    return int(match.group(1)) if match else -1


def is_a111_code(code: str) -> bool:
    return is_source_code_format(code) and 1 <= code_base_number(code) <= 137


def is_supplemental_code(code: str) -> bool:
    return bool(re.fullmatch(r"A1-1-\d+-\d+", text(code)))


def natural_code_key(code: str) -> Tuple[int, int]:
    match = re.fullmatch(r"A1-1-(\d+)(?:-(\d+))?", text(code))
    if not match:
        return (10_000, 10_000)
    base = int(match.group(1))
    supplement = int(match.group(2) or 0)
    return (base, supplement)


def section_for_code(code: str) -> Tuple[str, str]:
    base = code_base_number(code)
    if 1 <= base <= 66:
        return "A.1.1.1", "土方工程"
    if 67 <= base <= 125:
        return "A.1.1.2", "石方工程"
    if 126 <= base <= 137:
        return "A.1.1.3", "回填方及其他"
    return "", ""


def light_name(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def item_group_from_name(name: str) -> str:
    cleaned = light_name(name)
    if not cleaned:
        return ""
    split_tokens = [
        " 一、二类土",
        " 三类土",
        " 四类土",
        " 极软岩",
        " 软岩",
        " 较软岩",
        " 较硬岩",
        " 坚硬岩",
        " 运距",
        " 增运",
    ]
    for token in split_tokens:
        if token in cleaned:
            return cleaned.split(token, 1)[0]
    return cleaned


def keywords_from_name(name: str) -> str:
    tokens = [token for token in re.split(r"[\s、，；;（）()《》\"“”]+", name or "") if len(token) >= 2]
    seen = set()
    result = []
    for token in tokens:
        if token not in seen:
            seen.add(token)
            result.append(token)
    return ";".join(result[:10])


def has_mojibake(values: Iterable[Any]) -> bool:
    patterns = ("�", "Ã", "Â", "浣", "鐨", "乱码")
    return any(any(pattern in text(value) for pattern in patterns) for value in values)


def pricing_columns(headers: Sequence[str]) -> List[str]:
    pricing_words = ("人工费", "材料费", "机具费", "管理费", "合计", "单价", "合价")
    return [header for header in headers if any(word in header for word in pricing_words)]


def bill_code_fields(headers: Sequence[str]) -> List[str]:
    return [header for header in headers if "清单" in header or ("编码" in header and header != "定额编号")]


def row_dict(headers: Sequence[str], row: Sequence[Any]) -> Dict[str, Any]:
    return {headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))}


def value(row: Dict[str, Any], field: str) -> Any:
    return row.get(field) if row.get(field) is not None else ""


def collect_workbook_profile(excel_path: Path, workbook: Any, target_sheet: Any, headers: Sequence[str], a111_rows: Sequence[Dict[str, Any]]) -> Dict[str, Any]:
    sample_values: List[Any] = list(headers)
    for row in target_sheet.iter_rows(min_row=2, max_row=min(target_sheet.max_row, 30), values_only=True):
        sample_values.extend(row)
    formula_cells = 0
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    formula_cells += 1
    merged_ranges = sum(len(ws.merged_cells.ranges) for ws in workbook.worksheets)
    return {
        "file_exists": excel_path.exists(),
        "file_size_bytes": excel_path.stat().st_size if excel_path.exists() else "",
        "sha256": sha256_file(excel_path),
        "sheet_names": [ws.title for ws in workbook.worksheets],
        "sheet_count": len(workbook.worksheets),
        "target_sheet": target_sheet.title,
        "target_sheet_rows": target_sheet.max_row,
        "target_sheet_columns": target_sheet.max_column,
        "headers": list(headers),
        "hidden_sheet_count": sum(1 for ws in workbook.worksheets if ws.sheet_state != "visible"),
        "formula_cell_count": formula_cells,
        "merged_range_count": merged_ranges,
        "has_mojibake": has_mojibake(sample_values),
        "bill_code_fields": bill_code_fields(headers),
        "pricing_columns": pricing_columns(headers),
        "a111_source_code_count": len(a111_rows),
    }


def collect_a111_rows(ws: Any, headers: Sequence[str]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    code_col = headers.index("定额编号")
    rows: List[Dict[str, Any]] = []
    invalid_a11_rows: List[Dict[str, Any]] = []
    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        code = text(row[code_col] if code_col < len(row) else "")
        if is_a111_code(code):
            data = row_dict(headers, row)
            data["_row_no"] = row_no
            data["_source_code"] = code
            rows.append(data)
        elif code.startswith("A1-1-") and not is_a111_code(code):
            data = row_dict(headers, row)
            data["_row_no"] = row_no
            data["_source_code"] = code
            invalid_a11_rows.append(data)
    rows.sort(key=lambda r: natural_code_key(r["_source_code"]))
    return rows, invalid_a11_rows


def build_raw_rows(excel_path: Path, excel_hash: str, sheet_name: str, rows: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    output: List[Dict[str, Any]] = []
    for idx, row in enumerate(rows, start=1):
        code = text(row.get("_source_code"))
        parse_issues: List[str] = []
        if is_supplemental_code(code):
            parse_issues.append("supplemental_source_code")
        output.append(
            {
                "raw_row_id": f"GD2018_A111_RAW_{idx:03d}",
                "source_type": SOURCE_TYPE,
                "source_name": SOURCE_NAME,
                "extraction_source_file": excel_path.name,
                "extraction_source_file_hash": excel_hash,
                "extraction_source_sheet": sheet_name,
                "extraction_source_row": row["_row_no"],
                "raw_source_code": code,
                "raw_name": text(row.get("项目名称")),
                "raw_spec_model": text(row.get("规格型号")),
                "raw_unit": text(row.get("计量单位")),
                "raw_main_material_factor": value(row, "主材系数"),
                "raw_quantity": value(row, "工程数量"),
                "raw_main_material_price": value(row, "主材单价（元）"),
                "raw_labor_fee": value(row, "人工费"),
                "raw_material_fee": value(row, "材料费"),
                "raw_machine_fee": value(row, "机具费"),
                "raw_management_fee": value(row, "管理费"),
                "raw_total_fee": value(row, "合计"),
                "raw_main_material_total": value(row, "主材合价（元）"),
                "parse_issue": ";".join(parse_issues),
                "remark": "third_party_excel_raw_reference_row;not_official_truth",
            }
        )
    return output


def build_candidates(excel_path: Path, excel_hash: str, sheet_name: str, rows: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    candidates: List[Dict[str, Any]] = []
    for row in rows:
        code = text(row.get("_source_code"))
        section_code, section_name = section_for_code(code)
        raw_name = text(row.get("项目名称"))
        standard_name = light_name(raw_name)
        remark_parts = ["third_party_excel_reference_candidate", "official_pdf_page_pending"]
        if is_supplemental_code(code):
            remark_parts.append("supplemental_source_code")
        candidates.append(
            {
                "reference_id": f"GD2018_A111_EXCEL_{code}",
                "source_type": SOURCE_TYPE,
                "source_name": SOURCE_NAME,
                "extraction_source_file": excel_path.name,
                "extraction_source_file_hash": excel_hash,
                "extraction_source_sheet": sheet_name,
                "extraction_source_row": row["_row_no"],
                "official_source_file": OFFICIAL_SOURCE_FILE,
                "official_source_page": "",
                "official_source_page_evidence_type": OFFICIAL_SOURCE_PAGE_EVIDENCE_TYPE,
                "chapter_code": CHAPTER_CODE,
                "chapter_name": CHAPTER_NAME,
                "section_code": section_code,
                "section_name": section_name,
                "item_group_name": item_group_from_name(raw_name),
                "source_code": code,
                "raw_name": raw_name,
                "standard_name_candidate": standard_name,
                "unit": text(row.get("计量单位")),
                "keywords": keywords_from_name(standard_name),
                "aliases": "",
                "feature_template": "",
                "extraction_confidence": "0.86" if not is_supplemental_code(code) else "0.80",
                "source_trust_level": SOURCE_TRUST_LEVEL,
                "verification_status": VERIFICATION_STATUS,
                "review_status": REVIEW_STATUS,
                "reviewer": "",
                "remark": ";".join(remark_parts),
            }
        )
    return candidates


def build_pricing_snapshot(rows: Sequence[Dict[str, Any]], candidates: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows_by_code = {text(row.get("_source_code")): row for row in rows}
    output: List[Dict[str, Any]] = []
    for candidate in candidates:
        code = candidate["source_code"]
        row = rows_by_code[code]
        reference_id = candidate["reference_id"]
        output.append(
            {
                "pricing_snapshot_id": f"GD2018_A111_PRICE_{code}",
                "reference_id": reference_id,
                "source_code": code,
                "raw_name": candidate["raw_name"],
                "unit": candidate["unit"],
                "labor_fee": value(row, "人工费"),
                "material_fee": value(row, "材料费"),
                "machine_fee": value(row, "机具费"),
                "management_fee": value(row, "管理费"),
                "total_fee": value(row, "合计"),
                "pricing_source": PRICING_SOURCE,
                "pricing_version": PRICING_VERSION,
                "source_trust_level": SOURCE_TRUST_LEVEL,
                "review_status": REVIEW_STATUS,
                "remark": "reference_pricing_snapshot_only;do_not_generate_internal_price_library",
            }
        )
    return output


def build_issues(headers: Sequence[str], rows: Sequence[Dict[str, Any]], invalid_a11_rows: Sequence[Dict[str, Any]], candidates: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    issues: List[Dict[str, Any]] = []

    def add(source_code: str, raw_name: str, issue_type: str, detail: str, severity: str, action: str) -> None:
        issues.append(
            {
                "issue_id": f"ISSUE_GD2018_A111_{len(issues) + 1:03d}",
                "source_code": source_code,
                "raw_name": raw_name,
                "issue_type": issue_type,
                "issue_detail": detail,
                "severity": severity,
                "suggested_action": action,
            }
        )

    code_counts = Counter(text(row.get("_source_code")) for row in rows)
    for row in rows:
        code = text(row.get("_source_code"))
        raw_name = text(row.get("项目名称"))
        unit = text(row.get("计量单位"))
        if not code:
            add("", raw_name, "missing_source_code", "A.1.1 candidate row has no source_code.", "high", "Exclude until source_code is resolved.")
        elif not is_source_code_format(code):
            add(code, raw_name, "invalid_source_code", "source_code is not A1-1-* or A1-1-*-*.", "high", "Do not use until corrected.")
        if not raw_name:
            add(code, raw_name, "missing_name", "A.1.1 candidate row has no raw_name/project name.", "high", "Verify against source Excel/PDF.")
        if not unit:
            add(code, raw_name, "missing_unit", "A.1.1 candidate row has no unit.", "high", "Verify against source Excel/PDF.")
        if code_counts[code] > 1:
            add(code, raw_name, "duplicated_source_code", "source_code appears more than once in A.1.1 extraction.", "high", "Deduplicate or explain variant rows.")
        if is_supplemental_code(code):
            add(code, raw_name, "supplemental_source_code", "Supplemental A.1.1 source_code detected.", "medium", "Keep as pending candidate and verify manually; this is not a blocker.")

    for row in invalid_a11_rows:
        code = text(row.get("_source_code"))
        if code:
            add(code, text(row.get("项目名称")), "invalid_source_code", "Row starts with A1-1 but is outside A.1.1 accepted range or format.", "high", "Exclude from A.1.1 mapping pilot until classified.")

    if len(rows) != 143:
        add("", "", "unexpected_candidate_count", f"Expected approximately 143 A.1.1 rows, extracted {len(rows)}.", "high", "Do not hard-code count; review Excel filter before mapping.")

    add("", "", "non_official_source", "The source workbook is a third-party Excel, not the official PDF.", "medium", "Treat all rows as reference candidates requiring human review.")
    add("", "", "official_pdf_page_pending", "official_source_page is pending for candidate rows.", "medium", "Verify against official PDF pages before stronger trust status.")
    if pricing_columns(headers):
        add("", "", "pricing_column_present", "Pricing columns are present and exported only as reference snapshot.", "low", "Do not write internal_price_library from this output.")

    return issues


def write_csv(path: Path, fields: Sequence[str], rows: Sequence[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def count_table(counter: Counter) -> str:
    lines = ["| Item | Count |", "|---|---:|"]
    for key in sorted(counter):
        lines.append(f"| {key or '(blank)'} | {counter[key]} |")
    return "\n".join(lines)


def write_report(
    path: Path,
    excel_path: Path,
    profile: Dict[str, Any],
    raw_rows: Sequence[Dict[str, Any]],
    candidates: Sequence[Dict[str, Any]],
    pricing: Sequence[Dict[str, Any]],
    issues: Sequence[Dict[str, Any]],
) -> None:
    code_set = {row["source_code"] for row in candidates}
    missing_required = [code for code in REQUIRED_CODES if code not in code_set]
    missing_supplements = [code for code in REQUIRED_SUPPLEMENTAL_CODES if code not in code_set]
    supplemental_codes = sorted([code for code in code_set if is_supplemental_code(code)], key=natural_code_key)
    invalid_codes = [code for code in code_set if not is_source_code_format(code)]
    duplicate_codes = [code for code, count in Counter(row["source_code"] for row in candidates).items() if count > 1]
    non_pending_candidates = [row["source_code"] for row in candidates if row["review_status"] != REVIEW_STATUS]
    non_pending_pricing = [row["source_code"] for row in pricing if row["review_status"] != REVIEW_STATUS]
    missing_raw_name = [row["source_code"] for row in candidates if not text(row["raw_name"])]
    missing_unit = [row["source_code"] for row in candidates if not text(row["unit"])]
    issue_counts = Counter(row["issue_type"] for row in issues)
    section_counts = Counter(row["section_code"] for row in candidates)

    go = (
        len(candidates) == 143
        and not missing_required
        and not invalid_codes
        and not duplicate_codes
        and not missing_raw_name
        and not missing_unit
        and not non_pending_candidates
        and not non_pending_pricing
    )

    lines = [
        "# Stage GD2018-A111-FULL Report",
        "",
        "## 1. Task Scope",
        "",
        "Extract A.1.1 土石方工程 quota reference candidates from the third-party Guangdong 2018 Excel workbook. This run does not write a database, migration, existing pipeline, cost_items, knowledge_review_records, approved data, internal_price_library, quota_to_bill_mapping, bill_code, or quota-to-bill mapping.",
        "",
        "## 2. Input Excel Profile",
        "",
        f"- excel_file: `{excel_path}`",
        f"- file_readable: {str(profile['file_exists']).lower()}",
        f"- file_size_bytes: {profile['file_size_bytes']}",
        f"- sha256: `{profile['sha256']}`",
        f"- sheet_count: {profile['sheet_count']}",
        f"- sheet_names: {'; '.join(profile['sheet_names'])}",
        f"- target_sheet: {profile['target_sheet']}",
        f"- target_sheet_rows: {profile['target_sheet_rows']}",
        f"- target_sheet_columns: {profile['target_sheet_columns']}",
        f"- headers: {'; '.join(profile['headers'])}",
        f"- hidden_sheet_count: {profile['hidden_sheet_count']}",
        f"- formula_cell_count: {profile['formula_cell_count']}",
        f"- merged_range_count: {profile['merged_range_count']}",
        f"- has_mojibake: {str(profile['has_mojibake']).lower()}",
        f"- contains_quota_code_field: {str('定额编号' in profile['headers']).lower()}",
        f"- contains_project_name_field: {str('项目名称' in profile['headers']).lower()}",
        f"- contains_unit_field: {str('计量单位' in profile['headers']).lower()}",
        f"- pricing_columns: {'; '.join(profile['pricing_columns']) if profile['pricing_columns'] else 'none'}",
        "",
        "## 3. A.1.1 Extraction Rule",
        "",
        "- Filter only rows where `定额编号` matches `A1-1-*` or `A1-1-*-*` and the base number is 1 through 137.",
        "- Do not extract A.1.2 or other chapters.",
        "- Preserve Excel project name and unit as raw source fields.",
        "- Treat the Excel as third-party reference input; candidates remain `pending`.",
        "",
        "## 4. Candidate Row Summary",
        "",
        f"- raw_reference_rows: {len(raw_rows)}",
        f"- standard_cost_item_reference_candidates: {len(candidates)}",
        f"- expected_candidate_count: approximately 143",
        count_table(section_counts),
        "",
        "## 5. Required Code Coverage",
        "",
    ]
    for code in REQUIRED_CODES:
        lines.append(f"- {code}: {'present' if code in code_set else 'missing'}")
    lines.extend(
        [
            f"- missing_required_codes: {'; '.join(missing_required) if missing_required else 'none'}",
            "",
            "## 6. Supplemental Code Summary",
            "",
            f"- supplemental_source_code_count: {len(supplemental_codes)}",
            f"- supplemental_source_codes: {'; '.join(supplemental_codes) if supplemental_codes else 'none'}",
            f"- missing_expected_supplemental_codes: {'; '.join(missing_supplements) if missing_supplements else 'none'}",
            "",
            "## 7. Field Completeness Check",
            "",
            f"- invalid_source_code_count: {len(invalid_codes)}",
            f"- duplicated_source_code_count: {len(duplicate_codes)}",
            f"- missing_raw_name_count: {len(missing_raw_name)}",
            f"- missing_unit_count: {len(missing_unit)}",
            f"- non_pending_candidate_review_status_count: {len(non_pending_candidates)}",
            "",
            "## 8. Pricing Snapshot Summary",
            "",
            f"- pricing_snapshot_rows: {len(pricing)}",
            f"- pricing_source: {PRICING_SOURCE}",
            f"- pricing_version: {PRICING_VERSION}",
            f"- non_pending_pricing_review_status_count: {len(non_pending_pricing)}",
            "- Pricing fields are reference snapshots only and must not be loaded into `internal_price_library` from this stage.",
            "",
            "## 9. Issues and Risks",
            "",
            f"- issue_count: {len(issues)}",
            count_table(issue_counts) if issues else "No extraction issues generated.",
            "- Third-party Excel is not official truth; all candidates require manual QA.",
            "- official_source_page remains pending and should be verified against the official PDF before higher trust use.",
            "",
            "## 10. Manual QA Checklist",
            "",
            "- Verify `source_code` is truly an A1-1-* quota code.",
            "- Verify `raw_name` comes from the Excel project name.",
            "- Verify `unit` is complete.",
            "- Verify price fields are used only as reference snapshots.",
            "- Verify all `review_status` values are `pending`.",
            "- Verify no `bill_code` was generated.",
            "- Verify no mapping output was generated.",
            "",
            "## 11. Go / No-Go Recommendation for MAP-A111-0",
            "",
            "Go for MAP-A111-0 pre-mapping design after manual QA of supplemental codes and official PDF page evidence." if go else "No-Go for MAP-A111-0 until blocking extraction issues are resolved.",
            "",
            f"Generated at: {datetime.now().isoformat(timespec='seconds')}",
        ]
    )
    path.write_text("\n".join(lines), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Extract GD2018 A.1.1 full reference candidates from Excel.")
    parser.add_argument("--project-root", type=Path, default=PROJECT_ROOT_DEFAULT)
    parser.add_argument("--excel", type=Path, default=None)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    project_root = args.project_root
    excel_path = args.excel if args.excel else project_root / EXCEL_REL
    output_dir = project_root / OUTPUT_DIR_REL
    output_dir.mkdir(parents=True, exist_ok=True)

    if not excel_path.exists():
        raise SystemExit(f"Excel file not found: {excel_path}")

    workbook = load_workbook(excel_path, read_only=False, data_only=False)
    target_sheet = workbook.worksheets[0]
    headers = [norm_header(cell.value) for cell in target_sheet[1]]
    missing_headers = [header for header in EXPECTED_HEADERS if header not in headers]
    if missing_headers:
        raise SystemExit("Missing expected headers: " + "; ".join(missing_headers))

    a111_rows, invalid_a11_rows = collect_a111_rows(target_sheet, headers)
    excel_hash = sha256_file(excel_path)
    profile = collect_workbook_profile(excel_path, workbook, target_sheet, headers, a111_rows)
    raw_rows = build_raw_rows(excel_path, excel_hash, target_sheet.title, a111_rows)
    candidates = build_candidates(excel_path, excel_hash, target_sheet.title, a111_rows)
    pricing = build_pricing_snapshot(a111_rows, candidates)
    issues = build_issues(headers, a111_rows, invalid_a11_rows, candidates)

    write_csv(output_dir / "raw_reference_excel_rows_A111.csv", RAW_FIELDS, raw_rows)
    write_csv(output_dir / "standard_cost_item_reference_A111_candidate.csv", CANDIDATE_FIELDS, candidates)
    write_csv(output_dir / "reference_quota_pricing_snapshot_A111.csv", PRICING_FIELDS, pricing)
    write_csv(output_dir / "gd2018_a111_extraction_issues.csv", ISSUE_FIELDS, issues)
    write_report(output_dir / "stage_gd2018_a111_full_report.md", excel_path, profile, raw_rows, candidates, pricing, issues)

    print(f"raw_reference_rows={len(raw_rows)}")
    print(f"candidate_rows={len(candidates)}")
    print(f"pricing_snapshot_rows={len(pricing)}")
    print(f"issue_rows={len(issues)}")
    print("section_counts=" + json.dumps(dict(Counter(row["section_code"] for row in candidates)), ensure_ascii=False, sort_keys=True))
    print("issue_counts=" + json.dumps(dict(Counter(row["issue_type"] for row in issues)), ensure_ascii=False, sort_keys=True))
    print(f"output_dir={output_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
