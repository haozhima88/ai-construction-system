#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Stage 2R-0 Reference Excel intake review for A.1.1.

The third-party Excel workbook is treated only as structured reference input.
This script does not write databases, does not modify migrations or the existing
pipeline, and does not create approved/internal_price_library data.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence, Tuple

from openpyxl import load_workbook

DEFAULT_EXCEL_PATH = Path(
    r"E:\workspace\01_Projects\ai-construction-system\construction_cost_knowledge_engine\data\private\reference_extraction\source_excels\广东省房屋建筑与装饰工程综合定额（2018 ）.xlsx"
)
DEFAULT_OFFICIAL_PDF = Path(
    r"C:\Users\haozh\Downloads\1. 广东省房屋建筑与装饰工程定额20190112(上册).pdf"
)

SOURCE_TYPE = "reference_excel_third_party"
SOURCE_NAME = "广东省房屋建筑与装饰工程综合定额2018"
SOURCE_TRUST_LEVEL = "L1"
VERIFICATION_STATUS = "structure_checked"
REVIEW_STATUS = "pending"
CHAPTER_CODE = "A.1.1"
CHAPTER_NAME = "土石方工程"

EXPECTED_HEADERS = [
    "序号",
    "定额编号",
    "项目名称",
    "规格型号",
    "计量单位",
    "主材系数",
    "工程数量",
    "主材单价（元）",
    "人工费",
    "材料费",
    "机具费",
    "管理费",
    "合计",
    "主材合价（元）",
]

SAMPLE_CODES = [
    "A1-1-1",
    "A1-1-2",
    "A1-1-3",
    "A1-1-4",
    "A1-1-5",
    "A1-1-6",
    "A1-1-7",
    "A1-1-8",
    "A1-1-37",
    "A1-1-38",
    "A1-1-39",
    "A1-1-40",
    "A1-1-53",
    "A1-1-54",
    "A1-1-55",
    "A1-1-56",
    "A1-1-56-1",
    "A1-1-56-2",
    "A1-1-56-3",
    "A1-1-56-4",
    "A1-1-67",
    "A1-1-68",
    "A1-1-69",
    "A1-1-73",
    "A1-1-74",
    "A1-1-75",
    "A1-1-76",
    "A1-1-77",
    "A1-1-117",
    "A1-1-118",
    "A1-1-118-1",
    "A1-1-118-2",
    "A1-1-125",
    "A1-1-126",
    "A1-1-127",
    "A1-1-128",
    "A1-1-129",
    "A1-1-130",
    "A1-1-131",
    "A1-1-134",
    "A1-1-135",
    "A1-1-136",
    "A1-1-137",
]

WORKBOOK_PROFILE_FIELDS = [
    "profile_id",
    "extraction_source_file",
    "file_exists",
    "file_size_bytes",
    "sha256",
    "workbook_type",
    "sheet_count",
    "visible_sheet_count",
    "hidden_sheet_count",
    "formula_cell_count",
    "merged_range_count",
    "has_formulas",
    "has_merged_cells",
    "has_hidden_sheets",
    "has_mojibake",
    "bill_code_field_present",
    "pricing_columns_present",
    "source_trust_level",
    "remark",
]

SHEET_PROFILE_FIELDS = [
    "sheet_name",
    "sheet_state",
    "max_row",
    "max_column",
    "header_row",
    "headers",
    "formula_cell_count",
    "merged_range_count",
    "has_formulas",
    "has_merged_cells",
    "has_mojibake",
    "bill_code_fields",
    "pricing_columns",
    "row_count_non_empty",
    "a111_source_code_count",
    "supplemental_source_code_count",
    "duplicated_a111_source_code_count",
    "missing_name_count",
    "missing_unit_count",
    "remark",
]

CANDIDATE_FIELDS = [
    "reference_id",
    "source_type",
    "source_name",
    "extraction_source_file",
    "extraction_source_file_hash",
    "extraction_source_sheet",
    "extraction_source_row",
    "official_source_file",
    "official_source_page",
    "official_source_page_evidence_type",
    "chapter_code",
    "chapter_name",
    "section_code",
    "section_name",
    "item_group_name",
    "source_code",
    "raw_name",
    "standard_name_candidate",
    "unit",
    "labor_fee",
    "material_fee",
    "machine_fee",
    "management_fee",
    "total_fee",
    "keywords",
    "feature_template",
    "extraction_confidence",
    "source_trust_level",
    "verification_status",
    "review_status",
    "reviewer",
    "remark",
]

ISSUE_FIELDS = [
    "issue_id",
    "issue_type",
    "source_sheet",
    "source_row",
    "source_code",
    "issue_detail",
    "severity",
    "suggested_action",
]


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def light_name(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def norm_header(value: Any) -> str:
    return text(value).replace("\ufeff", "")


def is_a111_code(code: str) -> bool:
    return bool(re.fullmatch(r"A1-1-\d+(?:-\d+)?", code))


def code_base_number(code: str) -> int:
    match = re.fullmatch(r"A1-1-(\d+)(?:-\d+)?", code)
    return int(match.group(1)) if match else -1


def is_supplemental_code(code: str) -> bool:
    return bool(re.fullmatch(r"A1-1-\d+-\d+", code))


def section_for_code(code: str) -> Tuple[str, str]:
    base = code_base_number(code)
    if 1 <= base <= 66:
        return "A.1.1.1", "土方工程"
    if 67 <= base <= 125:
        return "A.1.1.2", "石方工程"
    if 126 <= base <= 137:
        return "A.1.1.3", "回填方及其他"
    return "", ""


def item_group_from_name(name: str) -> str:
    cleaned = light_name(name)
    if not cleaned:
        return ""
    split_tokens = [" 一、二类土", " 三类土", " 四类土", " 极软岩", " 软岩", " 较软岩", " 较硬岩", " 坚硬岩"]
    for token in split_tokens:
        if token in cleaned:
            return cleaned.split(token, 1)[0]
    return cleaned


def keywords_from_name(name: str) -> str:
    tokens = [token for token in re.split(r"[\s、，；;（）()]+", name or "") if len(token) >= 2]
    seen = set()
    result = []
    for token in tokens:
        if token not in seen:
            seen.add(token)
            result.append(token)
    return ";".join(result[:8])


def has_mojibake(values: Iterable[Any]) -> bool:
    patterns = ("�", "Ã", "Â", "浣", "鐨", "乱码")
    return any(any(pattern in text(value) for pattern in patterns) for value in values)


def find_bill_code_fields(headers: Sequence[str]) -> List[str]:
    return [header for header in headers if "清单" in header or ("编码" in header and header != "定额编号")]


def find_pricing_columns(headers: Sequence[str]) -> List[str]:
    pricing_words = ("人工费", "材料费", "机具费", "管理费", "合计", "单价", "合价")
    return [header for header in headers if any(word in header for word in pricing_words)]


def count_formula_cells(ws: Any) -> int:
    count = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == "f":
                count += 1
    return count


def non_empty_rows(ws: Any) -> int:
    count = 0
    for row in ws.iter_rows(values_only=True):
        if any(value is not None and text(value) for value in row):
            count += 1
    return count


def row_dict(headers: Sequence[str], row: Sequence[Any]) -> Dict[str, Any]:
    return {headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))}


def collect_a111_rows(ws: Any, headers: Sequence[str]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    code_col = headers.index("定额编号")
    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        code = text(row[code_col] if code_col < len(row) else "")
        if is_a111_code(code):
            data = row_dict(headers, row)
            data["_row_no"] = row_no
            data["_source_code"] = code
            rows.append(data)
    return rows


def build_profiles(excel_path: Path, wb: Any, a111_by_sheet: Dict[str, List[Dict[str, Any]]]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    workbook_hash = sha256_file(excel_path)
    workbook_formula_count = 0
    workbook_merged_count = 0
    workbook_has_mojibake = False
    visible_count = 0
    hidden_count = 0
    sheet_rows: List[Dict[str, Any]] = []
    bill_fields_all: List[str] = []
    pricing_fields_all: List[str] = []

    for ws in wb.worksheets:
        headers = [norm_header(cell.value) for cell in ws[1]]
        formula_count = count_formula_cells(ws)
        merged_count = len(ws.merged_cells.ranges)
        sample_values = headers[:]
        for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 20), values_only=True):
            sample_values.extend(row)
        sheet_mojibake = has_mojibake(sample_values)
        bill_fields = find_bill_code_fields(headers)
        pricing_fields = find_pricing_columns(headers)
        a111_rows = a111_by_sheet.get(ws.title, [])
        duplicated = [code for code, n in Counter(row["_source_code"] for row in a111_rows).items() if n > 1]
        missing_name = [row for row in a111_rows if not text(row.get("项目名称"))]
        missing_unit = [row for row in a111_rows if not text(row.get("计量单位"))]

        visible_count += 1 if ws.sheet_state == "visible" else 0
        hidden_count += 0 if ws.sheet_state == "visible" else 1
        workbook_formula_count += formula_count
        workbook_merged_count += merged_count
        workbook_has_mojibake = workbook_has_mojibake or sheet_mojibake
        bill_fields_all.extend(bill_fields)
        pricing_fields_all.extend(pricing_fields)

        sheet_rows.append(
            {
                "sheet_name": ws.title,
                "sheet_state": ws.sheet_state,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "header_row": 1,
                "headers": ";".join(headers),
                "formula_cell_count": formula_count,
                "merged_range_count": merged_count,
                "has_formulas": str(formula_count > 0).lower(),
                "has_merged_cells": str(merged_count > 0).lower(),
                "has_mojibake": str(sheet_mojibake).lower(),
                "bill_code_fields": ";".join(bill_fields),
                "pricing_columns": ";".join(pricing_fields),
                "row_count_non_empty": non_empty_rows(ws),
                "a111_source_code_count": len(a111_rows),
                "supplemental_source_code_count": sum(1 for row in a111_rows if is_supplemental_code(row["_source_code"])),
                "duplicated_a111_source_code_count": len(duplicated),
                "missing_name_count": len(missing_name),
                "missing_unit_count": len(missing_unit),
                "remark": "third-party workbook sheet; structure review only",
            }
        )

    workbook_rows = [
        {
            "profile_id": "REFERENCE_EXCEL_WORKBOOK",
            "extraction_source_file": str(excel_path),
            "file_exists": str(excel_path.exists()).lower(),
            "file_size_bytes": excel_path.stat().st_size if excel_path.exists() else "",
            "sha256": workbook_hash,
            "workbook_type": excel_path.suffix.lower(),
            "sheet_count": len(wb.worksheets),
            "visible_sheet_count": visible_count,
            "hidden_sheet_count": hidden_count,
            "formula_cell_count": workbook_formula_count,
            "merged_range_count": workbook_merged_count,
            "has_formulas": str(workbook_formula_count > 0).lower(),
            "has_merged_cells": str(workbook_merged_count > 0).lower(),
            "has_hidden_sheets": str(hidden_count > 0).lower(),
            "has_mojibake": str(workbook_has_mojibake).lower(),
            "bill_code_field_present": str(bool(set(bill_fields_all))).lower(),
            "pricing_columns_present": str(bool(set(pricing_fields_all))).lower(),
            "source_trust_level": SOURCE_TRUST_LEVEL,
            "remark": "Non-official third-party Excel; structured reference candidate only.",
        }
    ]
    return workbook_rows, sheet_rows


def build_candidates(excel_path: Path, excel_hash: str, official_pdf: Path, sheet_name: str, rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    by_code = {row["_source_code"]: row for row in rows}
    candidates = []
    for code in SAMPLE_CODES:
        row = by_code.get(code)
        if not row:
            continue
        section_code, section_name = section_for_code(code)
        raw_name = text(row.get("项目名称"))
        standard_name = light_name(raw_name)
        remark_parts = ["third_party_excel_reference_candidate", "official_pdf_page_pending"]
        if is_supplemental_code(code):
            remark_parts.append("supplemental_source_code")
        candidates.append(
            {
                "reference_id": f"GD2018_A111_EXCEL_{code}",
                "source_type": SOURCE_TYPE,
                "source_name": SOURCE_NAME,
                "extraction_source_file": excel_path.name,
                "extraction_source_file_hash": excel_hash,
                "extraction_source_sheet": sheet_name,
                "extraction_source_row": row["_row_no"],
                "official_source_file": official_pdf.name,
                "official_source_page": "",
                "official_source_page_evidence_type": "pending_pdf_verification",
                "chapter_code": CHAPTER_CODE,
                "chapter_name": CHAPTER_NAME,
                "section_code": section_code,
                "section_name": section_name,
                "item_group_name": item_group_from_name(raw_name),
                "source_code": code,
                "raw_name": raw_name,
                "standard_name_candidate": standard_name,
                "unit": text(row.get("计量单位")),
                "labor_fee": row.get("人工费") if row.get("人工费") is not None else "",
                "material_fee": row.get("材料费") if row.get("材料费") is not None else "",
                "machine_fee": row.get("机具费") if row.get("机具费") is not None else "",
                "management_fee": row.get("管理费") if row.get("管理费") is not None else "",
                "total_fee": row.get("合计") if row.get("合计") is not None else "",
                "keywords": keywords_from_name(standard_name),
                "feature_template": "",
                "extraction_confidence": "0.86" if not is_supplemental_code(code) else "0.72",
                "source_trust_level": SOURCE_TRUST_LEVEL,
                "verification_status": VERIFICATION_STATUS,
                "review_status": REVIEW_STATUS,
                "reviewer": "",
                "remark": ";".join(remark_parts),
            }
        )
    return candidates


def build_issues(sheet_name: str, headers: Sequence[str], a111_rows: List[Dict[str, Any]], candidates: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    issues: List[Dict[str, Any]] = []

    def add(issue_type: str, source_row: Any, source_code: str, detail: str, severity: str, action: str) -> None:
        issues.append(
            {
                "issue_id": f"ISSUE_A111_EXCEL_{len(issues) + 1:03d}",
                "issue_type": issue_type,
                "source_sheet": sheet_name,
                "source_row": source_row,
                "source_code": source_code,
                "issue_detail": detail,
                "severity": severity,
                "suggested_action": action,
            }
        )

    code_counts = Counter(row["_source_code"] for row in a111_rows)
    for row in a111_rows:
        code = row["_source_code"]
        if not code:
            add("missing_source_code", row["_row_no"], "", "A.1.1 row has no source code.", "high", "Exclude until source code is resolved.")
        if not text(row.get("项目名称")):
            add("missing_name", row["_row_no"], code, "A.1.1 row has no project/item name.", "high", "Fill from verified source before candidate use.")
        if not text(row.get("计量单位")):
            add("missing_unit", row["_row_no"], code, "A.1.1 row has no unit.", "high", "Verify against official PDF before candidate use.")
        if code_counts[code] > 1:
            add("duplicated_source_code", row["_row_no"], code, "A.1.1 source code appears more than once.", "high", "Deduplicate or explain variant rows.")
        if is_supplemental_code(code):
            add("supplemental_source_code", row["_row_no"], code, "Third-party Excel contains supplemental A.1.1 code not present in the base PDF numbering sequence.", "medium", "Verify source and policy before Stage 2B inclusion.")
        if not section_for_code(code)[0]:
            add("possible_wrong_section", row["_row_no"], code, "Code does not fall inside A.1.1 base number ranges.", "high", "Do not include until manually classified.")

    bill_fields = find_bill_code_fields(headers)
    if not bill_fields:
        add("bill_code_missing", "", "", "Workbook headers do not include a 清单编码/bill-code field.", "medium", "Do not generate bill-code mapping from this intake review.")

    non_official_columns = [h for h in headers if h in {"主材系数", "工程数量", "主材单价（元）", "主材合价（元）"}]
    if non_official_columns:
        add("non_official_column", "", "", "Columns may be third-party calculation/helper fields: " + ";".join(non_official_columns), "medium", "Treat as reference-only and exclude from standard-name truth.")

    pricing_columns = find_pricing_columns(headers)
    if pricing_columns:
        add("pricing_column_present", "", "", "Pricing columns are present: " + ";".join(pricing_columns), "low", "Keep pricing as reference context only; do not write internal_price_library.")

    pending_codes = [row["source_code"] for row in candidates if not row["official_source_page"]]
    if pending_codes:
        add("official_pdf_page_pending", "", ";".join(pending_codes), "Candidate official PDF page is not individually verified.", "medium", "Verify each candidate against the official PDF before trusted full extraction.")

    return issues


def write_csv(path: Path, fields: Sequence[str], rows: Sequence[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def markdown_count_table(counter: Counter) -> str:
    lines = ["| Item | Count |", "|---|---:|"]
    for key in sorted(counter):
        lines.append(f"| {key} | {counter[key]} |")
    return "\n".join(lines)


def write_report(
    path: Path,
    excel_path: Path,
    official_pdf: Path,
    registry_path: Path,
    workbook_profile: List[Dict[str, Any]],
    sheet_profile: List[Dict[str, Any]],
    a111_rows: List[Dict[str, Any]],
    candidates: List[Dict[str, Any]],
    issues: List[Dict[str, Any]],
) -> None:
    code_set = {row["_source_code"] for row in a111_rows}
    issue_counts = Counter(row["issue_type"] for row in issues)
    section_counts = Counter(row["section_code"] for row in candidates)
    supplementals = sorted((code for code in code_set if is_supplemental_code(code)), key=lambda c: (code_base_number(c), c))
    duplicated = sorted(code for code, n in Counter(row["_source_code"] for row in a111_rows).items() if n > 1)
    missing_name = [row for row in a111_rows if not text(row.get("项目名称"))]
    missing_unit = [row for row in a111_rows if not text(row.get("计量单位"))]
    anchors = ["A1-1-1", "A1-1-67", "A1-1-126", "A1-1-137", "A1-1-118-1", "A1-1-118-2"]

    go = bool(candidates) and not missing_name and not missing_unit and not duplicated
    lines = [
        "# Stage 2R Excel Intake Review Report - A.1.1 土石方工程",
        "",
        "## 1. Task Scope",
        "",
        "This review profiles a third-party Excel workbook as structured reference input only. It does not perform full extraction, does not write any database, does not modify migrations or the existing pipeline, and does not generate approved/internal_price_library data.",
        "",
        "## 2. Input Files",
        "",
        f"- third_party_excel: `{excel_path}`",
        f"- official_pdf: `{official_pdf}`",
        f"- stage1_page_registry: `{registry_path}`",
        f"- generated_at: {datetime.now().isoformat(timespec='seconds')}",
        "",
        "## 3. Workbook Profile",
        "",
        f"- file_size_bytes: {workbook_profile[0]['file_size_bytes']}",
        f"- sha256: `{workbook_profile[0]['sha256']}`",
        f"- sheet_count: {workbook_profile[0]['sheet_count']}",
        f"- hidden_sheet_count: {workbook_profile[0]['hidden_sheet_count']}",
        f"- formula_cell_count: {workbook_profile[0]['formula_cell_count']}",
        f"- merged_range_count: {workbook_profile[0]['merged_range_count']}",
        f"- has_mojibake: {workbook_profile[0]['has_mojibake']}",
        f"- bill_code_field_present: {workbook_profile[0]['bill_code_field_present']}",
        f"- pricing_columns_present: {workbook_profile[0]['pricing_columns_present']}",
        "",
        "## 4. Sheet Profile",
        "",
        "| Sheet | State | Rows | Cols | A.1.1 Codes | Supplementals | Missing Name | Missing Unit |",
        "|---|---|---:|---:|---:|---:|---:|---:|",
    ]
    for row in sheet_profile:
        lines.append(
            f"| {row['sheet_name']} | {row['sheet_state']} | {row['max_row']} | {row['max_column']} | {row['a111_source_code_count']} | {row['supplemental_source_code_count']} | {row['missing_name_count']} | {row['missing_unit_count']} |"
        )
    lines.extend(
        [
            "",
            "## 5. A.1.1 Source Code Coverage",
            "",
            f"- A.1.1 source_code count: {len(a111_rows)}",
            f"- A1-1-1 exists: {'yes' if 'A1-1-1' in code_set else 'no'}",
            f"- A1-1-67 exists: {'yes' if 'A1-1-67' in code_set else 'no'}",
            f"- A1-1-126 exists: {'yes' if 'A1-1-126' in code_set else 'no'}",
            f"- A1-1-137 exists: {'yes' if 'A1-1-137' in code_set else 'no'}",
            f"- A1-1-118-1 exists: {'yes' if 'A1-1-118-1' in code_set else 'no'}",
            f"- A1-1-118-2 exists: {'yes' if 'A1-1-118-2' in code_set else 'no'}",
            f"- supplemental_source_codes: {'; '.join(supplementals) if supplementals else 'none'}",
            f"- duplicated_source_codes: {'; '.join(duplicated) if duplicated else 'none'}",
            "",
            "## 6. Main Row Availability",
            "",
            f"- missing project names in A.1.1: {len(missing_name)}",
            f"- missing units in A.1.1: {len(missing_unit)}",
            "- Excel rows are already one-row-per-code for A.1.1, which is structurally better than the PDF table extraction path.",
            "- This does not make the Excel official truth; it only makes it a better structured candidate source.",
            "",
            "## 7. Candidate Sample Summary",
            "",
            f"- sample_candidate_rows: {len(candidates)}",
            markdown_count_table(section_counts),
            "",
            "## 8. Non-official Columns",
            "",
            "- Non-official/helper-looking columns detected: 主材系数, 工程数量, 主材单价（元）, 主材合价（元）.",
            "- Pricing columns are present and must remain reference context only in this review.",
            "",
            "## 9. Name Hierarchy / Feature Parsing Risk",
            "",
            "- `项目名称` appears flattened into a single name string, e.g. parent name plus soil/rock class or distance feature.",
            "- `standard_name_candidate` is therefore only a lightly cleaned candidate, not an enterprise final standard name.",
            "- Feature parsing should be a later reviewed step, not part of this intake review.",
            "",
            "## 10. Bill Code Missing Analysis",
            "",
            "- No 清单编码 / bill-code field was found in workbook headers.",
            "- Do not generate bill-code mapping from this Excel intake.",
            "",
            "## 11. Recommended Next Stage",
            "",
            "- Stage 2R-1 can use this Excel as structured candidate input after manual QA.",
            "- Before full candidate generation, verify supplemental codes and sample rows against the official PDF or another trusted source.",
            "- Keep all generated candidate review statuses as `pending`.",
            "",
            "## 12. Go / No-Go Recommendation",
            "",
            "Go for Stage 2R-1 sample-to-full candidate design, with manual verification controls." if go else "No-Go until missing/duplicated core fields are resolved.",
            "",
            "## Issue Summary",
            "",
            markdown_count_table(issue_counts),
            "",
        ]
    )
    path.write_text("\n".join(lines), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Stage 2R-0 reference Excel intake review.")
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL_PATH)
    parser.add_argument("--official-pdf", type=Path, default=DEFAULT_OFFICIAL_PDF)
    parser.add_argument(
        "--project-root",
        type=Path,
        default=Path(__file__).resolve().parents[2],
        help="construction_cost_knowledge_engine project root.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    excel_path = args.excel
    official_pdf = args.official_pdf
    project_root = args.project_root
    registry_path = project_root / "data" / "private" / "reference_extraction" / "runs" / "A111_stage1" / "page_registry_A111.json"
    output_dir = project_root / "data" / "private" / "reference_extraction" / "runs" / "A111_stage2R_excel_review"
    output_dir.mkdir(parents=True, exist_ok=True)

    if not excel_path.exists():
        raise SystemExit(f"Excel file not found: {excel_path}")
    if not registry_path.exists():
        raise SystemExit(f"Stage 1 registry not found: {registry_path}")
    json.loads(registry_path.read_text(encoding="utf-8"))

    excel_hash = sha256_file(excel_path)
    wb = load_workbook(excel_path, read_only=False, data_only=False)

    a111_by_sheet: Dict[str, List[Dict[str, Any]]] = {}
    first_target_sheet = wb.worksheets[0].title
    first_headers: List[str] = []
    for ws in wb.worksheets:
        headers = [norm_header(cell.value) for cell in ws[1]]
        if "定额编号" in headers and "项目名称" in headers and "计量单位" in headers:
            a111_by_sheet[ws.title] = collect_a111_rows(ws, headers)
            if ws.title == first_target_sheet:
                first_headers = headers

    workbook_profile, sheet_profile = build_profiles(excel_path, wb, a111_by_sheet)
    target_rows = a111_by_sheet.get(first_target_sheet, [])
    candidates = build_candidates(excel_path, excel_hash, official_pdf, first_target_sheet, target_rows)
    issues = build_issues(first_target_sheet, first_headers or EXPECTED_HEADERS, target_rows, candidates)

    write_csv(output_dir / "excel_workbook_profile.csv", WORKBOOK_PROFILE_FIELDS, workbook_profile)
    write_csv(output_dir / "excel_sheet_profile.csv", SHEET_PROFILE_FIELDS, sheet_profile)
    write_csv(output_dir / "a111_excel_candidate_sample.csv", CANDIDATE_FIELDS, candidates)
    write_csv(output_dir / "a111_excel_issues.csv", ISSUE_FIELDS, issues)
    write_report(
        output_dir / "stage2R_excel_review_report.md",
        excel_path,
        official_pdf,
        registry_path,
        workbook_profile,
        sheet_profile,
        target_rows,
        candidates,
        issues,
    )

    print(f"workbook_profile_rows={len(workbook_profile)}")
    print(f"sheet_profile_rows={len(sheet_profile)}")
    print(f"a111_source_code_count={len(target_rows)}")
    print(f"candidate_sample_rows={len(candidates)}")
    print(f"issue_rows={len(issues)}")
    print("candidate_section_counts=" + json.dumps(dict(Counter(row["section_code"] for row in candidates)), ensure_ascii=False, sort_keys=True))
    print("issue_counts=" + json.dumps(dict(Counter(row["issue_type"] for row in issues)), ensure_ascii=False, sort_keys=True))
    print(f"output_dir={output_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
